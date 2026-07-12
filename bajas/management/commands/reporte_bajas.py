# bajas/management/commands/reporte_bajas.py
#
# 1) Detecta los morosos de una compañía con mora dentro de la ventana
#    [días de gracia] .. [días de gracia + 2] (no agarra el backlog viejo).
# 2) Arma un "Reporte de bajas" (lista en el cuerpo + PDF adjunto) y lo envía
#    al email configurado para esa compañía en el panel de Correos de compañías.
# 3) Si el email se envió OK, da de baja esas pólizas automáticamente:
#       estado = cancelada · motivo = incumplimiento de pago
#       observación = "Dada de baja automática por falta de pago" (se ve en el perfil)
#       fecha de baja = hoy · queda en el historial de la póliza.
# 4) Manda un aviso al dueño (EMAIL_AVISO) con el resumen de lo que se dio de baja.
#
# Columnas del reporte (email y PDF): Nombre y apellido · Patente · Marca · Modelo.
#
# USO:
#   python manage.py reporte_bajas --compania NRE              (hace TODO)
#   python manage.py reporte_bajas --compania NRE --solo-email (manda, NO da de baja)
#   python manage.py reporte_bajas --dry-run                   (no manda ni da de baja)
#
# RAILWAY CRON (ej. 18 hs Argentina, días hábiles → corre DESPUÉS de los recordatorios):
#   schedule = "0 21 * * 1-5"
#   command  = "python manage.py reporte_bajas --compania NRE"

import os
import logging
from io import BytesIO

from django.conf import settings
from django.core.mail import EmailMessage
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone

from polizas.models import Poliza
from bajas.models import CorreoCompaniaBaja, BajaPoliza
from bajas.services import detectar_polizas_en_mora

logger = logging.getLogger(__name__)

COMPANIA_DEFAULT = "NRE"
BORDO = "#8B1E3F"
OBS_BAJA = "Dada de baja automática por falta de pago"

# Ventana de mora: se procesan las pólizas cuya mora va desde el día de gracia
# de la compañía hasta (día de gracia + VENTANA_EXTRA_DIAS) días. Así no se agarra
# el backlog viejo y se tolera que el cron no corra algún día (finde/feriado/falla).
VENTANA_EXTRA_DIAS = 2

# Email del dueño para el aviso de "se dieron de baja X pólizas".
# Se puede sobreescribir con la variable de entorno EMAIL_AVISO_BAJAS sin tocar el código.
_EMAILS_AVISO_DEFAULT = "francodavid_dev@outlook.com"
EMAILS_AVISO = [
    e.strip()
    for e in os.environ.get("EMAIL_AVISO_BAJAS", _EMAILS_AVISO_DEFAULT).split(",")
    if e.strip() and "@" in e
]

# Números de WhatsApp para el aviso interno. Override: env WHATSAPP_AVISO_BAJAS (coma).
_WHATSAPP_AVISO_DEFAULT = "1164235336"
WHATSAPP_NUMEROS = [
    n.strip()
    for n in os.environ.get("WHATSAPP_AVISO_BAJAS", _WHATSAPP_AVISO_DEFAULT).split(",")
    if n.strip()
]


def _oficina_con_whatsapp():
    """Id de una oficina activa con credenciales de UltraMsg cargadas (o None)."""
    try:
        from usuarios.models import Oficina
        ofi = (
            Oficina.objects.filter(activa=True)
            .exclude(ultramsg_instance_id__isnull=True).exclude(ultramsg_instance_id="")
            .exclude(ultramsg_token__isnull=True).exclude(ultramsg_token="")
            .order_by("id")
            .first()
        )
        return ofi.id if ofi else None
    except Exception:
        return None

# Texto del email a la compañía (con emojis; el PDF va sin emojis)
MENSAJE_EMAIL = ("Estimados 👋, adjuntamos el listado de pólizas a dar de baja. "
                 "El detalle va a continuación y en el PDF y el Excel adjuntos.")


def _generar_pdf(compania, polizas, fecha):
    """Arma el PDF 'Reporte de bajas' y devuelve los bytes."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=18 * mm, bottomMargin=18 * mm,
        leftMargin=16 * mm, rightMargin=16 * mm,
        title="Reporte de bajas",
    )

    styles = getSampleStyleSheet()
    h_title = ParagraphStyle(
        "TituloBajas", parent=styles["Title"],
        fontSize=18, textColor=colors.HexColor(BORDO), spaceAfter=4,
    )
    h_sub = ParagraphStyle(
        "SubBajas", parent=styles["Normal"],
        fontSize=10, textColor=colors.HexColor("#6b7280"), spaceAfter=14,
    )
    p_total = ParagraphStyle(
        "TotalBajas", parent=styles["Normal"],
        fontSize=11, textColor=colors.HexColor("#374151"), spaceAfter=10,
    )

    elementos = []
    elementos.append(Paragraph("Reporte de bajas", h_title))
    elementos.append(Paragraph(
        f"{settings.EMAIL_REMITENTE_NOMBRE} &nbsp;·&nbsp; {compania} &nbsp;·&nbsp; {fecha.strftime('%d/%m/%Y')}",
        h_sub,
    ))
    elementos.append(Paragraph(
        f"Total: <b>{len(polizas)} póliza{'s' if len(polizas) != 1 else ''}</b>",
        p_total,
    ))
    elementos.append(Spacer(1, 4))

    data = [["Nombre y apellido", "Patente", "Marca", "Modelo"]]
    for p in polizas:
        data.append([
            str(p["nombre_apellido"]),
            str(p["patente"]),
            str(p["marca"]),
            str(p["modelo"]),
        ])

    tabla = Table(data, colWidths=[62 * mm, 28 * mm, 38 * mm, 38 * mm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BORDO)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FBEFF3")]),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    return buffer.getvalue()


def _generar_excel(compania, polizas, fecha):
    """Arma el Excel 'Reporte de bajas' (mismas columnas que el PDF) y devuelve los bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Bajas"

    ws.append([f"Reporte de bajas — {compania} — {fecha.strftime('%d/%m/%Y')}"])
    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(bold=True, size=13, color="8B1E3F")

    ws.append([f"Total: {len(polizas)} póliza{'s' if len(polizas) != 1 else ''}"])
    ws.merge_cells("A2:D2")
    ws["A2"].font = Font(italic=True, size=10, color="6B7280")

    ws.append([])

    encabezados = ["Nombre y apellido", "Patente", "Marca", "Modelo"]
    ws.append(encabezados)
    fila_header = ws.max_row
    for col in range(1, len(encabezados) + 1):
        celda = ws.cell(row=fila_header, column=col)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="8B1E3F")
        celda.alignment = Alignment(horizontal="left", vertical="center")

    for p in polizas:
        ws.append([
            str(p["nombre_apellido"]),
            str(p["patente"]),
            str(p["marca"]),
            str(p["modelo"]),
        ])

    anchos = [32, 14, 20, 20]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _cuerpo_html(compania, polizas, fecha, titulo="Reporte de bajas", mensaje=MENSAJE_EMAIL):
    """Cuerpo del email: título + mensaje + tabla (Nombre/Patente/Marca/Modelo)."""
    filas = ""
    for p in polizas:
        filas += f"""
        <tr>
          <td style="padding:8px 12px;border-bottom:1px solid #e5e7eb;font-size:13px">{p['nombre_apellido']}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e5e7eb;font-family:monospace;font-size:13px;font-weight:600">{p['patente']}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e5e7eb;font-size:13px">{p['marca']}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e5e7eb;font-size:13px">{p['modelo']}</td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f9fafb;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f9fafb;padding:40px 20px">
    <tr><td>
      <table width="600" cellpadding="0" cellspacing="0" align="center"
             style="background:#ffffff;border-radius:8px;border:1px solid #e5e7eb;overflow:hidden">
        <tr>
          <td style="background:{BORDO};padding:24px 32px">
            <p style="margin:0;color:#ffffff;font-size:18px;font-weight:500">{titulo}</p>
            <p style="margin:4px 0 0;color:#f0d4dc;font-size:13px">{compania} · {fecha.strftime('%d/%m/%Y')}</p>
          </td>
        </tr>
        <tr>
          <td style="padding:32px">
            <p style="margin:0 0 20px;font-size:14px;color:#4b5563;line-height:1.6">{mensaje}</p>
            <table width="100%" cellpadding="0" cellspacing="0"
                   style="border:1px solid #e5e7eb;border-radius:6px;overflow:hidden;border-collapse:collapse">
              <thead>
                <tr style="background:#f3f4f6">
                  <th style="padding:10px 12px;text-align:left;font-size:11px;font-weight:600;color:#6b7280">Nombre y apellido</th>
                  <th style="padding:10px 12px;text-align:left;font-size:11px;font-weight:600;color:#6b7280">Patente</th>
                  <th style="padding:10px 12px;text-align:left;font-size:11px;font-weight:600;color:#6b7280">Marca</th>
                  <th style="padding:10px 12px;text-align:left;font-size:11px;font-weight:600;color:#6b7280">Modelo</th>
                </tr>
              </thead>
              <tbody>{filas}</tbody>
            </table>
          </td>
        </tr>
        <tr>
          <td style="background:#f9fafb;padding:20px 32px;border-top:1px solid #e5e7eb">
            <p style="margin:0;font-size:12px;color:#9ca3af">{settings.EMAIL_REMITENTE_NOMBRE} · Generado el {fecha.strftime('%d/%m/%Y')}</p>
          </td>
        </tr>
      </table>
    </td></tr>
  </table>
</body></html>"""


def _enriquecer(polizas):
    """Agrega nombre_apellido, marca y modelo a cada póliza (desde la base)."""
    ids = [p["id"] for p in polizas]
    extra = {
        row["id"]: row
        for row in Poliza.objects.filter(id__in=ids).values(
            "id", "marca", "modelo", "cliente__nombre", "cliente__apellido"
        )
    }
    for p in polizas:
        e = extra.get(p["id"], {})
        nombre = (e.get("cliente__nombre") or "").strip()
        apellido = (e.get("cliente__apellido") or "").strip()
        p["nombre_apellido"] = f"{nombre} {apellido}".strip() or "—"
        p["marca"] = (e.get("marca") or "").strip() or "—"
        p["modelo"] = (e.get("modelo") or "").strip() or "—"
    return polizas


def _marcar_dadas_de_baja(polizas, hoy):
    """Da de baja las pólizas reportadas (cancelada + motivo + observación + fecha)."""
    try:
        from polizas.utils.viewtools import hist_log as _hist_log
    except Exception:
        _hist_log = None

    now = timezone.now()
    marcadas = 0

    with transaction.atomic():
        for p in polizas:
            try:
                poliza = Poliza.objects.get(id=p["id"])
            except Poliza.DoesNotExist:
                continue

            if str(getattr(poliza, "estado", "")).strip().lower() == "cancelada":
                continue

            poliza.estado = "cancelada"
            update_fields = ["estado"]
            if hasattr(poliza, "fecha_baja"):
                poliza.fecha_baja = hoy
                update_fields.append("fecha_baja")
            if hasattr(poliza, "motivo_baja"):
                poliza.motivo_baja = "INCUMPLIMIENTO_PAGO"
                update_fields.append("motivo_baja")
            if hasattr(poliza, "observaciones_baja"):
                poliza.observaciones_baja = OBS_BAJA
                update_fields.append("observaciones_baja")
            poliza.save(update_fields=update_fields)

            baja, _ = BajaPoliza.objects.get_or_create(
                poliza=poliza,
                defaults={"estado": BajaPoliza.Estado.ENVIADA},
            )
            baja.estado = BajaPoliza.Estado.ENVIADA
            if hasattr(baja, "enviada_en") and not baja.enviada_en:
                baja.enviada_en = now
            baja.save()

            if _hist_log:
                try:
                    _hist_log(
                        poliza=poliza, tipo="POLIZA_BAJA_AUTOMATICA",
                        mensaje=OBS_BAJA, severidad="ACTION",
                        request=None, subject=poliza, categoria="POLIZA",
                    )
                except Exception:
                    pass

            marcadas += 1

    return marcadas


def _enviar_aviso_dueno(compania, polizas, marcadas, hoy, pdf_bytes, excel_bytes=None):
    """Aviso de confirmación al dueño con lo que se dio de baja."""
    if not EMAILS_AVISO or marcadas == 0:
        return

    plural = "s" if marcadas != 1 else ""
    asunto = f"✅ Bajas procesadas — {compania} · {hoy.strftime('%d/%m/%Y')} ({marcadas})"
    mensaje = (
        f"Se dieron de baja automáticamente <strong>{marcadas} póliza{plural}</strong> "
        f"de {compania} por falta de pago. El detalle va abajo, en el PDF y en el Excel adjuntos."
    )
    cuerpo = _cuerpo_html(
        compania, polizas, hoy,
        titulo="Bajas procesadas", mensaje=mensaje,
    )
    try:
        msg = EmailMessage(
            subject=asunto, body=cuerpo,
            from_email=settings.DEFAULT_FROM_EMAIL, to=EMAILS_AVISO,
        )
        msg.content_subtype = "html"
        msg.attach(f"bajas_{hoy.strftime('%d-%m-%Y')}.pdf", pdf_bytes, "application/pdf")
        if excel_bytes:
            msg.attach(
                f"bajas_{hoy.strftime('%d-%m-%Y')}.xlsx", excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        msg.send(fail_silently=False)
        logger.info(f"[reporte_bajas] Aviso enviado al dueño ({', '.join(EMAILS_AVISO)})")
    except Exception as e:
        logger.error(f"[reporte_bajas] Error al enviar aviso al dueño: {e}")


def _enviar_aviso_whatsapp(compania, marcadas, hoy):
    """
    Aviso corto por WhatsApp a los números de control cuando se dan bajas.
    Blindado: si algo falla (import, credenciales, red), NUNCA rompe el proceso.
    """
    if marcadas == 0:
        return
    try:
        from notificaciones.utils.mensajeria import enviar_whatsapp
    except Exception as e:
        logger.error(f"[reporte_bajas] No se pudo importar enviar_whatsapp: {e}")
        return

    plural = "s" if marcadas != 1 else ""
    msg = (
        f"⚠️ *Bajas automáticas*\n"
        f"Se dieron de baja {marcadas} póliza{plural} de {compania} "
        f"por falta de pago.\n"
        f"📅 {hoy.strftime('%d/%m/%Y')}"
    )
    ofi_wa = _oficina_con_whatsapp()
    for numero in WHATSAPP_NUMEROS:
        try:
            ok, info = enviar_whatsapp(numero, msg, oficina=ofi_wa)
            if not ok:
                logger.error(f"[reporte_bajas] WhatsApp a {numero} no se envió: {info}")
        except Exception as e:
            logger.error(f"[reporte_bajas] WhatsApp a {numero} excepción: {e}")


class Command(BaseCommand):
    help = "Envía el Reporte de bajas de una compañía y da de baja automáticamente esas pólizas."

    def add_arguments(self, parser):
        parser.add_argument(
            "--compania", default=COMPANIA_DEFAULT,
            help=f"Nombre de la compañía a procesar (default: {COMPANIA_DEFAULT}).",
        )
        parser.add_argument(
            "--dry-run", action="store_true", default=False,
            help="No manda el email ni da de baja. Solo muestra a quiénes incluiría.",
        )
        parser.add_argument(
            "--solo-email", action="store_true", default=False,
            help="Manda el email pero NO da de baja las pólizas (para probar el envío).",
        )

    def handle(self, *args, **options):
        compania = (options["compania"] or "").strip()
        dry = bool(options["dry_run"])
        hoy = timezone.localdate()

        try:
            correo_obj = CorreoCompaniaBaja.objects.get(compania__iexact=compania)
        except CorreoCompaniaBaja.DoesNotExist:
            self.stderr.write(self.style.ERROR(
                f"No hay correo configurado para '{compania}'. "
                f"Cargalo en el panel de Correos de compañías."
            ))
            return

        polizas_raw = detectar_polizas_en_mora(dias_default=correo_obj.dias_gracia)
        polizas_cia = [
            p for p in polizas_raw
            if (p["compania"] or "").strip().lower() == compania.lower()
        ]

        # Ventana de mora: de [días de gracia] a [días de gracia + VENTANA_EXTRA_DIAS].
        # El piso ya lo garantiza detectar_polizas_en_mora; acá sumamos el techo.
        mora_max = correo_obj.dias_gracia + VENTANA_EXTRA_DIAS
        polizas_cia = [p for p in polizas_cia if p["mora_dias"] <= mora_max]

        polizas_cia.sort(key=lambda p: p["mora_dias"], reverse=True)

        if not polizas_cia:
            self.stdout.write(f"[reporte_bajas] {compania}: sin pólizas en la ventana de mora para procesar.")
            return

        _enriquecer(polizas_cia)

        if dry:
            self.stdout.write(self.style.MIGRATE_HEADING(
                f"[DRY] {compania} → {correo_obj.email} · {len(polizas_cia)} pólizas "
                f"(mora {correo_obj.dias_gracia}-{mora_max} días; en real se envía Y se da de baja)"
            ))
            for p in polizas_cia:
                self.stdout.write(
                    f"  {p['nombre_apellido']} | {p['patente']} | {p['marca']} | {p['modelo']} | {p['mora_dias']}d"
                )
            return

        pdf_bytes = _generar_pdf(compania, polizas_cia, hoy)
        excel_bytes = _generar_excel(compania, polizas_cia, hoy)
        cuerpo = _cuerpo_html(compania, polizas_cia, hoy)
        asunto = f"📋 Reporte de bajas — {compania} · {hoy.strftime('%d/%m/%Y')}"

        # Parsea "a@x.com, b@y.com" a ["a@x.com", "b@y.com"] para mandar a TODOS
        # los destinatarios. Sin esto, con varios emails se mandaba a un único
        # destinatario malformado y el envío fallaba.
        destinatarios = (
            correo_obj.emails_lista()
            if hasattr(correo_obj, "emails_lista")
            else [correo_obj.email]
        ) or [correo_obj.email]

        try:
            msg = EmailMessage(
                subject=asunto,
                body=cuerpo,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=destinatarios,
            )
            msg.content_subtype = "html"
            msg.attach(
                f"reporte_bajas_{hoy.strftime('%d-%m-%Y')}.pdf",
                pdf_bytes,
                "application/pdf",
            )
            msg.attach(
                f"reporte_bajas_{hoy.strftime('%d-%m-%Y')}.xlsx",
                excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            msg.send(fail_silently=False)
        except Exception as e:
            self.stderr.write(self.style.ERROR(
                f"Error al enviar el reporte: {e}. NO se dio de baja ninguna póliza."
            ))
            logger.error(f"[reporte_bajas] Error al enviar a {correo_obj.email}: {e}")
            return

        self.stdout.write(self.style.SUCCESS(
            f"✓ Reporte enviado a {correo_obj.email} ({len(polizas_cia)} pólizas de {compania})"
        ))

        if options.get("solo_email"):
            self.stdout.write(self.style.WARNING(
                "Modo --solo-email: NO se dio de baja ninguna póliza."
            ))
            return

        marcadas = _marcar_dadas_de_baja(polizas_cia, hoy)
        self.stdout.write(self.style.SUCCESS(
            f"✓ Pólizas dadas de baja automáticamente: {marcadas}"
        ))

        _enviar_aviso_dueno(compania, polizas_cia, marcadas, hoy, pdf_bytes, excel_bytes)
        if marcadas:
            self.stdout.write(self.style.SUCCESS(f"✓ Aviso enviado a {', '.join(EMAILS_AVISO)}"))

        _enviar_aviso_whatsapp(compania, marcadas, hoy)
        if marcadas:
            self.stdout.write(self.style.SUCCESS(f"✓ WhatsApp enviado a {len(WHATSAPP_NUMEROS)} número(s)"))

        logger.info(
            f"[reporte_bajas] {compania}: enviadas={len(polizas_cia)} dadas_de_baja={marcadas}"
        )