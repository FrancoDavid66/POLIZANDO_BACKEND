# balanzes/views.py
import io
import re
import traceback
import xlsxwriter
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Optional

from django.db.models import Sum, Count, Q, Max
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.http import HttpResponse, FileResponse

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from seguros_project.pagination import LargeResultsSetPagination

from .models import Ingreso, Egreso, Categoria
from .serializers import IngresoSerializer, EgresoSerializer, CategoriaSerializer
from notificaciones.services_balanzes import enviar_balance_por_whatsapp
from polizas.models import Poliza
from usuarios.mixins import MultiTenantMixin

# 🚀 IMPORTAMOS EL MODELO OFICIAL DE OFICINAS
from usuarios.models import Oficina

def _d_to_str(v):
    if v is None: return "0"
    try: return str(v)
    except Exception: return "0"

# ==========================================
# 🚀 HELPER OPTIMIZADO (Devuelve IDs reales)
# ==========================================
def _get_todas_las_llaves_oficina(raw_or_obj):
    keys = []
    
    if isinstance(raw_or_obj, str) and raw_or_obj.strip().upper() == "ALL":
        return list(Oficina.objects.values_list('id', flat=True))

    if hasattr(raw_or_obj, 'id') and raw_or_obj.id:
        keys.append(raw_or_obj.id)
        
    if isinstance(raw_or_obj, str) and raw_or_obj.strip():
        val = raw_or_obj.strip()
        if val.isdigit():
            ofi = Oficina.objects.filter(Q(codigo=val) | Q(id=val)).first()
        else:
            ofi = Oficina.objects.filter(nombre__icontains=val).first()
        if ofi:
            keys.append(ofi.id)
            
    return list(set(k for k in keys if k))


# ==========================================
# 🚀 HELPERS DE EXPORT (compartidos por Ingresos y Egresos)
# ==========================================
def _parse_ymd(s):
    if not s:
        return None
    return parse_date(str(s).strip())


def _build_historial_xlsx_response(
    items,
    *,
    titulo,
    columnas,
    fila_extractor,
    filename,
    color_header="1E3A8A",
):
    """
    Genera Excel con TABLA REAL de openpyxl (objeto Table) — flechitas de filtro
    nativas en cada header al abrir el archivo.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    # Nombre de hoja seguro (Excel prohíbe: \ / ? * : [ ] y máx 31 chars)
    _safe_title = re.sub(r'[\\/\?\*:\[\]]', '', titulo)[:30] or "Historial"
    ws.title = _safe_title

    title_font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    title_fill = PatternFill("solid", fgColor=color_header)
    title_align = Alignment(horizontal="left", vertical="center", indent=1)

    subtitle_font = Font(color="475569", size=10, italic=True, name="Calibri")
    subtitle_align = Alignment(horizontal="left", vertical="center", indent=1)

    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    header_fill = PatternFill("solid", fgColor=color_header)
    header_align = Alignment(horizontal="center", vertical="center")

    thin = Side(border_style="thin", color="E2E8F0")
    cell_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    cell_font = Font(color="1E293B", size=10, name="Calibri")
    cell_align_left = Alignment(horizontal="left", vertical="center", indent=1)
    cell_align_right = Alignment(horizontal="right", vertical="center", indent=1)
    cell_align_center = Alignment(horizontal="center", vertical="center")

    total_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    total_fill = PatternFill("solid", fgColor="064E3B")
    total_align = Alignment(horizontal="right", vertical="center", indent=1)

    HEADERS = [label for _, label in columnas]
    CLAVES = [clave for clave, _ in columnas]
    n_cols = len(HEADERS)
    last_col_letter = get_column_letter(n_cols)

    monto_col_idx = None
    for i, clave in enumerate(CLAVES, 1):
        if clave == "monto":
            monto_col_idx = i
            break

    # Fila 1: título (merge)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = title_font
    c.fill = title_fill
    c.alignment = title_align
    ws.row_dimensions[1].height = 28

    # Fila 2: subtítulo
    total_count = len(items)
    generado_str = timezone.localtime().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(row=2, column=1, value=f"Generado: {generado_str}   ·   Total de registros: {total_count}")
    c.font = subtitle_font
    c.alignment = subtitle_align
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6

    # Fila 4: headers
    header_row = 4
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border
    ws.row_dimensions[header_row].height = 24

    row_alt_fill = PatternFill("solid", fgColor="F8FAFC")
    total_general = 0.0
    current_row = header_row

    for it in items:
        current_row += 1
        data = fila_extractor(it)

        is_alt = (current_row - header_row) % 2 == 0
        for col_idx, clave in enumerate(CLAVES, 1):
            val = data.get(clave, "") or ""

            if clave == "monto":
                try:
                    val = float(val or 0)
                    total_general += val
                except (TypeError, ValueError):
                    val = 0.0

            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = cell_font
            cell.border = cell_border
            if is_alt:
                cell.fill = row_alt_fill

            if clave == "monto":
                cell.number_format = '"$"#,##0.00'
                cell.alignment = cell_align_right
            elif clave in ("forma_pago", "fecha"):
                cell.alignment = cell_align_center
            else:
                cell.alignment = cell_align_left

    data_last_row = current_row
    has_data = data_last_row > header_row

    # Total general
    if has_data and monto_col_idx:
        total_row = data_last_row + 1
        if monto_col_idx > 1:
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=monto_col_idx - 1)
            c = ws.cell(row=total_row, column=1, value="TOTAL GENERAL")
            c.font = total_font
            c.fill = total_fill
            c.alignment = total_align
        c = ws.cell(row=total_row, column=monto_col_idx, value=total_general)
        c.font = total_font
        c.fill = total_fill
        c.alignment = total_align
        c.number_format = '"$"#,##0.00'
        ws.row_dimensions[total_row].height = 22

    # 🚀 TABLA REAL de Excel con flechitas de filtro nativas
    if has_data:
        table_ref = f"A{header_row}:{last_col_letter}{data_last_row}"
        # displayName: solo letras y números, máx 31 chars, debe empezar con letra
        safe_name = re.sub(r'[^A-Za-z0-9]', '', titulo) or "Historial"
        if not safe_name or not safe_name[0].isalpha():
            safe_name = "T" + safe_name
        table = Table(displayName=safe_name[:30], ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    # Anchos de columna
    anchos_default = {
        "fecha": 14, "descripcion": 40, "categoria": 22,
        "forma_pago": 16, "pagado_por": 26, "billetera": 22,
        "oficina": 20, "monto": 16,
    }
    for i, clave in enumerate(CLAVES, 1):
        ws.column_dimensions[get_column_letter(i)].width = anchos_default.get(clave, 18)

    ws.freeze_panes = "A5"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _build_historial_pdf_response(
    items,
    *,
    titulo,
    columnas,
    fila_extractor,
    filename,
    color_header_rgb=(30, 58, 138),
):
    """
    Genera PDF A4 horizontal con tabla, header repetido por página y total general.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    buff = io.BytesIO()
    pdf = canvas.Canvas(buff, pagesize=landscape(A4))
    width, height = landscape(A4)

    left = 12 * mm
    top = height - 12 * mm
    pdf.setTitle(filename)

    usable = width - 2 * left
    total_ratio = sum(c[2] for c in columnas)
    col_widths = [int(usable * (c[2] / total_ratio)) for c in columnas]
    xs = []
    cur = left
    for w in col_widths:
        xs.append(cur)
        cur += w

    monto_col_idx = None
    for i, (clave, _, _) in enumerate(columnas):
        if clave == "monto":
            monto_col_idx = i
            break

    def draw_title():
        pdf.setFillColorRGB(color_header_rgb[0]/255, color_header_rgb[1]/255, color_header_rgb[2]/255)
        pdf.rect(left, top - 14*mm, usable, 14*mm, fill=1, stroke=0)
        pdf.setFillColorRGB(1, 1, 1)
        pdf.setFont("Helvetica-Bold", 13)
        pdf.drawString(left + 4*mm, top - 6*mm, titulo)
        pdf.setFont("Helvetica", 9)
        pdf.drawString(left + 4*mm, top - 11*mm, f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}  ·  Total: {len(items)} registros")
        pdf.setFillColorRGB(0, 0, 0)

    def draw_header(y):
        pdf.setFillColorRGB(color_header_rgb[0]/255, color_header_rgb[1]/255, color_header_rgb[2]/255)
        pdf.rect(left, y - 5*mm, usable, 6*mm, fill=1, stroke=0)
        pdf.setFillColorRGB(1, 1, 1)
        pdf.setFont("Helvetica-Bold", 9)
        for i, (_, label, _) in enumerate(columnas):
            align = "right" if columnas[i][0] == "monto" else "left"
            if align == "right":
                pdf.drawRightString(xs[i] + col_widths[i] - 3, y - 1.5*mm, str(label))
            else:
                pdf.drawString(xs[i] + 3, y - 1.5*mm, str(label))
        pdf.setFillColorRGB(0, 0, 0)
        return y - 7*mm

    def draw_row(data, y, alt=False):
        if alt:
            pdf.setFillColorRGB(0.97, 0.98, 0.99)
            pdf.rect(left, y - 5*mm, usable, 5.5*mm, fill=1, stroke=0)
            pdf.setFillColorRGB(0, 0, 0)
        pdf.setFont("Helvetica", 8.5)
        for i, (clave, _, _) in enumerate(columnas):
            val = data.get(clave, "") or ""
            if clave == "monto":
                try:
                    n = float(val or 0)
                    val = "$ " + f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                except Exception:
                    val = "$ 0,00"
                pdf.drawRightString(xs[i] + col_widths[i] - 3, y - 2*mm, str(val))
            else:
                txt = str(val)
                max_w = col_widths[i] - 6
                while pdf.stringWidth(txt, "Helvetica", 8.5) > max_w and len(txt) > 1:
                    txt = txt[:-1]
                if len(txt) < len(str(val)):
                    txt = txt[:-1] + "…"
                pdf.drawString(xs[i] + 3, y - 2*mm, txt)
        pdf.setStrokeColorRGB(0.9, 0.92, 0.94)
        pdf.setLineWidth(0.3)
        pdf.line(left, y - 5*mm, left + usable, y - 5*mm)
        pdf.setStrokeColorRGB(0, 0, 0)
        return y - 5.5*mm

    draw_title()
    y = top - 18*mm
    y = draw_header(y)

    total_general = 0.0

    for idx, it in enumerate(items):
        if y < 15*mm:
            pdf.showPage()
            draw_title()
            y = top - 18*mm
            y = draw_header(y)
        data = fila_extractor(it)
        if monto_col_idx is not None:
            try:
                total_general += float(data.get("monto", 0) or 0)
            except Exception:
                pass
        y = draw_row(data, y, alt=(idx % 2 == 1))

    if items and monto_col_idx is not None:
        if y < 15*mm:
            pdf.showPage()
            draw_title()
            y = top - 18*mm

        pdf.setFillColorRGB(0.024, 0.306, 0.224)
        pdf.rect(left, y - 6*mm, usable, 7*mm, fill=1, stroke=0)
        pdf.setFillColorRGB(1, 1, 1)
        pdf.setFont("Helvetica-Bold", 11)
        label_x = xs[monto_col_idx] - 3
        pdf.drawRightString(label_x, y - 2*mm, "TOTAL GENERAL:")
        tot_str = "$ " + f"{total_general:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        pdf.drawRightString(xs[monto_col_idx] + col_widths[monto_col_idx] - 3, y - 2*mm, tot_str)
        pdf.setFillColorRGB(0, 0, 0)

    total_pages = pdf.getPageNumber()
    pdf.setFont("Helvetica", 8)
    pdf.setFillColorRGB(0.4, 0.45, 0.5)
    pdf.drawRightString(width - left, 8*mm, f"Página {total_pages}")

    pdf.save()
    buff.seek(0)
    return FileResponse(buff, as_attachment=True, filename=filename, content_type="application/pdf")


class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        tipo = self.request.query_params.get('tipo')
        if tipo:
            qs = qs.filter(tipo__in=[tipo.upper(), "AMBOS"])
        return qs

class IngresoViewSet(MultiTenantMixin, viewsets.ModelViewSet):
    queryset = Ingreso.objects.all().order_by("-fecha", "-id")
    serializer_class = IngresoSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = LargeResultsSetPagination

    def get_queryset(self):
        user = self.request.user
        es_admin = user.is_superuser or (hasattr(user, 'perfil') and user.perfil.rol == 'ADMIN')

        qs = Ingreso.objects.all().order_by("-fecha", "-id")

        # ── Filtro de oficina ────────────────────────────────────────
        if es_admin:
            oficina_param = self.request.query_params.get('oficina')
            if oficina_param and str(oficina_param).upper() not in ["ALL", ""]:
                keys = _get_todas_las_llaves_oficina(oficina_param)
                qs = qs.filter(oficina_id__in=keys)
        elif hasattr(user, 'perfil') and user.perfil.oficina:
            keys = _get_todas_las_llaves_oficina(user.perfil.oficina)
            qs = qs.filter(oficina_id__in=keys)
        else:
            return Ingreso.objects.none()

        # ── Filtros de historial ─────────────────────────────────────
        desde = self.request.query_params.get('fecha__gte')
        hasta  = self.request.query_params.get('fecha__lte')
        forma  = self.request.query_params.get('forma_pago')
        q      = self.request.query_params.get('search')

        if desde:
            qs = qs.filter(fecha__gte=desde)
        if hasta:
            qs = qs.filter(fecha__lte=hasta)
        if forma and forma.lower() not in ('todas', ''):
            qs = qs.filter(forma_pago__iexact=forma)
        if q:
            qs = qs.filter(
                Q(descripcion__icontains=q) |
                Q(pagado_por__icontains=q) |
                Q(categoria__icontains=q)
            )

        return qs

    def perform_create(self, serializer):
        user = self.request.user
        es_admin = user.is_superuser or (hasattr(user, 'perfil') and user.perfil.rol == 'ADMIN')
        if not es_admin and hasattr(user, 'perfil') and user.perfil.oficina:
            serializer.save(usuario=user, oficina=user.perfil.oficina)
        else:
            serializer.save(usuario=user)

    @action(detail=True, methods=["patch"], url_path="verificar")
    def verificar(self, request, pk=None):
        """
        PATCH /api/balanzes/ingresos/{id}/verificar/
        Marca una transferencia como verificada o no verificada.
        Body: { "verificada": true/false, "nota_verificacion": "..." }
        """
        ingreso = self.get_object()
        verificada = request.data.get("verificada", True)
        nota = request.data.get("nota_verificacion", "")

        if verificada:
            ingreso.verificada = True
            ingreso.verificada_por = request.user
            ingreso.verificada_en = timezone.now()
            ingreso.nota_verificacion = nota or ""
        else:
            # Desmarcar
            ingreso.verificada = False
            ingreso.verificada_por = None
            ingreso.verificada_en = None
            ingreso.nota_verificacion = ""

        ingreso.save(update_fields=["verificada", "verificada_por", "verificada_en", "nota_verificacion"])
        from .serializers import IngresoSerializer
        return Response(IngresoSerializer(ingreso).data)

    @action(detail=False, methods=["get"], url_path="transferencias")
    def transferencias(self, request):
        """
        GET /api/balanzes/ingresos/transferencias/
        Lista solo las transferencias (forma_pago != EFECTIVO) con filtros.
        """
        qs = self.get_queryset().exclude(forma_pago="EFECTIVO").exclude(forma_pago__isnull=True)

        verificada = request.query_params.get("verificada")
        if verificada == "true":
            qs = qs.filter(verificada=True)
        elif verificada == "false":
            qs = qs.filter(verificada=False)

        page = self.paginate_queryset(qs)
        if page is not None:
            from .serializers import IngresoSerializer
            return self.get_paginated_response(IngresoSerializer(page, many=True).data)
        from .serializers import IngresoSerializer
        return Response(IngresoSerializer(qs, many=True).data)

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """
        GET /api/balanzes/ingresos/export/
        Versión vieja (mantenida por compatibilidad).
        Para la NUEVA descarga prolija usá /api/balanzes/ingresos/historial/?export=xlsx
        """
        qs = self.get_queryset()

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse("openpyxl no está instalado. Ejecutá: pip install openpyxl", status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = "Historial de Pagos"

        headers = ["Fecha", "Hora", "Descripción", "Enviado por", "CUIT/CUIL remitente", "Cuenta destino", "N° Operación", "Forma de pago", "Categoría", "Oficina", "Monto"]
        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        fill_par   = PatternFill("solid", fgColor="F8FAFC")
        fill_impar = PatternFill("solid", fgColor="FFFFFF")

        for row_idx, ingreso in enumerate(qs.select_related("oficina"), 2):
            fecha_str = ingreso.fecha.strftime("%d/%m/%Y") if ingreso.fecha else "—"
            hora_str  = ingreso.created_at.strftime("%H:%M") if getattr(ingreso, 'created_at', None) else "—"
            ofi       = getattr(ingreso, 'oficina', None)
            ofi_nombre = ""
            if ofi:
                ofi_nombre = getattr(ofi, 'nombre', None) or getattr(ofi, 'codigo', None) or str(ofi)

            import re as _re
            obs_raw = ingreso.observaciones or ""
            cuit_m  = _re.search(r'CUIT:\s*([^\s|]+)', obs_raw)
            op_m    = _re.search(r'Op:\s*([^\s|]+)', obs_raw)
            cuit_val = (cuit_m.group(1) if cuit_m else "") or "—"
            op_val   = (op_m.group(1)   if op_m   else "") or "—"

            row_fill = fill_par if row_idx % 2 == 0 else fill_impar
            valores  = [
                fecha_str,
                hora_str,
                ingreso.descripcion or "",
                ingreso.pagado_por  or "",
                cuit_val,
                getattr(ingreso, 'billetera', None) or "—",
                op_val,
                (ingreso.forma_pago or "efectivo").upper(),
                ingreso.categoria   or "",
                ofi_nombre,
                float(ingreso.monto or 0),
            ]
            for col_idx, val in enumerate(valores, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = row_fill
                if col_idx == 11:
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

        total_row = qs.count() + 2
        ws.cell(row=total_row, column=10, value="TOTAL").font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=11, value=f"=SUM(K2:K{total_row - 1})")
        total_cell.font = Font(bold=True)
        total_cell.number_format = '"$"#,##0.00'
        total_cell.alignment = Alignment(horizontal="right")

        for i, width in enumerate([12, 8, 38, 26, 20, 24, 18, 16, 18, 14, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        hoy = timezone.localdate().strftime("%Y-%m-%d")
        filename = f"Historial_Pagos_{hoy}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    # ============================================================
    # 🚀 NUEVO: HISTORIAL UNIFICADO (JSON paginado + export Excel/PDF)
    # ============================================================
    @action(detail=False, methods=["get"], url_path="historial")
    def historial(self, request):
        """
        GET /api/balanzes/ingresos/historial/
        Params:
          desde, hasta (YYYY-MM-DD)
          oficina (id, "ALL" o vacío)
          forma_pago ("EFECTIVO", "TRANSFERENCIA", "TODAS")
          search
          page, page_size
          export = "xlsx" | "pdf" (si está, ignora paginación y devuelve archivo)
          all = 1 (sin paginar, devuelve hasta 50k items)
        """
        user = request.user
        es_admin = user.is_superuser or (hasattr(user, 'perfil') and user.perfil.rol == 'ADMIN')

        qs = Ingreso.objects.all().select_related("oficina", "usuario").order_by("-fecha", "-id")

        if es_admin:
            oficina_param = request.query_params.get('oficina')
            if oficina_param and str(oficina_param).upper() not in ["ALL", ""]:
                keys = _get_todas_las_llaves_oficina(oficina_param)
                qs = qs.filter(oficina_id__in=keys)
        elif hasattr(user, 'perfil') and user.perfil.oficina:
            keys = _get_todas_las_llaves_oficina(user.perfil.oficina)
            qs = qs.filter(oficina_id__in=keys)
        else:
            qs = Ingreso.objects.none()

        desde_raw = request.query_params.get('desde') or request.query_params.get('fecha__gte')
        hasta_raw = request.query_params.get('hasta') or request.query_params.get('fecha__lte')
        forma = request.query_params.get('forma_pago')
        q = request.query_params.get('search') or request.query_params.get('q')

        d_desde = _parse_ymd(desde_raw) if desde_raw else None
        d_hasta = _parse_ymd(hasta_raw) if hasta_raw else None

        if d_desde:
            qs = qs.filter(fecha__gte=d_desde)
        if d_hasta:
            qs = qs.filter(fecha__lte=d_hasta)
        if forma and forma.lower() not in ('todas', ''):
            qs = qs.filter(forma_pago__iexact=forma)
        if q:
            qs = qs.filter(
                Q(descripcion__icontains=q) |
                Q(pagado_por__icontains=q) |
                Q(categoria__icontains=q) |
                Q(billetera__icontains=q)
            )

        export = (request.query_params.get('export') or '').strip().lower()

        def _fila(ing):
            ofi = getattr(ing, 'oficina', None)
            ofi_nombre = getattr(ofi, 'nombre', None) or ""
            fecha_str = ing.fecha.strftime("%d/%m/%Y") if ing.fecha else "—"
            return {
                "fecha": fecha_str,
                "descripcion": ing.descripcion or "—",
                "categoria": ing.categoria or "—",
                "forma_pago": (ing.forma_pago or "EFECTIVO").upper(),
                "pagado_por": ing.pagado_por or "—",
                "billetera": ing.billetera or "—",
                "oficina": ofi_nombre or "—",
                "monto": float(ing.monto or 0),
            }

        columnas_xlsx = [
            ("fecha", "Fecha"),
            ("descripcion", "Descripción"),
            ("categoria", "Categoría"),
            ("forma_pago", "Forma de pago"),
            ("pagado_por", "Pagado por"),
            ("billetera", "Cuenta destino"),
            ("oficina", "Oficina"),
            ("monto", "Monto"),
        ]

        if export in ("xlsx", "excel"):
            try:
                rango = ""
                if d_desde and d_hasta:
                    rango = f"_{d_desde.strftime('%Y-%m-%d')}_a_{d_hasta.strftime('%Y-%m-%d')}"
                titulo = "Historial de Ingresos"
                if d_desde and d_hasta:
                    titulo += f" - {d_desde.strftime('%d/%m/%Y')} a {d_hasta.strftime('%d/%m/%Y')}"
                filename = f"Historial_Ingresos{rango}.xlsx"
                items = list(qs[:50000])
                return _build_historial_xlsx_response(
                    items,
                    titulo=titulo,
                    columnas=columnas_xlsx,
                    fila_extractor=_fila,
                    filename=filename,
                    color_header="047857",
                )
            except Exception as e:
                import traceback as _tb
                tb_str = _tb.format_exc()
                print("=" * 80)
                print("[HISTORIAL INGRESOS XLSX] ERROR:")
                print(tb_str)
                print("=" * 80)
                return Response(
                    {"detail": f"Error generando Excel: {str(e)}", "trace": tb_str},
                    status=500,
                )

        if export == "pdf":
            try:
                rango = ""
                if d_desde and d_hasta:
                    rango = f"_{d_desde.strftime('%Y-%m-%d')}_a_{d_hasta.strftime('%Y-%m-%d')}"
                titulo = "Historial de Ingresos"
                if d_desde and d_hasta:
                    titulo += f" - {d_desde.strftime('%d/%m/%Y')} a {d_hasta.strftime('%d/%m/%Y')}"
                filename = f"Historial_Ingresos{rango}.pdf"
                items = list(qs[:50000])
                columnas_pdf = [
                    ("fecha", "Fecha", 0.09),
                    ("descripcion", "Descripción", 0.22),
                    ("categoria", "Categoría", 0.13),
                    ("forma_pago", "Forma", 0.10),
                    ("pagado_por", "Pagado por", 0.15),
                    ("billetera", "Cta. destino", 0.13),
                    ("oficina", "Oficina", 0.10),
                    ("monto", "Monto", 0.08),
                ]
                return _build_historial_pdf_response(
                    items,
                    titulo=titulo,
                    columnas=columnas_pdf,
                    fila_extractor=_fila,
                    filename=filename,
                    color_header_rgb=(4, 120, 87),
                )
            except Exception as e:
                import traceback as _tb
                tb_str = _tb.format_exc()
                print("=" * 80)
                print("[HISTORIAL INGRESOS PDF] ERROR:")
                print(tb_str)
                print("=" * 80)
                return Response(
                    {"detail": f"Error generando PDF: {str(e)}", "trace": tb_str},
                    status=500,
                )

        # Respuesta JSON normal con paginación
        all_flag = str(request.query_params.get('all') or '').strip().lower() in ("1", "true", "yes", "y")
        if all_flag:
            items = list(qs[:50000])
            from .serializers import IngresoSerializer
            ser = IngresoSerializer(items, many=True)
            return Response({"count": len(items), "results": ser.data, "all": True}, status=200)

        page = self.paginate_queryset(qs)
        if page is not None:
            from .serializers import IngresoSerializer
            return self.get_paginated_response(IngresoSerializer(page, many=True).data)

        from .serializers import IngresoSerializer
        return Response(IngresoSerializer(qs, many=True).data)



class EgresoViewSet(MultiTenantMixin, viewsets.ModelViewSet): 
    queryset = Egreso.objects.all().order_by("-fecha", "-id")
    serializer_class = EgresoSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = LargeResultsSetPagination

    def get_queryset(self):
        user = self.request.user
        es_admin = user.is_superuser or (hasattr(user, 'perfil') and user.perfil.rol == 'ADMIN')
        
        qs = Egreso.objects.all().order_by("-fecha", "-id")

        if es_admin:
            oficina_param = self.request.query_params.get('oficina')
            if oficina_param and str(oficina_param).upper() not in ["ALL", ""]:
                keys = _get_todas_las_llaves_oficina(oficina_param)
                qs = qs.filter(oficina_id__in=keys)
        elif hasattr(user, 'perfil') and user.perfil.oficina:
            keys = _get_todas_las_llaves_oficina(user.perfil.oficina)
            qs = qs.filter(oficina_id__in=keys)
        else:
            return Egreso.objects.none()

        # Filtros de fecha
        desde = self.request.query_params.get('fecha__gte')
        hasta  = self.request.query_params.get('fecha__lte')
        if desde: qs = qs.filter(fecha__gte=desde)
        if hasta:  qs = qs.filter(fecha__lte=hasta)

        return qs
        
    def perform_create(self, serializer):
        user = self.request.user
        es_admin = user.is_superuser or (hasattr(user, 'perfil') and user.perfil.rol == 'ADMIN')
        
        if not es_admin and hasattr(user, 'perfil') and user.perfil.oficina:
            serializer.save(usuario=user, oficina=user.perfil.oficina)
        else:
            serializer.save(usuario=user)

    # ============================================================
    # 🚀 NUEVO: HISTORIAL UNIFICADO DE EGRESOS
    # ============================================================
    @action(detail=False, methods=["get"], url_path="historial")
    def historial(self, request):
        """
        GET /api/balanzes/egresos/historial/
        Params: desde, hasta, oficina, forma_pago, search, page, page_size,
                export = "xlsx" | "pdf", all = 1
        """
        user = request.user
        es_admin = user.is_superuser or (hasattr(user, 'perfil') and user.perfil.rol == 'ADMIN')

        qs = Egreso.objects.all().select_related("oficina", "usuario").order_by("-fecha", "-id")

        if es_admin:
            oficina_param = request.query_params.get('oficina')
            if oficina_param and str(oficina_param).upper() not in ["ALL", ""]:
                keys = _get_todas_las_llaves_oficina(oficina_param)
                qs = qs.filter(oficina_id__in=keys)
        elif hasattr(user, 'perfil') and user.perfil.oficina:
            keys = _get_todas_las_llaves_oficina(user.perfil.oficina)
            qs = qs.filter(oficina_id__in=keys)
        else:
            qs = Egreso.objects.none()

        desde_raw = request.query_params.get('desde') or request.query_params.get('fecha__gte')
        hasta_raw = request.query_params.get('hasta') or request.query_params.get('fecha__lte')
        forma = request.query_params.get('forma_pago')
        q = request.query_params.get('search') or request.query_params.get('q')

        d_desde = _parse_ymd(desde_raw) if desde_raw else None
        d_hasta = _parse_ymd(hasta_raw) if hasta_raw else None

        if d_desde:
            qs = qs.filter(fecha__gte=d_desde)
        if d_hasta:
            qs = qs.filter(fecha__lte=d_hasta)
        if forma and forma.lower() not in ('todas', ''):
            qs = qs.filter(forma_pago__iexact=forma)
        if q:
            qs = qs.filter(
                Q(descripcion__icontains=q) |
                Q(categoria__icontains=q)
            )

        export = (request.query_params.get('export') or '').strip().lower()

        def _fila(eg):
            ofi = getattr(eg, 'oficina', None)
            ofi_nombre = getattr(ofi, 'nombre', None) or ""
            fecha_str = eg.fecha.strftime("%d/%m/%Y") if eg.fecha else "—"
            return {
                "fecha": fecha_str,
                "descripcion": eg.descripcion or "—",
                "categoria": eg.categoria or "—",
                "forma_pago": (eg.forma_pago or "EFECTIVO").upper(),
                "oficina": ofi_nombre or "—",
                "monto": float(eg.monto or 0),
            }

        columnas_xlsx = [
            ("fecha", "Fecha"),
            ("descripcion", "Descripción"),
            ("categoria", "Categoría"),
            ("forma_pago", "Forma de pago"),
            ("oficina", "Oficina"),
            ("monto", "Monto"),
        ]

        if export in ("xlsx", "excel"):
            try:
                rango = ""
                if d_desde and d_hasta:
                    rango = f"_{d_desde.strftime('%Y-%m-%d')}_a_{d_hasta.strftime('%Y-%m-%d')}"
                titulo = "Historial de Egresos"
                if d_desde and d_hasta:
                    titulo += f" - {d_desde.strftime('%d/%m/%Y')} a {d_hasta.strftime('%d/%m/%Y')}"
                filename = f"Historial_Egresos{rango}.xlsx"
                items = list(qs[:50000])
                return _build_historial_xlsx_response(
                    items,
                    titulo=titulo,
                    columnas=columnas_xlsx,
                    fila_extractor=_fila,
                    filename=filename,
                    color_header="9F1239",
                )
            except Exception as e:
                import traceback as _tb
                tb_str = _tb.format_exc()
                print("=" * 80)
                print("[HISTORIAL EGRESOS XLSX] ERROR:")
                print(tb_str)
                print("=" * 80)
                return Response(
                    {"detail": f"Error generando Excel: {str(e)}", "trace": tb_str},
                    status=500,
                )

        if export == "pdf":
            try:
                rango = ""
                if d_desde and d_hasta:
                    rango = f"_{d_desde.strftime('%Y-%m-%d')}_a_{d_hasta.strftime('%Y-%m-%d')}"
                titulo = "Historial de Egresos"
                if d_desde and d_hasta:
                    titulo += f" - {d_desde.strftime('%d/%m/%Y')} a {d_hasta.strftime('%d/%m/%Y')}"
                filename = f"Historial_Egresos{rango}.pdf"
                items = list(qs[:50000])
                columnas_pdf = [
                    ("fecha", "Fecha", 0.10),
                    ("descripcion", "Descripción", 0.38),
                    ("categoria", "Categoría", 0.17),
                    ("forma_pago", "Forma", 0.12),
                    ("oficina", "Oficina", 0.13),
                    ("monto", "Monto", 0.10),
                ]
                return _build_historial_pdf_response(
                    items,
                    titulo=titulo,
                    columnas=columnas_pdf,
                    fila_extractor=_fila,
                    filename=filename,
                    color_header_rgb=(159, 18, 57),
                )
            except Exception as e:
                import traceback as _tb
                tb_str = _tb.format_exc()
                print("=" * 80)
                print("[HISTORIAL EGRESOS PDF] ERROR:")
                print(tb_str)
                print("=" * 80)
                return Response(
                    {"detail": f"Error generando PDF: {str(e)}", "trace": tb_str},
                    status=500,
                )

        all_flag = str(request.query_params.get('all') or '').strip().lower() in ("1", "true", "yes", "y")
        if all_flag:
            items = list(qs[:50000])
            from .serializers import EgresoSerializer
            ser = EgresoSerializer(items, many=True)
            return Response({"count": len(items), "results": ser.data, "all": True}, status=200)

        page = self.paginate_queryset(qs)
        if page is not None:
            from .serializers import EgresoSerializer
            return self.get_paginated_response(EgresoSerializer(page, many=True).data)

        from .serializers import EgresoSerializer
        return Response(EgresoSerializer(qs, many=True).data)


class BalanceViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated] 

    def _parse_fecha(self, request):
        raw = request.query_params.get("fecha") or (request.data.get("fecha") if hasattr(request, "data") else None)
        if raw:
            try: return datetime.fromisoformat(raw).date()
            except Exception: return None
        return timezone.localdate()

    def _get_seguridad_oficina(self, request, requested_oficina):
        user = request.user
        es_admin = user.is_superuser or (hasattr(user, 'perfil') and user.perfil.rol == 'ADMIN')
        
        if es_admin:
            if not requested_oficina or str(requested_oficina).upper() in ["ALL", "NULL", "UNDEFINED", ""]:
                return None
            return _get_todas_las_llaves_oficina(requested_oficina)
            
        if hasattr(user, 'perfil') and user.perfil.oficina:
            return _get_todas_las_llaves_oficina(user.perfil.oficina)
            
        return "BLOQUEADO"


    def _desde_ultimo_cierre_hoy(self, fecha, oficina_id):
        """
        Si ya hubo un cierre de caja HOY en esta oficina (ej: el de mediodía),
        devuelve la fecha/hora exacta de ese cierre. Sirve para que "Efectivo
        esperado" NO repita la plata que ya se contó y se cerró en ese cierre
        anterior — mismo criterio que usa recaudacion/views.py al armar el
        cierre de la noche. Para fechas que no son HOY no aplica (ya cerraron
        todo ese día, mostrar el total completo es lo correcto ahí).
        """
        if not oficina_id or fecha != timezone.localdate():
            return None
        try:
            from recaudacion.models import CierreCaja
            ultimo = (
                CierreCaja.objects.filter(oficina_id=oficina_id, creado_en__date=fecha)
                .order_by("-creado_en")
                .first()
            )
            return ultimo.creado_en if ultimo else None
        except Exception:
            return None
    def _build_balance_from_qs(self, fecha, ingresos_qs, egresos_qs):
        total_ingresos = ingresos_qs.aggregate(t=Coalesce(Sum("monto"), Decimal("0")))["t"]
        total_egresos = egresos_qs.aggregate(t=Coalesce(Sum("monto"), Decimal("0")))["t"]
        ingresos_cantidad = ingresos_qs.aggregate(c=Coalesce(Count("id"), 0))["c"]
        egresos_cantidad = egresos_qs.aggregate(c=Coalesce(Count("id"), 0))["c"]

        try:
            pagadores_distintos = ingresos_qs.exclude(pagado_por__isnull=True).exclude(pagado_por__exact="").values("pagado_por").distinct().count()
        except Exception:
            pagadores_distintos = 0

        ingresos_por_forma_dict = {}
        for item in ingresos_qs.values("forma_pago", "monto"):
            forma_key = (item.get("forma_pago") or "SIN FORMA").upper()
            monto = item.get("monto") or Decimal("0")
            if forma_key not in ingresos_por_forma_dict:
                ingresos_por_forma_dict[forma_key] = {"forma_pago": forma_key, "total": Decimal("0"), "cantidad": 0}
            ingresos_por_forma_dict[forma_key]["total"] += monto
            ingresos_por_forma_dict[forma_key]["cantidad"] += 1

        ingresos_por_forma_out = [
            {"forma_pago": it["forma_pago"], "total": _d_to_str(it["total"]), "cantidad": it["cantidad"]} 
            for it in sorted(ingresos_por_forma_dict.values(), key=lambda x: x["forma_pago"])
        ]
            
        egresos_por_forma_dict = {}
        for item in egresos_qs.values("forma_pago", "monto"):
            forma_key = (item.get("forma_pago") or "EFECTIVO").upper()
            monto = item.get("monto") or Decimal("0")
            if forma_key not in egresos_por_forma_dict:
                egresos_por_forma_dict[forma_key] = {"forma_pago": forma_key, "total": Decimal("0"), "cantidad": 0}
            egresos_por_forma_dict[forma_key]["total"] += monto
            egresos_por_forma_dict[forma_key]["cantidad"] += 1
            
        egresos_por_forma_out = [
            {"forma_pago": it["forma_pago"], "total": _d_to_str(it["total"]), "cantidad": it["cantidad"]} 
            for it in sorted(egresos_por_forma_dict.values(), key=lambda x: x["forma_pago"])
        ]

        ingresos_efectivo = ingresos_por_forma_dict.get("EFECTIVO", {}).get("total", Decimal("0"))
        egresos_efectivo = egresos_por_forma_dict.get("EFECTIVO", {}).get("total", Decimal("0"))
        saldo_caja_chica = ingresos_efectivo - egresos_efectivo

        # 🚀 DETALLE de ingresos EN EFECTIVO (para el ticket de cierre): nombre, monto y hora exacta.
        detalle_efectivo = []
        try:
            ef_qs = (
                ingresos_qs.filter(forma_pago__iexact="EFECTIVO")
                .order_by("created_at", "id")
                .values("pagado_por", "monto", "created_at")
            )
            for ing in ef_qs:
                ca = ing.get("created_at")
                hora = ""
                if ca:
                    try:
                        hora = timezone.localtime(ca).strftime("%H:%M")
                    except Exception:
                        try:
                            hora = ca.strftime("%H:%M")
                        except Exception:
                            hora = ""
                detalle_efectivo.append({
                    "pagado_por": (ing.get("pagado_por") or "").strip(),
                    "monto": _d_to_str(ing.get("monto") or Decimal("0")),
                    "hora": hora,
                })
        except Exception:
            detalle_efectivo = []

        return {
            "fecha_iso": fecha.isoformat(),
            "fecha_hum": fecha.strftime("%d/%m/%Y"),
            "totales": {
                "ingresos": _d_to_str(total_ingresos),
                "egresos": _d_to_str(total_egresos),
                "balance": _d_to_str((total_ingresos or 0) - (total_egresos or 0)),
                "saldo_caja_chica": _d_to_str(saldo_caja_chica), 
                "ingresos_cantidad": int(ingresos_cantidad or 0),
                "egresos_cantidad": int(egresos_cantidad or 0),
                "pagadores_distintos": int(pagadores_distintos or 0),
            },
            "ingresos": {"por_forma_pago": ingresos_por_forma_out, "detalle_efectivo": detalle_efectivo},
            "egresos": {"por_forma_pago": egresos_por_forma_out}
        }

    def _build_balance(self, fecha, oficina_keys=None):
        if oficina_keys:
            ingresos_qs = Ingreso.objects.filter(fecha=fecha, oficina_id__in=oficina_keys)
            egresos_qs = Egreso.objects.filter(fecha=fecha, oficina_id__in=oficina_keys)

            # 🆕 No repetir la plata ya contada en un cierre anterior de HOY (ej: mediodía).
            if len(oficina_keys) == 1:
                desde = self._desde_ultimo_cierre_hoy(fecha, oficina_keys[0])
                if desde:
                    ingresos_qs = ingresos_qs.filter(created_at__gt=desde)
                    egresos_qs = egresos_qs.filter(created_at__gt=desde)

            ofi_obj = Oficina.objects.filter(id=oficina_keys[0]).first() if oficina_keys else None
            ofi_label = ofi_obj.nombre if ofi_obj else "Tu Sucursal"
            
            payload = self._build_balance_from_qs(fecha, ingresos_qs, egresos_qs)
            payload["scope"] = {"oficina": oficina_keys[0], "oficina_nombre": ofi_label}
            return payload

        ingresos_all = Ingreso.objects.filter(fecha=fecha)
        egresos_all = Egreso.objects.filter(fecha=fecha)
        general = self._build_balance_from_qs(fecha, ingresos_all, egresos_all)

        por_oficina = []
        for ofi in Oficina.objects.all():
            ing_ofi = ingresos_all.filter(oficina=ofi)
            egr_ofi = egresos_all.filter(oficina=ofi)

            # 🆕 Mismo criterio, por sucursal: no repetir lo ya cerrado hoy.
            desde = self._desde_ultimo_cierre_hoy(fecha, ofi.id)
            if desde:
                ing_ofi = ing_ofi.filter(created_at__gt=desde)
                egr_ofi = egr_ofi.filter(created_at__gt=desde)

            block = self._build_balance_from_qs(fecha, ing_ofi, egr_ofi)
            block["scope"] = {"oficina": ofi.id, "oficina_nombre": ofi.nombre}
            por_oficina.append(block)

        sin_oficina = None
        iq_none = ingresos_all.filter(oficina__isnull=True)
        eq_none = egresos_all.filter(oficina__isnull=True)
        if iq_none.exists() or eq_none.exists():
            sin_oficina = self._build_balance_from_qs(fecha, iq_none, eq_none)
            sin_oficina["scope"] = {"oficina": None, "oficina_nombre": "SIN OFICINA"}

        general["por_oficina"] = por_oficina
        if sin_oficina:
            general["sin_oficina"] = sin_oficina

        return general

    @action(detail=False, methods=["get"])
    def balance_diario(self, request):
        try:
            fecha = self._parse_fecha(request)
            if fecha is None: return Response({"detail": "Fecha inválida."}, status=400)

            req_ofi = request.query_params.get("oficina")
            keys = self._get_seguridad_oficina(request, req_ofi)
            if keys == "BLOQUEADO": return Response({"detail": "No autorizado."}, status=403)

            return Response(self._build_balance(fecha, oficina_keys=keys), status=200)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    def _build_balance_rango(self, desde, hasta, oficina_keys=None):
        """
        Igual que _build_balance pero sumando un RANGO de fechas (mes completo),
        no un solo día. Reutiliza _build_balance_from_qs (suma en el backend).
        """
        rango = Q(fecha__gte=desde, fecha__lte=hasta)

        if oficina_keys:
            ingresos_qs = Ingreso.objects.filter(rango, oficina_id__in=oficina_keys)
            egresos_qs = Egreso.objects.filter(rango, oficina_id__in=oficina_keys)

            ofi_obj = Oficina.objects.filter(id=oficina_keys[0]).first() if oficina_keys else None
            ofi_label = ofi_obj.nombre if ofi_obj else "Tu Sucursal"

            payload = self._build_balance_from_qs(desde, ingresos_qs, egresos_qs)
            payload["scope"] = {"oficina": oficina_keys[0], "oficina_nombre": ofi_label}
            payload["rango"] = {"desde": desde.isoformat(), "hasta": hasta.isoformat()}
            return payload

        ingresos_all = Ingreso.objects.filter(rango)
        egresos_all = Egreso.objects.filter(rango)
        general = self._build_balance_from_qs(desde, ingresos_all, egresos_all)

        por_oficina = []
        for ofi in Oficina.objects.all():
            block = self._build_balance_from_qs(
                desde,
                ingresos_all.filter(oficina=ofi),
                egresos_all.filter(oficina=ofi),
            )
            block["scope"] = {"oficina": ofi.id, "oficina_nombre": ofi.nombre}
            por_oficina.append(block)

        sin_oficina = None
        iq_none = ingresos_all.filter(oficina__isnull=True)
        eq_none = egresos_all.filter(oficina__isnull=True)
        if iq_none.exists() or eq_none.exists():
            sin_oficina = self._build_balance_from_qs(desde, iq_none, eq_none)
            sin_oficina["scope"] = {"oficina": None, "oficina_nombre": "SIN OFICINA"}

        general["por_oficina"] = por_oficina
        if sin_oficina:
            general["sin_oficina"] = sin_oficina
        general["rango"] = {"desde": desde.isoformat(), "hasta": hasta.isoformat()}
        return general

    @action(detail=False, methods=["get"], url_path="balance-mensual")
    def balance_mensual(self, request):
        """
        Totales sumados de TODO un mes (o un rango), calculados en el backend.
        Params:
          ?mes=YYYY-MM           (default: mes actual)
          ?desde=&hasta=         (opcional, rango explícito YYYY-MM-DD)
          ?oficina=<id|ALL>      (respeta el escudo de sucursal)
        Misma estructura que balance-diario (totales, por_forma_pago, por_oficina).
        """
        try:
            hoy = timezone.localdate()
            mes_raw   = (request.query_params.get("mes")   or "").strip()
            desde_raw = (request.query_params.get("desde") or "").strip()
            hasta_raw = (request.query_params.get("hasta") or "").strip()

            if desde_raw and hasta_raw:
                desde = parse_date(desde_raw)
                hasta = parse_date(hasta_raw)
            else:
                if mes_raw:
                    try:
                        y, m = mes_raw.split("-")[:2]
                        y, m = int(y), int(m)
                    except Exception:
                        y, m = hoy.year, hoy.month
                else:
                    y, m = hoy.year, hoy.month
                desde = datetime(y, m, 1).date()
                if m == 12:
                    hasta = datetime(y, 12, 31).date()
                else:
                    hasta = (datetime(y, m + 1, 1) - timedelta(days=1)).date()

            if not desde or not hasta:
                return Response({"detail": "Rango inválido."}, status=400)

            req_ofi = request.query_params.get("oficina")
            keys = self._get_seguridad_oficina(request, req_ofi)
            if keys == "BLOQUEADO":
                return Response({"detail": "No autorizado."}, status=403)

            return Response(self._build_balance_rango(desde, hasta, oficina_keys=keys), status=200)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    @action(detail=False, methods=["post"])
    def enviar_balance(self, request):
        try:
            fecha = self._parse_fecha(request)
            keys = self._get_seguridad_oficina(request, request.data.get("oficina"))
            if keys == "BLOQUEADO": return Response({"detail": "No autorizado."}, status=403)

            data = self._build_balance(fecha, oficina_keys=keys)
            destinatarios = ["1164235336"]
            resultados = []
            all_ok = True

            for numero in destinatarios:
                ok, info = enviar_balance_por_whatsapp(fecha=fecha, data=data, destinatario=numero)
                resultados.append({"numero": numero, "ok": bool(ok), "info": info})
                if not ok: all_ok = False

            status_code = 200 if all_ok else 502
            return Response({"detail": "Envío procesado", "resultados": resultados}, status=status_code)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    @action(detail=False, methods=["get"])
    def exportar_excel(self, request):
        try:
            fecha = self._parse_fecha(request)
            keys = self._get_seguridad_oficina(request, request.query_params.get("oficina"))
            if keys == "BLOQUEADO": return Response({"detail": "No autorizado."}, status=403)

            if keys:
                ingresos = Ingreso.objects.filter(fecha=fecha, oficina_id__in=keys).order_by('created_at')
                egresos = Egreso.objects.filter(fecha=fecha, oficina_id__in=keys).order_by('created_at')
                polizas_nuevas = Poliza.objects.filter(oficina_id__in=keys, fecha_emision=fecha)
                polizas_bajas = Poliza.objects.filter(oficina_id__in=keys, fecha_baja=fecha)
            else:
                ingresos = Ingreso.objects.filter(fecha=fecha).order_by('created_at')
                egresos = Egreso.objects.filter(fecha=fecha).order_by('created_at')
                polizas_nuevas = Poliza.objects.filter(fecha_emision=fecha)
                polizas_bajas = Poliza.objects.filter(fecha_baja=fecha)

            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            
            formato_titulo = workbook.add_format({'bold': True, 'font_size': 16, 'bg_color': '#0F172A', 'font_color': 'white', 'align': 'center'})
            formato_subtitulo = workbook.add_format({'bold': True, 'font_size': 12, 'bg_color': '#1E293B', 'font_color': 'white', 'bottom': 2})
            formato_cabecera = workbook.add_format({'bold': True, 'bg_color': '#334155', 'font_color': 'white', 'border': 1, 'align': 'center'})
            formato_cabecera_verde = workbook.add_format({'bold': True, 'bg_color': '#059669', 'font_color': 'white', 'border': 1, 'align': 'center'})
            formato_cabecera_rojo = workbook.add_format({'bold': True, 'bg_color': '#E11D48', 'font_color': 'white', 'border': 1, 'align': 'center'})
            formato_moneda = workbook.add_format({'num_format': '$#,##0.00', 'border': 1})
            formato_moneda_bold = workbook.add_format({'bold': True, 'num_format': '$#,##0.00', 'border': 1, 'bg_color': '#F1F5F9'})
            formato_normal = workbook.add_format({'border': 1})
            formato_centro = workbook.add_format({'border': 1, 'align': 'center'})

            ws_dash = workbook.add_worksheet('Dashboard')
            ws_dash.set_column('A:A', 30)
            ws_dash.set_column('B:F', 20)
            ws_dash.set_column('H:K', 20)
            
            ws_dash.merge_range('A1:F1', f'REPORTE GERENCIAL - {fecha.strftime("%d/%m/%Y")}', formato_titulo)
            
            tot_ing = sum(i.monto for i in ingresos)
            tot_eg = sum(e.monto for e in egresos)
            neto = tot_ing - tot_eg
            
            ws_dash.write('A3', 'TOTALES DEL DÍA', formato_subtitulo)
            ws_dash.write('A4', 'Total Ingresos', formato_cabecera_verde)
            ws_dash.write('B4', tot_ing, formato_moneda_bold)
            ws_dash.write('A5', 'Total Egresos', formato_cabecera_rojo)
            ws_dash.write('B5', tot_eg, formato_moneda_bold)
            ws_dash.write('A6', 'Balance Neto', formato_cabecera)
            ws_dash.write('B6', neto, formato_moneda_bold)

            ws_dash.write('A8', 'RECAUDACIÓN POR MEDIO DE PAGO', formato_subtitulo)
            ws_dash.write('A9', 'Medio de Pago', formato_cabecera)
            ws_dash.write('B9', 'Monto Total', formato_cabecera)
            ws_dash.write('C9', 'Cantidad Ops.', formato_cabecera)
            
            pagos_map = {}
            for i in ingresos:
                fp = i.forma_pago or "EFECTIVO"
                if fp not in pagos_map: pagos_map[fp] = {'monto': Decimal('0'), 'cant': 0}
                pagos_map[fp]['monto'] += i.monto
                pagos_map[fp]['cant'] += 1
                
            fila = 9
            for fp, data in sorted(pagos_map.items(), key=lambda x: x[1]['monto'], reverse=True):
                ws_dash.write(fila, 0, fp, formato_normal)
                ws_dash.write(fila, 1, data['monto'], formato_moneda)
                ws_dash.write(fila, 2, data['cant'], formato_centro)
                fila += 1

            if pagos_map:
                chart = workbook.add_chart({'type': 'pie'})
                chart.add_series({
                    'name': 'Ingresos por Medio de Pago',
                    'categories': ['Dashboard', 9, 0, fila - 1, 0],
                    'values':     ['Dashboard', 9, 1, fila - 1, 1],
                    'data_labels': {'percentage': True, 'leader_lines': True}
                })
                chart.set_title({'name': 'Distribución de Ingresos'})
                chart.set_style(10)
                ws_dash.insert_chart('E3', chart)

            fila += 2
            ws_dash.write(fila, 0, 'COMPARATIVA DE RENDIMIENTO POR SUCURSAL', formato_subtitulo)
            ws_dash.write(fila + 1, 0, 'Sucursal', formato_cabecera)
            ws_dash.write(fila + 1, 1, 'Recaudación ($)', formato_cabecera)
            ws_dash.write(fila + 1, 2, 'Cant. Cobros', formato_cabecera)
            ws_dash.write(fila + 1, 3, 'Gastos / Egresos ($)', formato_cabecera)
            ws_dash.write(fila + 1, 4, 'Pólizas Nuevas', formato_cabecera)
            ws_dash.write(fila + 1, 5, 'Bajas', formato_cabecera)
            
            ofi_map = {}
            
            def _get_ofi_name(obj):
                return obj.oficina.nombre if obj.oficina else "Sin Sucursal"

            for i in ingresos:
                ofi_name = _get_ofi_name(i)
                if ofi_name not in ofi_map: ofi_map[ofi_name] = {'recaudacion': Decimal('0'), 'cobros': 0, 'egresos': Decimal('0'), 'nuevas': 0, 'bajas': 0}
                ofi_map[ofi_name]['recaudacion'] += i.monto
                ofi_map[ofi_name]['cobros'] += 1

            for e in egresos:
                ofi_name = _get_ofi_name(e)
                if ofi_name not in ofi_map: ofi_map[ofi_name] = {'recaudacion': Decimal('0'), 'cobros': 0, 'egresos': Decimal('0'), 'nuevas': 0, 'bajas': 0}
                ofi_map[ofi_name]['egresos'] += e.monto

            for p in polizas_nuevas:
                ofi_name = _get_ofi_name(p)
                if ofi_name not in ofi_map: ofi_map[ofi_name] = {'recaudacion': Decimal('0'), 'cobros': 0, 'egresos': Decimal('0'), 'nuevas': 0, 'bajas': 0}
                ofi_map[ofi_name]['nuevas'] += 1

            for p in polizas_bajas:
                ofi_name = _get_ofi_name(p)
                if ofi_name not in ofi_map: ofi_map[ofi_name] = {'recaudacion': Decimal('0'), 'cobros': 0, 'egresos': Decimal('0'), 'nuevas': 0, 'bajas': 0}
                ofi_map[ofi_name]['bajas'] += 1

            fila_ofi = fila + 2
            for ofi_name, data in sorted(ofi_map.items(), key=lambda x: x[1]['recaudacion'], reverse=True):
                ws_dash.write(fila_ofi, 0, ofi_name, formato_normal)
                ws_dash.write(fila_ofi, 1, data['recaudacion'], formato_moneda)
                ws_dash.write(fila_ofi, 2, data['cobros'], formato_centro)
                ws_dash.write(fila_ofi, 3, data['egresos'], formato_moneda)
                ws_dash.write(fila_ofi, 4, data['nuevas'], formato_centro)
                ws_dash.write(fila_ofi, 5, data['bajas'], formato_centro)
                fila_ofi += 1

            ws_ing = workbook.add_worksheet('Detalle Ingresos')
            cabeceras_ing = ['Hora Exacta', 'Sucursal', 'Descripción', 'Categoría', 'Forma Pago', 'Monto', 'Usuario']
            ws_ing.write_row('A1', cabeceras_ing, formato_cabecera_verde)
            ws_ing.set_column('A:A', 15)
            ws_ing.set_column('B:G', 20)
            ws_ing.set_column('C:C', 40)
            
            for row, i in enumerate(ingresos, start=1):
                hora_str = timezone.localtime(i.created_at).strftime('%H:%M') if i.created_at else "—"
                ofi_str = _get_ofi_name(i)
                usr_str = i.usuario.username if i.usuario else "Sistema"
                ws_ing.write_row(row, 0, [hora_str, ofi_str, i.descripcion, i.categoria, i.forma_pago], formato_normal)
                ws_ing.write(row, 5, i.monto, formato_moneda)
                ws_ing.write(row, 6, usr_str, formato_normal)
                
            ws_ing.autofilter(0, 0, len(ingresos), len(cabeceras_ing) - 1)

            ws_eg = workbook.add_worksheet('Detalle Egresos')
            cabeceras_eg = ['Hora Exacta', 'Sucursal', 'Descripción', 'Categoría', 'Forma Pago', 'Monto', 'Usuario']
            ws_eg.write_row('A1', cabeceras_eg, formato_cabecera_rojo)
            ws_eg.set_column('A:A', 15)
            ws_eg.set_column('B:G', 20)
            ws_eg.set_column('C:C', 40)
            
            for row, e in enumerate(egresos, start=1):
                hora_str = timezone.localtime(e.created_at).strftime('%H:%M') if e.created_at else "—"
                ofi_str = _get_ofi_name(e)
                usr_str = e.usuario.username if e.usuario else "Sistema"
                ws_eg.write_row(row, 0, [hora_str, ofi_str, e.descripcion, e.categoria, e.forma_pago], formato_normal)
                ws_eg.write(row, 5, e.monto, formato_moneda)
                ws_eg.write(row, 6, usr_str, formato_normal)
                
            ws_eg.autofilter(0, 0, len(egresos), len(cabeceras_eg) - 1)

            workbook.close()
            output.seek(0)

            oficina_label = f"_Ofi_{keys[0]}" if keys else "_Todas_Las_Cajas"
            filename = f"Reporte_Diario_{fecha.strftime('%d-%m-%Y')}{oficina_label}.xlsx"
            
            response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response

        except Exception as e:
            return Response({"error": str(e)}, status=500)