# estadisticas/views.py
from calendar import monthrange
from datetime import date, timedelta, datetime, time
import csv
from io import BytesIO
import math

from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import OuterRef, Subquery
from django.core.exceptions import FieldDoesNotExist
from django.db.models import DateTimeField
from django.db.models.functions import TruncDay, TruncMonth, Trunc, TruncWeek, TruncHour

from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from polizas.models import Poliza
from clientes.models import Cliente
from pagos.models import Pago, Cuota 
from .models import PolizaOficinaSnapshot, ExportLog


# -------------------------
# 🚀 AUTO-ACTUALIZADOR DE ESTADOS 
# -------------------------
def auto_marcar_vencidas():
    """
    Delegamos al sincronizador central que está en polizas/views/poliza.py.
    Mantiene compatibilidad con las llamadas existentes.
    """
    try:
        from polizas.views.poliza import auto_marcar_vencidas as _sync_estados
        _sync_estados()
    except Exception:
        # Fallback silencioso para no romper ningún endpoint
        pass


# -------------------------
# 🚀 SUPER TRADUCTOR MULTI-TENANT (A PRUEBA DE FALLOS)
# -------------------------
def _get_seguridad_oficina(request, requested_oficina=""):
    user = request.user
    if not user.is_authenticated:
        return ["BLOQUEADO"]
        
    es_admin = user.is_superuser or (hasattr(user, 'perfil') and user.perfil.rol == 'ADMIN')
    
    target = None
    if es_admin:
        val = str(requested_oficina or "").strip()
        if not val or val.upper() == "ALL":
            return []
        target = val
    else:
        if hasattr(user, 'perfil') and user.perfil.oficina:
            target = user.perfil.oficina
        else:
            return ["BLOQUEADO"]
            
    synonyms = set()
    
    if hasattr(target, 'codigo') and target.codigo:
        synonyms.add(str(target.codigo).strip().lower())
    if hasattr(target, 'id') and target.id:
        synonyms.add(str(target.id).strip().lower())
    if hasattr(target, 'nombre') and target.nombre:
        synonyms.add(str(target.nombre).strip().lower())
        
    if isinstance(target, str):
        s = target.strip().lower()
        synonyms.add(s)
        try:
            from django.apps import apps
            Oficina = apps.get_model("usuarios", "Oficina")
            if s.isdigit():
                ofi = Oficina.objects.filter(Q(codigo=s) | Q(id=s)).first()
            else:
                ofi = Oficina.objects.filter(nombre__icontains=s).first()
            if ofi:
                synonyms.add(str(ofi.codigo).strip().lower())
                synonyms.add(str(ofi.id).strip().lower())
                synonyms.add(str(ofi.nombre).strip().lower())
        except Exception:
            pass
            
    final_synonyms = set(synonyms)
    for s in synonyms:
        if "1" == s or "esquina" in s or "5 esquinas" in s:
            final_synonyms.update(["1", "5 esquinas", "ofi 1", "ofi1"])
        elif "2" == s or "axion" in s:
            final_synonyms.update(["2", "axion", "ofi 2", "ofi2"])
        elif "3" == s or "39" in s or "kilometro" in s:
            final_synonyms.update(["3", "39", "kilometro 39", "ofi 3", "ofi3"])
            
    return list(final_synonyms)


def _is_poliza_oficina_fk() -> bool:
    try:
        f = Poliza._meta.get_field("oficina")
        return bool(getattr(f, "is_relation", False))
    except (FieldDoesNotExist, Exception):
        return False


def _build_oficina_q_from_keys(keys, prefix="poliza__"):
    if not keys: return Q()
    if "BLOQUEADO" in keys: return Q(pk__isnull=True)
    
    is_fk = _is_poliza_oficina_fk()
    q_final = Q()
    
    for k in keys:
        s = str(k).strip()
        if not s: continue
        
        if is_fk:
            if s.isdigit():
                q_final |= Q(**{f"{prefix}oficina_id": int(s)})
            q_final |= Q(**{f"{prefix}oficina__nombre__icontains": s})
            try: q_final |= Q(**{f"{prefix}oficina__codigo": s})
            except Exception: pass
        else:
            q_final |= Q(**{f"{prefix}oficina__icontains": s})
            q_final |= Q(**{f"{prefix}oficina__iexact": s})
            
    return q_final

def _apply_oficina_filter(qs, oficina_keys, is_poliza_model=True):
    if not oficina_keys:
        return qs
    if "BLOQUEADO" in oficina_keys:
        return qs.none()
        
    prefix = "" if is_poliza_model else "poliza__"
    return qs.filter(_build_oficina_q_from_keys(oficina_keys, prefix=prefix))


# -------------------------
# Helpers oficina (normalización)
# -------------------------
def _merge_sum_dict(dst: dict, src: dict):
    if not src: return
    for k, v in src.items():
        try: vv = int(v)
        except Exception:
            try: vv = int(float(v))
            except Exception: vv = 0
        dst[k] = int(dst.get(k, 0)) + vv

def _bucket_oficina(raw) -> str:
    """
    🚀 AHORA DEVUELVE EL ID REAL DE LA OFICINA DIRECTAMENTE.
    """
    if hasattr(raw, 'id') and raw.id is not None: return str(raw.id).strip()
    elif hasattr(raw, 'codigo') and raw.codigo is not None: return str(raw.codigo).strip()
    
    s = str(raw if raw is not None else "").strip()
    if not s: return "SIN_OFICINA"
    
    s_l = s.lower()
    if "5 esquinas" in s_l: return "1"
    if "axion" in s_l: return "2"
    if "kilometro 39" in s_l or "km 39" in s_l or "kilómetro 39" in s_l: return "3"
    
    return s

def _to_bool(v):
    if v is None: return False
    if isinstance(v, bool): return v
    return str(v).strip().lower() in {"1", "true", "t", "yes", "y", "on", "si", "sí"}

def _apply_vehiculos_filters(qs, params, oficina_segura_keys=None):
    tipo = str(params.get("tipo") or params.get("tipo_vehiculo") or "").strip()
    anio = str(params.get("anio") or params.get("año") or "").strip()
    anio_desde = str(params.get("anio_desde") or "").strip()
    anio_hasta = str(params.get("anio_hasta") or "").strip()
    marca = str(params.get("marca") or "").strip()
    modelo = str(params.get("modelo") or "").strip()
    patente = str(params.get("patente") or "").strip()
    solo_activas = _to_bool(params.get("solo_activas"))

    if solo_activas: qs = qs.filter(estado="activa")
    if oficina_segura_keys: qs = _apply_oficina_filter(qs, oficina_segura_keys, is_poliza_model=True)
    if tipo: qs = qs.filter(tipo__iexact=tipo)
    if anio.isdigit(): qs = qs.filter(anio=int(anio))
    if anio_desde.isdigit(): qs = qs.filter(anio__gte=int(anio_desde))
    if anio_hasta.isdigit(): qs = qs.filter(anio__lte=int(anio_hasta))
    if marca: qs = qs.filter(marca__icontains=marca)
    if modelo: qs = qs.filter(modelo__icontains=modelo)
    if patente: qs = qs.filter(patente__icontains=patente)

    return qs

def _clamp_int(v, default, min_v=None, max_v=None):
    try: n = int(v)
    except Exception: n = default
    if min_v is not None: n = max(min_v, n)
    if max_v is not None: n = min(max_v, n)
    return n


# -------------------------
# 🚀 Polizas por oficina (NORMALIZADO Y BLINDADO)
# -------------------------
class PolizasPorOficinaAPIView(APIView):
    """
    GET /api/estadisticas/polizas/por-oficina/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        auto_marcar_vencidas()

        oficina_filtro_raw = str(request.query_params.get("oficina") or "").strip()
        oficina_segura_keys = _get_seguridad_oficina(request, oficina_filtro_raw)
        
        if "BLOQUEADO" in oficina_segura_keys:
            return Response({"error": "Acceso denegado"}, status=403)

        hoy = timezone.localdate()

        try:
            anio = int(request.query_params.get("anio") or hoy.year)
        except (TypeError, ValueError):
            anio = hoy.year

        try:
            mes = int(request.query_params.get("mes") or hoy.month)
            if mes < 1 or mes > 12: mes = hoy.month
        except (TypeError, ValueError):
            mes = hoy.month

        primero_mes = date(anio, mes, 1)
        ultimo_dia = monthrange(anio, mes)[1]
        ultimo_mes = date(anio, mes, ultimo_dia)
        
        dt_primero = timezone.make_aware(datetime.combine(primero_mes, time.min))
        dt_ultimo = timezone.make_aware(datetime.combine(ultimo_mes, time.max))

        compania_filtro = str(request.query_params.get("compania") or "").strip()

        usar_snapshot = _to_bool(request.query_params.get("usar_snapshot"))
        guardar_snapshot = _to_bool(request.query_params.get("guardar_snapshot"))

        if usar_snapshot:
            qs_snap = PolizaOficinaSnapshot.objects.filter(anio=anio, mes=mes)
            if oficina_filtro_raw and oficina_filtro_raw.upper() != "ALL":
                bucket = _bucket_oficina(oficina_filtro_raw)
                qs_snap = qs_snap.filter(oficina=bucket)

            if qs_snap.exists():
                acc = {}
                for snap in qs_snap:
                    code = _bucket_oficina(getattr(snap, "oficina", "") or "")
                    a = acc.setdefault(
                        code,
                        {
                            "oficina": code,
                            "oficina_nombre": getattr(snap, "oficina_nombre", code) or code,
                            "polizas_total": 0, "polizas_activas": 0,
                            "nuevas_mes": 0, "bajas_mes": 0,
                            "por_compania": {}, "por_cobertura": {}, "antiguedad": {},
                            "renovaciones_mes": 0, "en_mora": 0, "calidad_datos": {}
                        },
                    )

                    a["polizas_total"] += int(getattr(snap, "total_polizas", 0) or 0)
                    a["polizas_activas"] += int(getattr(snap, "total_activas", 0) or 0)
                    a["nuevas_mes"] += int(getattr(snap, "nuevas_mes", 0) or 0)
                    a["bajas_mes"] += int(getattr(snap, "bajas_mes", 0) or 0)

                    extra = getattr(snap, "data_extra", None) or {}
                    _merge_sum_dict(a["por_compania"], extra.get("por_compania") or {})
                    _merge_sum_dict(a["por_cobertura"], extra.get("por_cobertura") or {})
                    _merge_sum_dict(a["antiguedad"], extra.get("antiguedad") or {})
                    a["renovaciones_mes"] += int(extra.get("renovaciones_mes") or 0)
                    a["en_mora"] += int(extra.get("en_mora") or 0)
                    
                    cal = extra.get("calidad_datos") or {}
                    if not a["calidad_datos"]: a["calidad_datos"] = cal
                    else: _merge_sum_dict(a["calidad_datos"], cal)

                oficinas_data = []
                for code, a in acc.items():
                    total_polizas = int(a["polizas_total"] or 0)
                    bajas_mes = int(a["bajas_mes"] or 0)
                    en_mora = int(a["en_mora"] or 0)

                    if total_polizas > 0:
                        churn_rate = (bajas_mes + en_mora) / float(total_polizas)
                        churn_porcentaje = round(churn_rate * 100.0, 2)
                    else:
                        churn_rate = 0.0
                        churn_porcentaje = 0.0

                    oficinas_data.append({**a, "churn_rate": churn_rate, "churn_porcentaje": churn_porcentaje})

                return Response({
                    "periodo": f"{anio:04d}-{mes:02d}", "desde": primero_mes.isoformat(),
                    "hasta": ultimo_mes.isoformat(), "oficinas": oficinas_data, "fuente": "snapshot",
                })

        qs_base = Poliza.objects.all()

        if oficina_segura_keys:
            qs_base = _apply_oficina_filter(qs_base, oficina_segura_keys, is_poliza_model=True)
        if compania_filtro:
            qs_base = qs_base.filter(compania__iexact=compania_filtro)

        if not qs_base.exists():
            return Response({
                "periodo": f"{anio:04d}-{mes:02d}", "desde": primero_mes.isoformat(),
                "hasta": ultimo_mes.isoformat(), "oficinas": [], "fuente": "live",
            })

        limite_1 = ultimo_mes - timedelta(days=365)
        limite_3 = ultimo_mes - timedelta(days=365 * 3)
        limite_5 = ultimo_mes - timedelta(days=365 * 5)
        tiene_es_renovacion = hasattr(Poliza, "es_renovacion")

        def _build_metrics(qs, code: str):
            total_polizas = qs.count()
            total_activas = qs.filter(estado="activa").count()

            # ── Última cuota de cada póliza ──────────────────────────────
            # Una póliza está en mora solo si su cuota MÁS RECIENTE
            # está vencida y sin pagar. Si la cuota de enero no se pagó
            # pero la de marzo sí, la póliza está al día.
            ultima_cuota_vto = Cuota.objects.filter(
                poliza=OuterRef("pk")
            ).order_by("-fecha_vencimiento").values("fecha_vencimiento")[:1]

            ultima_cuota_pagada = Cuota.objects.filter(
                poliza=OuterRef("pk")
            ).order_by("-fecha_vencimiento").values("pagado")[:1]

            activas_qs = qs.filter(estado="activa").annotate(
                ultima_vto=Subquery(ultima_cuota_vto),
                ultima_pagada=Subquery(ultima_cuota_pagada),
            )

            # Al día: última cuota pagada O última cuota aún no vencida
            activas_al_dia = activas_qs.filter(
                Q(ultima_pagada=True) | Q(ultima_vto__gte=hoy) | Q(ultima_vto__isnull=True)
            ).count()

            # En mora: última cuota vencida Y sin pagar
            activas_en_mora = activas_qs.filter(
                ultima_pagada=False,
                ultima_vto__lt=hoy,
            ).count()

            # 🔧 FIX: "nuevas_mes" ahora cuenta SOLO altas reales (sin renovaciones).
            # Antes contaba todas las emisiones del mes — incluidas las renovaciones —
            # lo que inflaba el KPI de altas. Ahora filtra es_renovacion=False cuando
            # ese campo existe en el modelo.
            nuevas_qs_mes = qs.filter(fecha_emision__gte=primero_mes, fecha_emision__lte=ultimo_mes)
            if tiene_es_renovacion:
                nuevas_qs_mes = nuevas_qs_mes.filter(es_renovacion=False)
            nuevas_mes = nuevas_qs_mes.count()

            bajas_mes = qs.filter(
                Q(fecha_baja__isnull=False, fecha_baja__gte=primero_mes, fecha_baja__lte=ultimo_mes) |
                Q(baja_operativa__realizada_en__gte=dt_primero, baja_operativa__realizada_en__lte=dt_ultimo)
            ).distinct().count()

            # en_mora — pólizas (cualquier estado) cuya última cuota venció e impaga
            # Usamos la misma lógica que activas_en_mora para consistencia
            en_mora_qs = qs.annotate(
                ultima_vto_mora=Subquery(ultima_cuota_vto),
                ultima_pagada_mora=Subquery(ultima_cuota_pagada),
            )
            en_mora = en_mora_qs.filter(
                ultima_pagada_mora=False,
                ultima_vto_mora__lt=hoy,
            ).count()

            sin_patente = qs.filter(Q(patente__isnull=True) | Q(patente="") | Q(patente="-")).count()
            sin_vehiculo = qs.filter(
                Q(marca__isnull=True) | Q(marca="") | Q(marca="-") | 
                Q(modelo__isnull=True) | Q(modelo="") | Q(modelo="-")
            ).count()
            sin_compania = qs.filter(Q(compania__isnull=True) | Q(compania="") | Q(compania="-")).count()
            
            calidad_datos = {
                "sin_patente": sin_patente,
                "sin_vehiculo": sin_vehiculo,
                "sin_compania": sin_compania
            }

            dist_compania = { str(row["compania"] or "—"): row["c"] for row in qs.values("compania").annotate(c=Count("id")).order_by("-c") }
            
            dist_cobertura = {}
            if hasattr(Poliza, "cobertura"):
                dist_cobertura = { str(row["cobertura"] or "—"): row["c"] for row in qs.values("cobertura").annotate(c=Count("id")).order_by("-c") }

            antiguedad_data = {
                "0_1": qs.filter(fecha_emision__gte=limite_1).count(),
                "1_3": qs.filter(fecha_emision__lt=limite_1, fecha_emision__gte=limite_3).count(),
                "3_5": qs.filter(fecha_emision__lt=limite_3, fecha_emision__gte=limite_5).count(),
                "5_plus": qs.filter(fecha_emision__lt=limite_5).count(),
            }

            renovaciones_mes = 0
            if tiene_es_renovacion:
                renovaciones_mes = qs.filter(es_renovacion=True, fecha_emision__gte=primero_mes, fecha_emision__lte=ultimo_mes).count()

            # Altas nuevas del mes (excluyendo renovaciones si el campo existe)
            if tiene_es_renovacion:
                altas_nuevas_mes = qs.filter(
                    es_renovacion=False,
                    fecha_emision__gte=primero_mes,
                    fecha_emision__lte=ultimo_mes,
                ).count()
            else:
                altas_nuevas_mes = nuevas_mes

            if total_polizas > 0:
                churn_rate = (bajas_mes + en_mora) / float(total_polizas)
                churn_porcentaje = round(churn_rate * 100.0, 2)
            else:
                churn_rate = 0.0
                churn_porcentaje = 0.0

            # Resolver nombre real de la oficina desde la DB
            oficina_nombre_real = code
            try:
                from django.apps import apps
                Oficina = apps.get_model("usuarios", "Oficina")
                if code.isdigit():
                    ofi_obj = Oficina.objects.filter(id=int(code)).first()
                else:
                    ofi_obj = Oficina.objects.filter(nombre__iexact=code).first()
                if ofi_obj:
                    oficina_nombre_real = getattr(ofi_obj, "nombre", code) or code
            except Exception:
                pass

            return {
                "oficina": code,
                "oficina_nombre": oficina_nombre_real,
                "polizas_total": total_polizas,
                "polizas_activas": total_activas,
                "activas_al_dia": activas_al_dia,       # ← activas SIN cuotas vencidas
                "activas_en_mora": activas_en_mora,     # ← activas CON cuota vencida impaga
                "nuevas_mes": nuevas_mes, "altas_nuevas_mes": altas_nuevas_mes,
                "bajas_mes": bajas_mes, "renovaciones_mes": renovaciones_mes,
                "por_compania": dist_compania, "por_cobertura": dist_cobertura,
                "antiguedad": antiguedad_data,
                "en_mora": en_mora, "calidad_datos": calidad_datos,
                "churn_rate": churn_rate, "churn_porcentaje": churn_porcentaje,
            }

        oficinas_data = []
        snapshots_a_guardar = []

        if oficina_segura_keys:
            code = _bucket_oficina(oficina_segura_keys[0])
            data = _build_metrics(qs_base, code)
            oficinas_data.append(data)

            snapshots_a_guardar.append(PolizaOficinaSnapshot(
                oficina=code, anio=anio, mes=mes,
                total_polizas=data["polizas_total"], total_activas=data["polizas_activas"],
                nuevas_mes=data["nuevas_mes"], bajas_mes=data["bajas_mes"],
                data_extra={
                    "por_compania": data["por_compania"], "por_cobertura": data["por_cobertura"],
                    "antiguedad": data["antiguedad"], "renovaciones_mes": data["renovaciones_mes"],
                    "churn_rate": data["churn_rate"], "churn_porcentaje": data["churn_porcentaje"],
                    "en_mora": data["en_mora"], "calidad_datos": data["calidad_datos"],
                    "activas_al_dia": data["activas_al_dia"],
                    "activas_en_mora": data["activas_en_mora"],
                },
            ))
        else:
            # Admin sin filtro de oficina — itera por cada oficina
            qs_sin = qs_base.filter(oficina_id__isnull=True)
            oficinas_con_polizas = qs_base.filter(oficina_id__isnull=False).values_list('oficina_id', flat=True).distinct()

            for ofi_id in oficinas_con_polizas:
                code = str(ofi_id)
                qs_ofi = qs_base.filter(oficina_id=ofi_id)

                if qs_ofi.exists():
                    data = _build_metrics(qs_ofi, code)
                    oficinas_data.append(data)

                    snapshots_a_guardar.append(PolizaOficinaSnapshot(
                        oficina=code, anio=anio, mes=mes,
                        total_polizas=data["polizas_total"], total_activas=data["polizas_activas"],
                        nuevas_mes=data["nuevas_mes"], bajas_mes=data["bajas_mes"],
                        data_extra={
                            "por_compania": data["por_compania"], "por_cobertura": data["por_cobertura"],
                            "antiguedad": data["antiguedad"], "renovaciones_mes": data["renovaciones_mes"],
                            "churn_rate": data["churn_rate"], "churn_porcentaje": data["churn_porcentaje"],
                            "en_mora": data["en_mora"], "calidad_datos": data["calidad_datos"],
                            "activas_al_dia": data["activas_al_dia"],
                            "activas_en_mora": data["activas_en_mora"],
                        },
                    ))

            if qs_sin.exists():
                data = _build_metrics(qs_sin, "SIN_OFICINA")
                oficinas_data.append(data)

        if guardar_snapshot:
            for snap in snapshots_a_guardar:
                PolizaOficinaSnapshot.objects.update_or_create(
                    oficina=snap.oficina, anio=snap.anio, mes=snap.mes,
                    defaults={
                        "total_polizas": snap.total_polizas, "total_activas": snap.total_activas,
                        "nuevas_mes": snap.nuevas_mes, "bajas_mes": snap.bajas_mes, "data_extra": snap.data_extra,
                    },
                )

        try:
            usuario = getattr(request, "user", None)
            ExportLog.objects.create(
                usuario=usuario if getattr(usuario, "is_authenticated", False) else None,
                tipo="polizas_oficina",
                parametros={
                    "anio": anio, "mes": mes, "oficina": oficina_filtro_raw or None,
                    "compania": compania_filtro or None, "usar_snapshot": usar_snapshot,
                    "guardar_snapshot": guardar_snapshot,
                },
            )
        except Exception: pass

        return Response({
            "periodo": f"{anio:04d}-{mes:02d}", "desde": primero_mes.isoformat(),
            "hasta": ultimo_mes.isoformat(), "oficinas": oficinas_data, "fuente": "live",
        })


# -------------------------
# Polizas: emisiones serie por oficina (hora/día/semana/mes) usando fecha_emision
# -------------------------
def _parse_date_iso(v: str):
    try: return date.fromisoformat((v or "").strip())
    except Exception: return None

class EmisionesSeriePorOficinaAPIView(APIView):
    """
    GET /api/estadisticas/polizas/emisiones/serie/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        params = request.query_params

        oficina_filtro_raw = str(params.get("oficina") or "").strip()
        oficina_segura_keys = _get_seguridad_oficina(request, oficina_filtro_raw)
        
        if "BLOQUEADO" in oficina_segura_keys: return Response({"error": "Acceso denegado"}, status=403)

        agrupacion = str(params.get("agrupacion") or "dia").strip().lower()
        if agrupacion not in {"hora", "dia", "semana", "mes"}: agrupacion = "dia"

        desde = _parse_date_iso(params.get("desde"))
        hasta = _parse_date_iso(params.get("hasta")) or timezone.localdate()

        if desde is None:
            if agrupacion in {"hora", "dia"}: desde = hasta - timedelta(days=30)
            elif agrupacion == "semana": desde = hasta - timedelta(weeks=12)
            else: desde = hasta - timedelta(days=365)

        if desde > hasta: desde, hasta = hasta, desde

        try:
            fecha_field = Poliza._meta.get_field("fecha_emision")
            is_datetime = fecha_field.get_internal_type() == "DateTimeField"
        except Exception:
            is_datetime = False

        if agrupacion == "hora" and not is_datetime: agrupacion = "dia"

        qs = Poliza.objects.all()
        if oficina_segura_keys: qs = _apply_oficina_filter(qs, oficina_segura_keys, is_poliza_model=True)

        # Filtro opcional: solo altas nuevas (excluir renovaciones)
        es_renovacion_param = str(params.get("es_renovacion") or "").strip().lower()
        if es_renovacion_param == "false" and hasattr(Poliza, "es_renovacion"):
            qs = qs.filter(es_renovacion=False)
        elif es_renovacion_param == "true" and hasattr(Poliza, "es_renovacion"):
            qs = qs.filter(es_renovacion=True)

        if is_datetime:
            dt_from = timezone.make_aware(datetime.combine(desde, time.min))
            dt_to_excl = timezone.make_aware(datetime.combine(hasta + timedelta(days=1), time.min))
            qs = qs.filter(fecha_emision__gte=dt_from, fecha_emision__lt=dt_to_excl)
        else:
            qs = qs.filter(fecha_emision__gte=desde, fecha_emision__lte=hasta)

        if agrupacion == "hora": trunc_expr = TruncHour("fecha_emision")
        elif agrupacion == "dia": trunc_expr = TruncDay("fecha_emision")
        elif agrupacion == "mes": trunc_expr = TruncMonth("fecha_emision")
        else:
            try: trunc_expr = TruncWeek("fecha_emision")
            except Exception: trunc_expr = Trunc("fecha_emision", "week", output_field=DateTimeField())

        rows = qs.annotate(periodo=trunc_expr).values("periodo", "oficina_id").annotate(c=Count("id")).order_by("periodo")

        # Pre-cargar nombres de oficinas para no hacer una query por cada fila
        try:
            from django.apps import apps
            Oficina = apps.get_model("usuarios", "Oficina")
            oficinas_map = {str(o.id): getattr(o, "nombre", str(o.id)) for o in Oficina.objects.all()}
        except Exception:
            oficinas_map = {}

        period_set = set()
        acc = {}

        for r in rows:
            per = r.get("periodo")
            if per is None: continue

            if hasattr(per, "isoformat"):
                per_str = per.isoformat()
                if agrupacion in {"dia", "semana", "mes"} and hasattr(per, "date"): per_str = per.date().isoformat()
            else: per_str = str(per)

            bucket = _bucket_oficina(r.get("oficina_id"))
            acc.setdefault(bucket, {})
            acc[bucket][per_str] = int(acc[bucket].get(per_str, 0)) + int(r.get("c") or 0)
            period_set.add(per_str)

        periodos = sorted(period_set)
        
        oficinas_payload = []
        for b in acc.keys():
            serie = [{"periodo": p, "cantidad": int(acc[b].get(p, 0))} for p in periodos]
            total = sum(it["cantidad"] for it in serie)
            nombre = oficinas_map.get(b, b)
            oficinas_payload.append({ "oficina": b, "oficina_nombre": nombre, "total": int(total), "serie": serie })

        try:
            usuario = getattr(request, "user", None)
            ExportLog.objects.create(
                usuario=usuario if getattr(usuario, "is_authenticated", False) else None,
                tipo="polizas_emisiones_serie",
                parametros={ "date_field": "fecha_emision", "agrupacion": agrupacion, "desde": desde.isoformat(), "hasta": hasta.isoformat(), "oficina": oficina_filtro_raw or None },
            )
        except Exception: pass

        return Response({ "date_field": "fecha_emision", "agrupacion": agrupacion, "desde": desde.isoformat(), "hasta": hasta.isoformat(), "periodos": periodos, "oficinas": oficinas_payload, "fuente": "live" })


# -------------------------
# Vehículos: resumen + export + LISTADO asegurados
# -------------------------
class VehiculosResumenAPIView(APIView):
    """
    GET /api/estadisticas/vehiculos/resumen/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        params = request.query_params
        
        oficina_filtro_raw = str(params.get("oficina") or "").strip()
        oficina_segura_keys = _get_seguridad_oficina(request, oficina_filtro_raw)
        
        if "BLOQUEADO" in oficina_segura_keys: return Response({"error": "Acceso denegado"}, status=403)

        base = Poliza.objects.all()
        qs = _apply_vehiculos_filters(base, params, oficina_segura_keys=oficina_segura_keys)

        total = qs.count()
        total_activas = qs.filter(estado="activa").count()

        por_tipo = { str(row["tipo"] or "—"): int(row["c"]) for row in qs.values("tipo").annotate(c=Count("id")).order_by("-c") }
        por_anio = { str(row["anio"] or "—"): int(row["c"]) for row in qs.values("anio").annotate(c=Count("id")).order_by("-c") }
        por_oficina = { str(row.get("oficina_id") or row.get("oficina") or "—"): int(row["c"]) for row in qs.values("oficina_id").annotate(c=Count("id")).order_by("-c") }
        por_compania = { str(row["compania"] or "—"): int(row["c"]) for row in qs.values("compania").annotate(c=Count("id")).order_by("-c") }
        
        por_cobertura = {}
        if hasattr(Poliza, "cobertura"):
            por_cobertura = { str(row["cobertura"] or "—"): int(row["c"]) for row in qs.values("cobertura").annotate(c=Count("id")).order_by("-c") }

        payload = {
            "filtros": {
                "oficina": oficina_filtro_raw or None, "tipo": str(params.get("tipo") or params.get("tipo_vehiculo") or "").strip() or None,
                "anio": str(params.get("anio") or "").strip() or None, "anio_desde": str(params.get("anio_desde") or "").strip() or None,
                "anio_hasta": str(params.get("anio_hasta") or "").strip() or None, "marca": str(params.get("marca") or "").strip() or None,
                "modelo": str(params.get("modelo") or "").strip() or None, "patente": str(params.get("patente") or "").strip() or None,
                "solo_activas": _to_bool(params.get("solo_activas")),
            },
            "total_polizas": int(total), "total_activas": int(total_activas), "por_tipo": por_tipo,
            "por_anio": por_anio, "por_oficina": por_oficina, "por_compania": por_compania, "por_cobertura": por_cobertura,
        }
        return Response(payload)


class VehiculosListAPIView(APIView):
    """
    GET /api/estadisticas/vehiculos/list/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        params = request.query_params
        
        oficina_filtro_raw = str(params.get("oficina") or "").strip()
        oficina_segura_keys = _get_seguridad_oficina(request, oficina_filtro_raw)
        
        if "BLOQUEADO" in oficina_segura_keys: return Response({"error": "Acceso denegado"}, status=403)

        hoy = timezone.localdate()
        page = _clamp_int(params.get("page"), 1, min_v=1)
        page_size = _clamp_int(params.get("page_size"), 25, min_v=1, max_v=200)

        q_txt = str(params.get("q") or "").strip()
        orden = str(params.get("orden") or "id").strip()
        direction = str(params.get("dir") or "desc").strip().lower()
        desc = direction == "desc"

        qs = Poliza.objects.select_related("cliente").all()
        qs = _apply_vehiculos_filters(qs, params, oficina_segura_keys=oficina_segura_keys)

        # Filtros de estado financiero — solo admins
        user = request.user
        es_admin = user.is_superuser or (hasattr(user, 'perfil') and user.perfil.rol == 'ADMIN')

        solo_al_dia   = _to_bool(params.get("solo_al_dia"))
        solo_en_mora  = _to_bool(params.get("solo_en_mora"))

        if (solo_al_dia or solo_en_mora) and not es_admin:
            return Response({"error": "Acceso denegado"}, status=403)

        if solo_al_dia or solo_en_mora:
            ultima_vto = Cuota.objects.filter(
                poliza=OuterRef("pk")
            ).order_by("-fecha_vencimiento").values("fecha_vencimiento")[:1]
            ultima_pagada = Cuota.objects.filter(
                poliza=OuterRef("pk")
            ).order_by("-fecha_vencimiento").values("pagado")[:1]
            qs = qs.filter(estado__iexact="activa").annotate(
                ultima_vto=Subquery(ultima_vto),
                ultima_pagada=Subquery(ultima_pagada),
            )
            if solo_al_dia:
                qs = qs.filter(
                    Q(ultima_pagada=True) | Q(ultima_vto__gte=hoy) | Q(ultima_vto__isnull=True)
                )
            elif solo_en_mora:
                qs = qs.filter(ultima_pagada=False, ultima_vto__lt=hoy)

        if q_txt:
            qs = qs.filter(
                Q(patente__icontains=q_txt) | Q(marca__icontains=q_txt) | Q(modelo__icontains=q_txt) | Q(numero_poliza__icontains=q_txt) |
                Q(cliente__apellido__icontains=q_txt) | Q(cliente__nombre__icontains=q_txt) | Q(cliente__dni_cuit_cuil__icontains=q_txt)
            )

        order_map = {
            "id": "id", "numero_poliza": "numero_poliza", "estado": "estado", "oficina": "oficina",
            "tipo": "tipo", "anio": "anio", "marca": "marca", "modelo": "modelo", "patente": "patente",
            "asegurado": "cliente__apellido", "dni": "cliente__dni_cuit_cuil",
        }

        order_field = order_map.get(orden, "id")
        if orden == "asegurado": ordering = [("-" if desc else "") + "cliente__apellido", ("-" if desc else "") + "cliente__nombre"]
        else: ordering = [("-" if desc else "") + order_field]

        qs = qs.order_by(*ordering)

        total = qs.count()
        total_pages = int(math.ceil(total / float(page_size))) if total else 1
        if page > total_pages: page = total_pages

        offset = (page - 1) * page_size
        items = list(qs[offset : offset + page_size])

        def _asegurado(p):
            c = getattr(p, "cliente", None)
            if not c: return ""
            return f"{getattr(c, 'apellido', '')} {getattr(c, 'nombre', '')}".strip()

        results = []
        for p in items:
            c = getattr(p, "cliente", None)
            oficina_raw = getattr(p, "oficina_id", None) or getattr(p, "oficina", "")
            oficina_bucket = _bucket_oficina(oficina_raw)
            results.append({
                "poliza_id": getattr(p, "id", None), "numero_poliza": str(getattr(p, "numero_poliza", "") or ""),
                "estado": str(getattr(p, "estado", "") or ""), "oficina": str(oficina_raw),
                "oficina_bucket": oficina_bucket, "oficina_nombre": oficina_bucket,
                "tipo": str(getattr(p, "tipo", "") or ""), "anio": getattr(p, "anio", None),
                "marca": str(getattr(p, "marca", "") or ""), "modelo": str(getattr(p, "modelo", "") or ""),
                "patente": str(getattr(p, "patente", "") or ""), "cliente_id": getattr(c, "id", None) if c else None,
                "asegurado": _asegurado(p), "dni_cuit_cuil": str(getattr(c, "dni_cuit_cuil", "") if c else ""),
            })

        return Response({ "page": page, "page_size": page_size, "count": int(total), "total_pages": int(total_pages), "results": results })


class VehiculosExportAPIView(APIView):
    """
    GET /api/estadisticas/vehiculos/export/

    Exporta Excel/CSV completo según la tarjeta seleccionada.

    Importante:
    - NO usa paginación.
    - La tabla/listado de pantalla puede seguir usando page/page_size.
    - anio y mes son el período del reporte, NO el año del vehículo.
    - Si se quiere filtrar por año del vehículo, usar anio_vehiculo.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        params = request.query_params

        tipo_listado = str(params.get("tipo_listado") or "TOTALES").strip().upper()
        formato = str(params.get("formato") or "csv").strip().lower()

        tipos_validos = {"TOTALES", "ACTIVAS", "ACTIVAS_AL_DIA", "ACTIVAS_EN_MORA", "ALTAS", "VENCIDAS", "BAJAS"}
        if tipo_listado not in tipos_validos:
            tipo_listado = "TOTALES"

        # ACTIVAS_AL_DIA y ACTIVAS_EN_MORA solo para admins
        user = request.user
        es_admin = user.is_superuser or (hasattr(user, 'perfil') and user.perfil.rol == 'ADMIN')
        if tipo_listado in ("ACTIVAS_AL_DIA", "ACTIVAS_EN_MORA") and not es_admin:
            return HttpResponse("Acceso denegado", status=403)

        oficina_filtro_raw = str(params.get("oficina") or "").strip()
        oficina_segura_keys = _get_seguridad_oficina(request, oficina_filtro_raw)

        if "BLOQUEADO" in oficina_segura_keys:
            return HttpResponse("Acceso denegado", status=403)

        hoy = timezone.localdate()

        try:
            anio_periodo = int(params.get("anio") or params.get("periodo_anio") or hoy.year)
        except (TypeError, ValueError):
            anio_periodo = hoy.year

        try:
            mes_periodo = int(params.get("mes") or params.get("periodo_mes") or hoy.month)
            if mes_periodo < 1 or mes_periodo > 12:
                mes_periodo = hoy.month
        except (TypeError, ValueError):
            mes_periodo = hoy.month

        primero_mes = date(anio_periodo, mes_periodo, 1)
        ultimo_dia = monthrange(anio_periodo, mes_periodo)[1]
        ultimo_mes = date(anio_periodo, mes_periodo, ultimo_dia)

        dt_primero = timezone.make_aware(datetime.combine(primero_mes, time.min))
        dt_ultimo = timezone.make_aware(datetime.combine(ultimo_mes, time.max))

        # Consulta base completa: SIN page, SIN page_size.
        try:
            if _is_poliza_oficina_fk():
                qs = Poliza.objects.select_related("cliente", "oficina").all()
            else:
                qs = Poliza.objects.select_related("cliente").all()
        except Exception:
            qs = Poliza.objects.select_related("cliente").all()

        # Filtro propio de la tarjeta.
        if tipo_listado == "ACTIVAS":
            qs = qs.filter(estado__iexact="activa")

        elif tipo_listado == "ACTIVAS_AL_DIA":
            # Activas cuya ÚLTIMA cuota está pagada o aún no venció
            ultima_vto = Cuota.objects.filter(
                poliza=OuterRef("pk")
            ).order_by("-fecha_vencimiento").values("fecha_vencimiento")[:1]
            ultima_pagada = Cuota.objects.filter(
                poliza=OuterRef("pk")
            ).order_by("-fecha_vencimiento").values("pagado")[:1]
            qs = qs.filter(estado__iexact="activa").annotate(
                ultima_vto=Subquery(ultima_vto),
                ultima_pagada=Subquery(ultima_pagada),
            ).filter(
                Q(ultima_pagada=True) | Q(ultima_vto__gte=hoy) | Q(ultima_vto__isnull=True)
            )

        elif tipo_listado == "ACTIVAS_EN_MORA":
            # Activas cuya ÚLTIMA cuota venció y no está pagada
            ultima_vto = Cuota.objects.filter(
                poliza=OuterRef("pk")
            ).order_by("-fecha_vencimiento").values("fecha_vencimiento")[:1]
            ultima_pagada = Cuota.objects.filter(
                poliza=OuterRef("pk")
            ).order_by("-fecha_vencimiento").values("pagado")[:1]
            qs = qs.filter(estado__iexact="activa").annotate(
                ultima_vto=Subquery(ultima_vto),
                ultima_pagada=Subquery(ultima_pagada),
            ).filter(
                ultima_pagada=False,
                ultima_vto__lt=hoy,
            )

        elif tipo_listado == "ALTAS":
            qs = qs.filter(fecha_emision__gte=primero_mes, fecha_emision__lte=ultimo_mes)

        elif tipo_listado == "VENCIDAS":
            qs = qs.filter(
                Q(cuotas__pagado=False, cuotas__fecha_vencimiento__lt=hoy) |
                Q(estado__iexact="vencida") |
                Q(estado__iexact="VENCIDA")
            ).distinct()

        elif tipo_listado == "BAJAS":
            qs = qs.filter(
                Q(fecha_baja__isnull=False, fecha_baja__gte=primero_mes, fecha_baja__lte=ultimo_mes) |
                Q(baja_operativa__realizada_en__gte=dt_primero, baja_operativa__realizada_en__lte=dt_ultimo)
            ).distinct()

        # MUY IMPORTANTE:
        # _apply_vehiculos_filters() usa params["anio"] como año del vehículo.
        # En Estadísticas, anio/mes son período del reporte. Si no los quitamos,
        # el Excel queda vacío cuando no hay vehículos modelo 2026, 2025, etc.
        vehiculos_params = params.copy()
        for key in ("anio", "año", "mes", "page", "page_size", "limit", "offset", "formato", "tipo_listado", "export_all"):
            try:
                vehiculos_params.pop(key, None)
            except Exception:
                pass

        # Compatibilidad opcional: si algún filtro realmente quiere año del vehículo,
        # debe mandarlo como anio_vehiculo.
        anio_vehiculo = str(params.get("anio_vehiculo") or "").strip()
        if anio_vehiculo:
            vehiculos_params["anio"] = anio_vehiculo

        qs = _apply_vehiculos_filters(qs, vehiculos_params, oficina_segura_keys=oficina_segura_keys)

        # Búsqueda opcional para que el export respete q si algún modal/listado lo envía.
        q_txt = str(params.get("q") or "").strip()
        if q_txt:
            qs = qs.filter(
                Q(patente__icontains=q_txt) |
                Q(marca__icontains=q_txt) |
                Q(modelo__icontains=q_txt) |
                Q(numero_poliza__icontains=q_txt) |
                Q(compania__icontains=q_txt) |
                Q(cliente__apellido__icontains=q_txt) |
                Q(cliente__nombre__icontains=q_txt) |
                Q(cliente__dni_cuit_cuil__icontains=q_txt) |
                Q(cliente__telefono__icontains=q_txt) |
                Q(cliente__email__icontains=q_txt)
            )

        try:
            qs = qs.order_by("oficina", "cliente__apellido", "cliente__nombre", "patente", "id")
        except Exception:
            qs = qs.order_by("cliente__apellido", "cliente__nombre", "patente", "id")

        def _has_poliza_field(field_name):
            try:
                Poliza._meta.get_field(field_name)
                return True
            except Exception:
                return False

        def _fmt_date(v):
            if not v:
                return ""
            try:
                if hasattr(v, "date"):
                    v = v.date()
                return v.isoformat()
            except Exception:
                return str(v)

        def _fmt_decimal(v):
            if v is None:
                return ""
            try:
                return float(v)
            except Exception:
                return str(v)

        def _asegurado(poliza):
            c = getattr(poliza, "cliente", None)
            if not c:
                return ""
            return f"{getattr(c, 'apellido', '')} {getattr(c, 'nombre', '')}".strip()

        def _oficina_valor(poliza):
            raw_id = getattr(poliza, "oficina_id", None)
            raw = getattr(poliza, "oficina", "")
            if raw_id:
                return str(raw_id)
            if hasattr(raw, "id") and raw.id is not None:
                return str(raw.id)
            return str(raw or "")

        def _oficina_nombre(poliza):
            raw = getattr(poliza, "oficina", "")
            if hasattr(raw, "nombre") and raw.nombre:
                return str(raw.nombre)
            if hasattr(raw, "codigo") and raw.codigo:
                return str(raw.codigo)
            return _bucket_oficina(_oficina_valor(poliza))

        tiene_cobertura = _has_poliza_field("cobertura")
        tiene_fecha_emision = _has_poliza_field("fecha_emision")
        tiene_fecha_vencimiento = _has_poliza_field("fecha_vencimiento")
        tiene_fecha_baja = _has_poliza_field("fecha_baja")
        tiene_precio_cuota = _has_poliza_field("precio_cuota")
        tiene_cantidad_cuotas = _has_poliza_field("cantidad_cuotas")
        tiene_observaciones_baja = _has_poliza_field("observaciones_baja")

        headers = [
            ("poliza_id", "ID póliza"),
            ("numero_poliza", "N° póliza"),
            ("estado", "Estado"),
            ("oficina", "Oficina"),
            ("oficina_nombre", "Nombre oficina"),
            ("compania", "Compañía"),
            ("cobertura", "Cobertura"),
            ("tipo", "Tipo"),
            ("patente", "Patente"),
            ("marca", "Marca"),
            ("modelo", "Modelo"),
            ("anio_vehiculo", "Año vehículo"),
            ("asegurado", "Asegurado"),
            ("dni_cuit_cuil", "DNI / CUIT / CUIL"),
            ("telefono", "Teléfono"),
            ("email", "Email"),
            ("fecha_emision", "Fecha emisión"),
            ("fecha_vencimiento", "Fecha vencimiento"),
            ("fecha_baja", "Fecha baja"),
            ("precio_cuota", "Precio cuota"),
            ("cantidad_cuotas", "Cantidad cuotas"),
            ("observaciones_baja", "Observaciones baja"),
        ]

        rows = []
        for poliza in qs.iterator(chunk_size=1000):
            cliente = getattr(poliza, "cliente", None)
            rows.append({
                "poliza_id": getattr(poliza, "id", ""),
                "numero_poliza": str(getattr(poliza, "numero_poliza", "") or ""),
                "estado": str(getattr(poliza, "estado", "") or ""),
                "oficina": _oficina_valor(poliza),
                "oficina_nombre": _oficina_nombre(poliza),
                "compania": str(getattr(poliza, "compania", "") or ""),
                "cobertura": str(getattr(poliza, "cobertura", "") or "") if tiene_cobertura else "",
                "tipo": str(getattr(poliza, "tipo", "") or ""),
                "patente": str(getattr(poliza, "patente", "") or ""),
                "marca": str(getattr(poliza, "marca", "") or ""),
                "modelo": str(getattr(poliza, "modelo", "") or ""),
                "anio_vehiculo": getattr(poliza, "anio", "") or "",
                "asegurado": _asegurado(poliza),
                "dni_cuit_cuil": str(getattr(cliente, "dni_cuit_cuil", "") if cliente else ""),
                "telefono": str(getattr(cliente, "telefono", "") if cliente else ""),
                "email": str(getattr(cliente, "email", "") if cliente else ""),
                "fecha_emision": _fmt_date(getattr(poliza, "fecha_emision", None)) if tiene_fecha_emision else "",
                "fecha_vencimiento": _fmt_date(getattr(poliza, "fecha_vencimiento", None)) if tiene_fecha_vencimiento else "",
                "fecha_baja": _fmt_date(getattr(poliza, "fecha_baja", None)) if tiene_fecha_baja else "",
                "precio_cuota": _fmt_decimal(getattr(poliza, "precio_cuota", None)) if tiene_precio_cuota else "",
                "cantidad_cuotas": getattr(poliza, "cantidad_cuotas", "") if tiene_cantidad_cuotas else "",
                "observaciones_baja": str(getattr(poliza, "observaciones_baja", "") or "") if tiene_observaciones_baja else "",
            })

        total_exportado = len(rows)
        filename_base = f"reporte_{tipo_listado.lower()}_{anio_periodo}_{mes_periodo:02d}"

        try:
            usuario = getattr(request, "user", None)
            ExportLog.objects.create(
                usuario=usuario if getattr(usuario, "is_authenticated", False) else None,
                tipo=f"vehiculos_export_{tipo_listado.lower()}",
                parametros={
                    "formato": formato,
                    "tipo_listado": tipo_listado,
                    "anio": anio_periodo,
                    "mes": mes_periodo,
                    "oficina": oficina_filtro_raw or None,
                    "total_exportado": total_exportado,
                    "sin_paginacion": True,
                    "nota": "anio/mes usados como período; anio_vehiculo reservado para año del vehículo",
                },
            )
        except Exception:
            pass

        if formato == "xlsx":
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
            except Exception:
                formato = "csv"

        if formato == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = f"{tipo_listado.title()}"

            titulo = f"Reporte de asegurados - {tipo_listado.title()}"
            subtitulo = f"Período: {anio_periodo:04d}-{mes_periodo:02d} | Oficina: {oficina_filtro_raw or 'Todas'} | Total exportado: {total_exportado}"

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            ws.cell(row=1, column=1, value=titulo)
            ws.cell(row=1, column=1).font = Font(bold=True, size=16)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            ws.cell(row=2, column=1, value=subtitulo)
            ws.cell(row=2, column=1).font = Font(size=11)
            ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

            ws.append([])
            ws.append([label for _, label in headers])

            header_row = 4
            header_fill = PatternFill("solid", fgColor="1F2937")
            header_font = Font(bold=True, color="FFFFFF")
            thin = Side(style="thin", color="D1D5DB")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for cell in ws[header_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            if rows:
                for row in rows:
                    ws.append([row.get(key, "") for key, _ in headers])
            else:
                ws.append(["Sin resultados para los filtros seleccionados"] + [""] * (len(headers) - 1))

            for row_cells in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row_cells:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            precio_idx = [i for i, (key, _) in enumerate(headers, start=1) if key == "precio_cuota"]
            if precio_idx:
                col_idx = precio_idx[0]
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=5, max_row=ws.max_row):
                    for c in cell:
                        if isinstance(c.value, (int, float)):
                            c.number_format = '$ #,##0.00'

            ws.freeze_panes = "A5"
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"

            for col_idx, (key, label) in enumerate(headers, start=1):
                letter = get_column_letter(col_idx)
                max_len = len(label)
                for cell in ws[letter]:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(value), 45))
                ws.column_dimensions[letter].width = max(12, min(max_len + 2, 48))

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)

            resp = HttpResponse(
                bio.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
            return resp

        resp = HttpResponse(content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'

        writer = csv.writer(resp)
        writer.writerow([label for _, label in headers])
        for row in rows:
            writer.writerow([row.get(key, "") for key, _ in headers])

        return resp

# -------------------------
# Duplicados: helpers
# -------------------------
def _only_digits(v: str) -> str:
    try: import re; return re.sub(r"\D", "", str(v or ""))
    except Exception: return "".join(ch for ch in str(v or "") if ch.isdigit())

def _norm_dni(v: str) -> str:
    s = str(v or "").strip()
    if not s: return ""
    try: import re; s = re.sub(r"[^0-9A-Za-z]", "", s)
    except Exception: s = "".join(ch for ch in s if ch.isalnum())
    return s.upper()

def _norm_email(v: str) -> str: return str(v or "").strip().lower()

def _norm_tel(v: str) -> str:
    d = _only_digits(v)
    if not d: return ""
    if d.startswith("0") and len(d) > 8: d = d[1:]
    if d.startswith("15") and len(d) >= 10: d = d[2:]
    return d

def _clamp_groups(v, default=200, min_v=1, max_v=2000): return _clamp_int(v, default, min_v=min_v, max_v=max_v)
def _clamp_items(v, default=20, min_v=1, max_v=200): return _clamp_int(v, default, min_v=min_v, max_v=max_v)


# -------------------------
# Duplicados: clientes
# -------------------------
class ClientesDuplicadosAPIView(APIView):
    """
    GET /api/estadisticas/duplicados/clientes/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        p = request.query_params

        oficina_filtro_raw = str(p.get("oficina") or "").strip()
        oficina_segura_keys = _get_seguridad_oficina(request, oficina_filtro_raw)
        
        if "BLOQUEADO" in oficina_segura_keys: return Response({"error": "Acceso denegado"}, status=403)

        modos_raw = str(p.get("modos") or "dni,telefono,email").strip().lower()
        modos = [m.strip() for m in modos_raw.split(",") if m.strip()]
        allowed = {"dni", "telefono", "email"}
        modos = [m for m in modos if m in allowed] or ["dni", "telefono", "email"]

        max_groups = _clamp_groups(p.get("max_groups"), default=200)
        max_items = _clamp_items(p.get("max_items"), default=20)

        solo_completo = str(p.get("solo_completo") or "").strip().lower() in { "1", "true", "t", "yes", "y", "si", "sí" }

        qs = Cliente.objects.all().only("id", "nombre", "apellido", "telefono", "email", "dni_cuit_cuil", "estado")
        if solo_completo and hasattr(Cliente, "estado"): qs = qs.filter(estado="COMPLETO")

        if oficina_segura_keys:
            pol_qs = _apply_oficina_filter(Poliza.objects.all(), oficina_segura_keys, is_poliza_model=True)
            try: pol_qs = pol_qs.exclude(cliente_id__isnull=True)
            except Exception: pass

            cliente_ids = list(pol_qs.values_list("cliente_id", flat=True).distinct())
            if cliente_ids: qs = qs.filter(id__in=cliente_ids)
            else: qs = qs.none()

        items = list(qs.values("id", "nombre", "apellido", "telefono", "email", "dni_cuit_cuil", "estado"))

        groups = []
        def add_group(modo: str, key: str, clientes: list):
            if not key or len(clientes) <= 1: return
            groups.append({ "modo": modo, "key": key, "count": len(clientes), "clientes": clientes[:max_items], "truncated": len(clientes) > max_items })

        for modo in modos:
            mp = {}
            for c in items:
                if modo == "dni": k = _norm_dni(c.get("dni_cuit_cuil"))
                elif modo == "telefono": k = _norm_tel(c.get("telefono"))
                else: k = _norm_email(c.get("email"))
                if not k: continue
                mp.setdefault(k, []).append(c)

            for k, arr in mp.items():
                if len(arr) > 1:
                    arr_sorted = sorted(arr, key=lambda x: (str(x.get("apellido") or "").lower(), str(x.get("nombre") or "").lower(), int(x.get("id") or 0)))
                    add_group(modo, k, arr_sorted)

        groups.sort(key=lambda g: (-int(g.get("count") or 0), str(g.get("modo") or ""), str(g.get("key") or "")))

        total_groups = len(groups)
        groups = groups[:max_groups]
        truncated_groups = total_groups > len(groups)

        total_clientes_in_groups = sum(int(g.get("count") or 0) for g in groups)

        payload = {
            "modos": modos, "solo_completo": bool(solo_completo), "oficina": oficina_filtro_raw or None,
            "total_grupos": int(total_groups), "total_grupos_devuelto": int(len(groups)),
            "grupos_truncados": bool(truncated_groups), "max_groups": int(max_groups), "max_items": int(max_items),
            "total_clientes_en_grupos_devueltos": int(total_clientes_in_groups), "grupos": groups,
        }
        return Response(payload)


# -------------------------
# Duplicados: pólizas (por patente)
# -------------------------
class PolizasDuplicadasAPIView(APIView):
    """
    GET /api/estadisticas/duplicados/polizas/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        p = request.query_params

        oficina_filtro_raw = str(p.get("oficina") or "").strip()
        oficina_segura_keys = _get_seguridad_oficina(request, oficina_filtro_raw)
        
        if "BLOQUEADO" in oficina_segura_keys: return Response({"error": "Acceso denegado"}, status=403)

        modo = str(p.get("modo") or "patente").strip().lower()
        if modo != "patente": modo = "patente"

        solo_activas = str(p.get("solo_activas") or "1").strip().lower() in { "1", "true", "t", "yes", "y", "si", "sí" }

        max_groups = _clamp_groups(p.get("max_groups"), default=200)
        max_items = _clamp_items(p.get("max_items"), default=20)

        qs = Poliza.objects.select_related("cliente").all()
        if solo_activas: qs = qs.filter(estado="activa")
        if oficina_segura_keys: qs = _apply_oficina_filter(qs, oficina_segura_keys, is_poliza_model=True)

        qs = qs.exclude(Q(patente__isnull=True) | Q(patente__exact=""))

        items = list(qs.values("id", "numero_poliza", "estado", "compania", "oficina", "patente", "marca", "modelo", "anio", "cliente_id", "cliente__nombre", "cliente__apellido", "cliente__dni_cuit_cuil"))

        mp = {}
        for r in items:
            pat = str(r.get("patente") or "").strip().replace(" ", "").upper()
            if not pat: continue
            mp.setdefault(pat, []).append(r)

        groups = []
        for pat, arr in mp.items():
            if len(arr) <= 1: continue

            arr_sorted = sorted(arr, key=lambda x: (str(x.get("estado") or "").lower(), str(x.get("compania") or "").lower(), str(x.get("numero_poliza") or ""), int(x.get("id") or 0)))

            polizas_payload = []
            for it in arr_sorted[:max_items]:
                polizas_payload.append({
                    "poliza_id": it.get("id"), "numero_poliza": str(it.get("numero_poliza") or ""),
                    "estado": str(it.get("estado") or ""), "compania": str(it.get("compania") or ""),
                    "oficina": str(it.get("oficina") or ""), "patente": str(it.get("patente") or ""),
                    "marca": str(it.get("marca") or ""), "modelo": str(it.get("modelo") or ""),
                    "anio": it.get("anio"), "cliente_id": it.get("cliente_id"),
                    "asegurado": f"{str(it.get('cliente__apellido') or '').strip()} {str(it.get('cliente__nombre') or '').strip()}".strip(),
                    "dni_cuit_cuil": str(it.get("cliente__dni_cuit_cuil") or ""),
                })

            groups.append({ "modo": "patente", "key": pat, "count": len(arr_sorted), "polizas": polizas_payload, "truncated": len(arr_sorted) > max_items })

        groups.sort(key=lambda g: (-int(g.get("count") or 0), str(g.get("key") or "")))

        total_groups = len(groups)
        groups = groups[:max_groups]
        truncated_groups = total_groups > len(groups)

        total_polizas_in_groups = sum(int(g.get("count") or 0) for g in groups)

        payload = {
            "modo": modo, "solo_activas": bool(solo_activas), "oficina": oficina_filtro_raw or None,
            "total_grupos": int(total_groups), "total_grupos_devuelto": int(len(groups)),
            "grupos_truncados": bool(truncated_groups), "max_groups": int(max_groups), "max_items": int(max_items),
            "total_polizas_en_grupos_devueltos": int(total_polizas_in_groups), "grupos": groups,
        }
        return Response(payload)


# -------------------------
# ✅ Contabilidad y Caja (Pagos y Cuotas)
# -------------------------


# -------------------------
# Bajas del mes + Tasa de retención
# -------------------------
class BajasRetenciónAPIView(APIView):
    """
    GET /api/estadisticas/polizas/bajas-retencion/

    Devuelve por cada mes del rango:
      - bajas_mes: pólizas dadas de baja ese mes
      - altas_mes: pólizas emitidas ese mes (sin renovaciones si es_renovacion existe)
      - activas_inicio: stock activo al inicio del mes
      - activas_fin: stock activo al fin del mes
      - retencion_pct: (activas_inicio - bajas_mes) / activas_inicio * 100

    Query params:
      - oficina: id de oficina (vacío = todas)
      - desde: YYYY-MM-DD (default: hace 12 meses)
      - hasta: YYYY-MM-DD (default: hoy)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        params = request.query_params

        oficina_filtro_raw = str(params.get("oficina") or "").strip()
        oficina_segura_keys = _get_seguridad_oficina(request, oficina_filtro_raw)
        if "BLOQUEADO" in oficina_segura_keys:
            return Response({"error": "Acceso denegado"}, status=403)

        hoy = timezone.localdate()
        desde = _parse_date_iso(params.get("desde")) or (hoy.replace(day=1) - timedelta(days=365))
        hasta = _parse_date_iso(params.get("hasta")) or hoy

        if desde > hasta:
            desde, hasta = hasta, desde

        qs = Poliza.objects.all()
        if oficina_segura_keys:
            qs = _apply_oficina_filter(qs, oficina_segura_keys, is_poliza_model=True)

        tiene_es_renovacion = hasattr(Poliza, "es_renovacion")

        # Iteramos mes a mes dentro del rango
        meses = []
        cur = date(desde.year, desde.month, 1)
        fin_rango = date(hasta.year, hasta.month, 1)

        while cur <= fin_rango:
            ultimo_dia = monthrange(cur.year, cur.month)[1]
            fin_mes = date(cur.year, cur.month, ultimo_dia)
            inicio_mes = cur

            # Bajas del mes — pólizas canceladas/dadas de baja en este mes
            bajas_mes = qs.filter(
                Q(fecha_baja__gte=inicio_mes, fecha_baja__lte=fin_mes) |
                Q(baja_operativa__realizada_en__date__gte=inicio_mes,
                  baja_operativa__realizada_en__date__lte=fin_mes)
            ).distinct().count()

            # Altas del mes — emitidas en este mes
            altas_qs = qs.filter(fecha_emision__gte=inicio_mes, fecha_emision__lte=fin_mes)
            altas_mes = altas_qs.count()

            # Altas nuevas (sin renovaciones)
            if tiene_es_renovacion:
                altas_nuevas_mes = altas_qs.filter(es_renovacion=False).count()
                renovaciones_mes = altas_qs.filter(es_renovacion=True).count()
            else:
                altas_nuevas_mes = altas_mes
                renovaciones_mes = 0

            # Stock activo al inicio del mes (pólizas emitidas antes del mes y no canceladas antes)
            activas_inicio = qs.filter(
                fecha_emision__lt=inicio_mes,
            ).exclude(
                Q(fecha_baja__lt=inicio_mes) |
                Q(estado__in=["cancelada", "finalizada"])
            ).count()

            # Stock activo al fin del mes
            activas_fin = qs.filter(estado__in=["activa", "vencida", "en_verificacion"]).count() \
                if cur.month == hoy.month and cur.year == hoy.year \
                else qs.filter(
                    fecha_emision__lte=fin_mes,
                ).exclude(
                    Q(fecha_baja__lte=fin_mes) |
                    Q(estado__in=["cancelada", "finalizada"])
                ).count()

            # Tasa de retención
            if activas_inicio > 0:
                retencion_pct = round(((activas_inicio - bajas_mes) / activas_inicio) * 100, 1)
            else:
                retencion_pct = 100.0

            meses.append({
                "periodo":          cur.strftime("%Y-%m"),
                "bajas_mes":        bajas_mes,
                "altas_mes":        altas_mes,
                "altas_nuevas_mes": altas_nuevas_mes,
                "renovaciones_mes": renovaciones_mes,
                "activas_inicio":   activas_inicio,
                "activas_fin":      activas_fin,
                "retencion_pct":    retencion_pct,
            })

            # Avanzar al próximo mes
            if cur.month == 12:
                cur = date(cur.year + 1, 1, 1)
            else:
                cur = date(cur.year, cur.month + 1, 1)

        return Response({
            "desde":   desde.isoformat(),
            "hasta":   hasta.isoformat(),
            "oficina": oficina_filtro_raw or None,
            "meses":   meses,
        })


# -------------------------
# Contabilidad y Caja (Pagos y Cuotas)
# -------------------------
class ContabilidadResumenAPIView(APIView):
    """
    GET /api/estadisticas/contabilidad/resumen/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        auto_marcar_vencidas()

        params = request.query_params
        
        # 🚀 APLICAMOS ESCUDO MULTI-TENANT
        oficina_filtro_raw = str(params.get("oficina") or "").strip()
        oficina_segura_keys = _get_seguridad_oficina(request, oficina_filtro_raw)
        
        if "BLOQUEADO" in oficina_segura_keys: return Response({"error": "Acceso denegado"}, status=403)

        hoy = timezone.localdate()
        try: anio = int(params.get("anio") or hoy.year)
        except ValueError: anio = hoy.year
            
        try: mes = int(params.get("mes") or hoy.month)
        except ValueError: mes = hoy.month

        pagos_qs = Pago.objects.filter(fecha__year=anio, fecha__month=mes)
        cuotas_mes_qs = Cuota.objects.filter(fecha_vencimiento__year=anio, fecha_vencimiento__month=mes)
        deuda_qs = Cuota.objects.filter(pagado=False, fecha_vencimiento__lt=hoy)

        if oficina_segura_keys:
            pagos_qs = pagos_qs.filter(_build_oficina_q_from_keys(oficina_segura_keys, prefix="poliza__"))
            cuotas_mes_qs = cuotas_mes_qs.filter(_build_oficina_q_from_keys(oficina_segura_keys, prefix="poliza__"))
            deuda_qs = deuda_qs.filter(_build_oficina_q_from_keys(oficina_segura_keys, prefix="poliza__"))

        recaudacion_total = pagos_qs.aggregate(t=Sum('monto'))['t'] or 0
        recaudacion_efectivo = pagos_qs.filter(metodo='efectivo').aggregate(t=Sum('monto'))['t'] or 0
        recaudacion_transf = pagos_qs.filter(metodo='transferencia').aggregate(t=Sum('monto'))['t'] or 0

        esperado_mes = cuotas_mes_qs.aggregate(t=Sum('monto'))['t'] or 0
        pendiente_mes = cuotas_mes_qs.filter(pagado=False).aggregate(t=Sum('monto'))['t'] or 0

        deuda_total = deuda_qs.aggregate(t=Sum('monto'))['t'] or 0

        return Response({
            "periodo": f"{anio:04d}-{mes:02d}", "oficina": oficina_filtro_raw or "Todas",
            "recaudacion": { "total": float(recaudacion_total), "efectivo": float(recaudacion_efectivo), "transferencia": float(recaudacion_transf) },
            "mes_actual": { "esperado": float(esperado_mes), "pendiente": float(pendiente_mes), "cobrado": float(esperado_mes) - float(pendiente_mes) },
            "morosidad_historica": float(deuda_total)
        })

# ════════════════════════════════════════════════════════════════════
# 🆕 EXPORT EXCEL DE EMISIONES (Altas + Renovaciones) — Reporte detallado
# ════════════════════════════════════════════════════════════════════
class EmisionesExportExcelAPIView(APIView):
    """
    GET /api/estadisticas/polizas/emisiones/export-excel/

    Exporta a Excel TODAS las pólizas emitidas en el rango con todos sus datos
    para análisis posterior en planilla.

    Query params:
      - desde=YYYY-MM-DD       (default: hace 30 días)
      - hasta=YYYY-MM-DD       (default: hoy)
      - oficina=<id|nombre>    (opcional; admin puede filtrar, no-admin solo su oficina)
      - tipo=nuevas|renovaciones|todas  (default: todas)
      - compania=<nombre>      (opcional)

    Hojas del Excel:
      1) "Pólizas"  → fila por póliza con 23 columnas
      2) "Resumen"  → totales, desglose por compañía y por oficina
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        # Lazy import por si xlsxwriter no está instalado
        try:
            import xlsxwriter
        except ImportError:
            return Response(
                {"detail": "El módulo xlsxwriter no está instalado en el servidor."},
                status=500,
            )

        params = request.query_params

        # ── Parse fechas ────────────────────────────────────────
        def _parse_date(v):
            try:
                return date.fromisoformat((v or "").strip())
            except Exception:
                return None

        hasta = _parse_date(params.get("hasta")) or timezone.localdate()
        desde = _parse_date(params.get("desde")) or (hasta - timedelta(days=30))
        if desde > hasta:
            desde, hasta = hasta, desde

        # ── Escudo de sucursal ──────────────────────────────────
        user = request.user
        perfil = getattr(user, "perfil", None)
        is_admin = bool(getattr(user, "is_superuser", False) or getattr(perfil, "rol", "") == "ADMIN")

        # ── Queryset base ───────────────────────────────────────
        try:
            fecha_field = Poliza._meta.get_field("fecha_emision")
            is_datetime = fecha_field.get_internal_type() == "DateTimeField"
        except Exception:
            is_datetime = False

        if is_datetime:
            dt_from = timezone.make_aware(datetime.combine(desde, time.min))
            dt_to_excl = timezone.make_aware(datetime.combine(hasta + timedelta(days=1), time.min))
            qs = Poliza.objects.filter(
                fecha_emision__gte=dt_from,
                fecha_emision__lt=dt_to_excl,
            )
        else:
            qs = Poliza.objects.filter(
                fecha_emision__gte=desde,
                fecha_emision__lte=hasta,
            )

        # No-admin: sólo su oficina
        if not is_admin:
            ofi_id = getattr(perfil, "oficina_id", None)
            if ofi_id:
                qs = qs.filter(oficina_id=ofi_id)
        else:
            # Admin puede filtrar por oficina (id o nombre)
            ofi_raw = str(params.get("oficina") or "").strip()
            if ofi_raw:
                if ofi_raw.isdigit():
                    qs = qs.filter(oficina_id=int(ofi_raw))
                else:
                    qs = qs.filter(oficina__nombre__iexact=ofi_raw)

        # Filtro de compañía
        compania = str(params.get("compania") or "").strip()
        if compania:
            qs = qs.filter(
                Q(compania__iexact=compania)
                | Q(compania_obj__nombre__iexact=compania)
            )

        # Filtro de tipo (nuevas / renovaciones / todas)
        tipo_param = str(params.get("tipo") or "todas").strip().lower()
        if hasattr(Poliza, "es_renovacion"):
            if tipo_param == "nuevas":
                qs = qs.filter(es_renovacion=False)
            elif tipo_param == "renovaciones":
                qs = qs.filter(es_renovacion=True)

        # Optimizar consultas
        qs = qs.select_related("cliente", "oficina", "compania_obj", "cobertura_obj", "vendedor")
        qs = qs.order_by("fecha_emision", "id")

        # ── Construir Excel ─────────────────────────────────────
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True, "remove_timezone": True})

        # Formatos
        fmt_title = workbook.add_format({
            "bold": True, "font_size": 14, "font_color": "white",
            "bg_color": "#0F172A", "align": "center", "valign": "vcenter",
            "border": 1,
        })
        fmt_subtitle = workbook.add_format({
            "italic": True, "font_size": 10, "font_color": "#475569",
            "align": "center", "valign": "vcenter",
        })
        fmt_header = workbook.add_format({
            "bold": True, "font_color": "white", "bg_color": "#1E293B",
            "align": "center", "valign": "vcenter", "border": 1,
            "text_wrap": True,
        })
        fmt_cell = workbook.add_format({
            "valign": "vcenter", "border": 1, "border_color": "#E2E8F0",
        })
        fmt_cell_alt = workbook.add_format({
            "valign": "vcenter", "border": 1, "border_color": "#E2E8F0",
            "bg_color": "#F8FAFC",
        })
        fmt_date = workbook.add_format({
            "num_format": "dd/mm/yyyy", "valign": "vcenter", "border": 1,
            "border_color": "#E2E8F0",
        })
        fmt_date_alt = workbook.add_format({
            "num_format": "dd/mm/yyyy", "valign": "vcenter", "border": 1,
            "border_color": "#E2E8F0", "bg_color": "#F8FAFC",
        })
        fmt_money = workbook.add_format({
            "num_format": '"$"#,##0.00', "valign": "vcenter", "border": 1,
            "border_color": "#E2E8F0",
        })
        fmt_money_alt = workbook.add_format({
            "num_format": '"$"#,##0.00', "valign": "vcenter", "border": 1,
            "border_color": "#E2E8F0", "bg_color": "#F8FAFC",
        })
        fmt_total = workbook.add_format({
            "bold": True, "font_color": "white", "bg_color": "#064E3B",
            "align": "right", "valign": "vcenter", "border": 1,
        })
        fmt_total_count = workbook.add_format({
            "bold": True, "font_color": "white", "bg_color": "#064E3B",
            "align": "center", "valign": "vcenter", "border": 1,
        })

        # ── Helpers ─────────────────────────────────────────────
        def _get_cliente_nombre(p):
            c = getattr(p, "cliente", None)
            if not c:
                return ""
            ap = getattr(c, "apellido", "") or ""
            no = getattr(c, "nombre", "") or ""
            return f"{ap}, {no}".strip(", ").strip()

        def _get_cliente_dni(p):
            c = getattr(p, "cliente", None)
            if not c:
                return ""
            return (
                getattr(c, "dni", "")
                or getattr(c, "dni_cuit_cuil", "")
                or getattr(c, "documento", "")
                or ""
            )

        def _get_cliente_tel(p):
            c = getattr(p, "cliente", None)
            if not c:
                return ""
            for k in ("telefono", "celular", "whatsapp", "telefono1"):
                v = getattr(c, k, None)
                if v:
                    return str(v).strip()
            return ""

        def _get_cliente_email(p):
            c = getattr(p, "cliente", None)
            if not c:
                return ""
            return getattr(c, "email", "") or ""

        def _get_compania(p):
            if getattr(p, "compania_obj", None):
                return p.compania_obj.nombre
            return getattr(p, "compania", "") or ""

        def _get_cobertura(p):
            if getattr(p, "cobertura_obj", None):
                return p.cobertura_obj.nombre
            return getattr(p, "cobertura", "") or ""

        def _get_oficina(p):
            ofi = getattr(p, "oficina", None)
            if not ofi:
                return ""
            return getattr(ofi, "nombre", str(ofi))

        def _get_vendedor(p):
            v = getattr(p, "vendedor", None)
            if not v:
                return ""
            usuario = getattr(v, "usuario", None) or getattr(v, "user", None)
            if usuario:
                full = f"{getattr(usuario, 'first_name', '')} {getattr(usuario, 'last_name', '')}".strip()
                return full or getattr(usuario, "username", "") or ""
            return str(v)

        def _get_tipo_alta(p):
            return "Renovación" if getattr(p, "es_renovacion", False) else "Nueva"

        def _get_poliza_origen(p):
            origen = getattr(p, "poliza_origen", None)
            if origen:
                return getattr(origen, "numero_poliza", "") or f"ID {origen.id}"
            return ""

        def _get_fase(p):
            f = getattr(p, "fase", "") or ""
            return f.capitalize() if f else ""

        def _get_estado(p):
            e = getattr(p, "estado", "") or ""
            return e.capitalize()

        def _format_date_val(v):
            if not v:
                return ""
            if hasattr(v, "date") and callable(getattr(v, "date", None)):
                try:
                    return v.date()
                except Exception:
                    return v
            return v

        # ── Definición de columnas (23 columnas) ────────────────
        columns = [
            ("Patente",           12, lambda p: getattr(p, "patente", "") or "", "text"),
            ("Día de alta",       12, lambda p: _format_date_val(getattr(p, "fecha_emision", None)), "date"),
            ("Compañía",          22, _get_compania, "text"),
            ("Oficina",           18, _get_oficina, "text"),
            ("Marca",             16, lambda p: getattr(p, "marca", "") or "", "text"),
            ("Modelo",            18, lambda p: getattr(p, "modelo", "") or "", "text"),
            ("Año",                8, lambda p: getattr(p, "anio", "") or "", "text"),
            ("Tipo vehículo",     14, lambda p: getattr(p, "tipo", "") or "", "text"),
            ("Nº Póliza",         18, lambda p: getattr(p, "numero_poliza", "") or "", "text"),
            ("Cobertura",         20, _get_cobertura, "text"),
            ("Cliente",           28, _get_cliente_nombre, "text"),
            ("DNI/CUIT",          14, _get_cliente_dni, "text"),
            ("Teléfono",          16, _get_cliente_tel, "text"),
            ("Email",             24, _get_cliente_email, "text"),
            ("Precio cuota",      14, lambda p: float(p.precio_cuota) if p.precio_cuota else 0, "money"),
            ("Cant. cuotas",      10, lambda p: getattr(p, "cantidad_cuotas", 0) or 0, "text"),
            ("Primer pago",       12, lambda p: _format_date_val(getattr(p, "primer_pago", None)), "date"),
            ("Vencimiento",       12, lambda p: _format_date_val(getattr(p, "fecha_vencimiento", None)), "date"),
            ("Estado",            14, _get_estado, "text"),
            ("Fase",              12, _get_fase, "text"),
            ("Tipo alta",         12, _get_tipo_alta, "text"),
            ("Póliza origen",     16, _get_poliza_origen, "text"),
            ("Vendedor",          22, _get_vendedor, "text"),
        ]

        # ── Hoja 1: Pólizas ─────────────────────────────────────
        ws = workbook.add_worksheet("Pólizas")

        # Título (row 0)
        ws.merge_range(0, 0, 0, len(columns) - 1, "REPORTE DE EMISIONES", fmt_title)
        ws.set_row(0, 28)

        # Subtítulo (row 1)
        tipo_label = {
            "nuevas": "Solo Altas Nuevas",
            "renovaciones": "Solo Renovaciones",
            "todas": "Todas (Nuevas + Renovaciones)",
        }.get(tipo_param, "Todas")

        subtitle = (
            f"Período: {desde.strftime('%d/%m/%Y')} al {hasta.strftime('%d/%m/%Y')}  ·  "
            f"Filtro: {tipo_label}"
            + (f"  ·  Compañía: {compania}" if compania else "")
        )
        ws.merge_range(1, 0, 1, len(columns) - 1, subtitle, fmt_subtitle)
        ws.set_row(1, 18)

        # Cabecera (row 2)
        header_row = 2
        for col_idx, (header, width, _, _) in enumerate(columns):
            ws.write(header_row, col_idx, header, fmt_header)
            ws.set_column(col_idx, col_idx, width)
        ws.set_row(header_row, 30)

        # Filas de datos
        row_count = 0
        for poliza in qs.iterator():
            row = header_row + 1 + row_count
            is_alt = (row_count % 2) == 1

            for col_idx, (_, _, getter, ftype) in enumerate(columns):
                try:
                    value = getter(poliza)
                except Exception:
                    value = ""

                if ftype == "date":
                    fmt = fmt_date_alt if is_alt else fmt_date
                    if value and not isinstance(value, str):
                        try:
                            ws.write_datetime(row, col_idx, value, fmt)
                        except Exception:
                            ws.write(row, col_idx, str(value) if value else "", fmt)
                    else:
                        ws.write(row, col_idx, "", fmt)

                elif ftype == "money":
                    fmt = fmt_money_alt if is_alt else fmt_money
                    try:
                        ws.write_number(row, col_idx, float(value or 0), fmt)
                    except Exception:
                        ws.write(row, col_idx, 0, fmt)

                else:
                    fmt = fmt_cell_alt if is_alt else fmt_cell
                    ws.write(row, col_idx, value, fmt)

            row_count += 1

        # Fila TOTAL al final
        if row_count > 0:
            total_row = header_row + 1 + row_count
            ws.merge_range(
                total_row, 0, total_row, len(columns) - 2,
                "TOTAL DE PÓLIZAS:",
                fmt_total,
            )
            ws.write(total_row, len(columns) - 1, row_count, fmt_total_count)
            ws.set_row(total_row, 24)

        # Freeze panes (header fijo)
        ws.freeze_panes(header_row + 1, 0)

        # Autofiltro
        if row_count > 0:
            ws.autofilter(header_row, 0, header_row + row_count, len(columns) - 1)

        # ── Hoja 2: Resumen ─────────────────────────────────────
        ws2 = workbook.add_worksheet("Resumen")
        ws2.set_column(0, 0, 32)
        ws2.set_column(1, 1, 18)

        fmt_h2 = workbook.add_format({
            "bold": True, "font_color": "white", "bg_color": "#1E293B",
            "align": "left", "valign": "vcenter", "border": 1,
        })
        fmt_label = workbook.add_format({
            "bold": True, "valign": "vcenter", "border": 1,
            "border_color": "#E2E8F0", "bg_color": "#F1F5F9",
        })
        fmt_value = workbook.add_format({
            "valign": "vcenter", "border": 1,
            "border_color": "#E2E8F0", "align": "right",
        })

        ws2.merge_range(0, 0, 0, 1, "RESUMEN", fmt_title)
        ws2.set_row(0, 28)

        # Conteos
        total = row_count
        nuevas = qs.filter(es_renovacion=False).count() if hasattr(Poliza, "es_renovacion") else 0
        renovaciones = qs.filter(es_renovacion=True).count() if hasattr(Poliza, "es_renovacion") else 0

        summary_rows = [
            ("Período desde", desde.strftime("%d/%m/%Y")),
            ("Período hasta", hasta.strftime("%d/%m/%Y")),
            ("Total pólizas emitidas", total),
            ("Altas nuevas (es_renovacion=False)", nuevas),
            ("Renovaciones (es_renovacion=True)", renovaciones),
            ("Filtro aplicado", tipo_label),
            ("Compañía filtrada", compania or "Todas"),
            ("Generado el", timezone.localtime().strftime("%d/%m/%Y %H:%M")),
            ("Generado por", getattr(user, "username", "—") or "—"),
        ]

        for i, (label, value) in enumerate(summary_rows):
            ws2.write(i + 1, 0, label, fmt_label)
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                ws2.write_number(i + 1, 1, value, fmt_value)
            else:
                ws2.write(i + 1, 1, str(value), fmt_value)
            ws2.set_row(i + 1, 20)

        # Desglose por compañía
        offset = len(summary_rows) + 3
        ws2.merge_range(offset, 0, offset, 1, "POR COMPAÑÍA", fmt_h2)
        ws2.set_row(offset, 22)

        from collections import defaultdict
        por_compania = defaultdict(int)
        for p in qs.iterator():
            por_compania[_get_compania(p) or "—"] += 1

        for i, (cia, cnt) in enumerate(sorted(por_compania.items(), key=lambda x: -x[1])):
            ws2.write(offset + 1 + i, 0, cia, fmt_label)
            ws2.write_number(offset + 1 + i, 1, cnt, fmt_value)
            ws2.set_row(offset + 1 + i, 18)

        # Desglose por oficina
        offset2 = offset + 1 + len(por_compania) + 2
        ws2.merge_range(offset2, 0, offset2, 1, "POR OFICINA", fmt_h2)
        ws2.set_row(offset2, 22)

        por_oficina = defaultdict(int)
        for p in qs.iterator():
            por_oficina[_get_oficina(p) or "—"] += 1

        for i, (ofi, cnt) in enumerate(sorted(por_oficina.items(), key=lambda x: -x[1])):
            ws2.write(offset2 + 1 + i, 0, ofi, fmt_label)
            ws2.write_number(offset2 + 1 + i, 1, cnt, fmt_value)
            ws2.set_row(offset2 + 1 + i, 18)

        workbook.close()
        output.seek(0)

        # ── Log del export ──────────────────────────────────────
        try:
            ExportLog.objects.create(
                usuario=user if getattr(user, "is_authenticated", False) else None,
                tipo="polizas_emisiones_excel",
                parametros={
                    "desde": desde.isoformat(),
                    "hasta": hasta.isoformat(),
                    "tipo": tipo_param,
                    "oficina": str(params.get("oficina") or "") or None,
                    "compania": compania or None,
                    "total_filas": row_count,
                },
            )
        except Exception:
            pass

        # ── Devolver archivo ────────────────────────────────────
        filename = f"Emisiones_{desde.strftime('%Y%m%d')}_{hasta.strftime('%Y%m%d')}_{tipo_param}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response