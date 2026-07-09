# notificaciones/services_reporte_contactos.py
"""
Reporte de contactos pendientes para gestión manual.

Genera PDF o Excel con clientes que tienen cuotas vencidas/por vencer
en los deltas: -7, -3, 0, +3, +7 días respecto a la FECHA DE PAGO
(= fin de cobertura de la cuota anterior, o sea cuándo el cliente tiene que
venir a pagar la cuota). El vencimiento propio de la cuota queda como referencia.

Útil para que el operador llame/escriba manualmente a los clientes que el
sistema no contactó automáticamente (porque solo se mandan recordatorios
en -3, 0, +3).
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date, timedelta
from typing import Any, Dict, List, Optional, Tuple

from django.utils import timezone

from notificaciones.services_cuotas import (
    obtener_cuotas_candidatas,
    _apply_oficina_filter,
    _normalize_oficina_bucket,
    _resolver_alias_transferencia,
    _get_numero_whatsapp,
    _descripcion_cuota,
)
from pagos.models import Cuota


def _es_primera_cuota_renovacion(cuota) -> bool:
    """
    True si la cuota es la PRIMERA (nro 1) de una póliza marcada como renovación.
    Definido localmente para que este reporte no dependa de cambios en services_cuotas.py.
    """
    pol = getattr(cuota, "poliza", None)
    if not pol or not getattr(pol, "es_renovacion", False):
        return False
    return getattr(cuota, "cuota_nro", None) == 1


def _fecha_pago_objetivo(cuota) -> Optional[date]:
    """
    Fecha en la que el cliente tiene que VENIR A PAGAR esta cuota
    = fin de cobertura de la cuota ANTERIOR (cuota_nro - 1), que en una póliza
    al día es la última cuota PAGADA.

    Cuota #1 (no hay anterior): usamos su propio vencimiento. En una renovación
    eso es justo el día en que termina la cobertura de la póliza vieja.

    Definido localmente para no acoplar el reporte a services_cuotas.py.
    """
    nro = getattr(cuota, "cuota_nro", None)
    if nro and nro > 1:
        vto_anterior = (
            Cuota.objects
            .filter(poliza_id=cuota.poliza_id, cuota_nro=nro - 1)
            .values_list("fecha_vencimiento", flat=True)
            .first()
        )
        if vto_anterior:
            return vto_anterior
    return getattr(cuota, "fecha_vencimiento", None)


# Deltas que se reportan SIEMPRE (independientes de los que se envían automáticos).
# Respecto a la FECHA DE PAGO (fin de cobertura de la cuota anterior).
# POSITIVO = faltan días para pagar (+1 = vence mañana) · NEGATIVO = ya venció.
REPORT_DELTAS: List[int] = [-30, -7, -3, -2, 0, 1, 3, 7]

# 🆕 Venta cruzada: clientes que pagaron hace N días → buen momento para ofrecerles otros seguros.
OFERTA_DIAS_DESPUES_PAGO = 14

# Mensaje sugerido para que el operador ofrezca otros seguros (lo copia y manda a mano).
MENSAJE_OFERTA = (
    "¡Hola! 😊 Además del seguro de tu auto, en Thames también tenemos seguros de "
    "celular, de comercio y servicios para profesionales (estudios/abogados). "
    "Si te interesa cotizar alguno, escribinos y te asesoramos sin compromiso."
)


# ─────────────────────────────────────────────────────────────────────
#  Recolección de datos
# ─────────────────────────────────────────────────────────────────────

def _safe(v) -> str:
    return ("" if v is None else str(v)).strip()


def _fmt_money(v) -> str:
    try:
        n = float(v or 0)
    except (TypeError, ValueError):
        return "-"
    # 12345.67 -> $12.345,67
    s = f"{n:,.2f}"
    return "$" + s.replace(",", "X").replace(".", ",").replace("X", ".")


def _delta_label(delta: int) -> str:
    if delta == 0:
        return "Vence HOY"
    if delta == 1:
        return "Vence MAÑANA"
    if delta > 0:
        return f"En +{delta} días"
    return f"Vencida hace {abs(delta)} días"


def _truncar_ancho(texto, max_pts, font="Helvetica", size=8) -> str:
    """
    Acorta un texto agregando '…' para que NO se pase del ancho de su columna
    en el PDF (así no pisa la columna de al lado). Mide el ancho real del texto.
    """
    from reportlab.pdfbase.pdfmetrics import stringWidth
    t = str(texto or "")
    if not t:
        return ""
    if stringWidth(t, font, size) <= max_pts:
        return t
    while t and stringWidth(t + "…", font, size) > max_pts:
        t = t[:-1]
    return (t.rstrip() + "…") if t else "…"


def recolectar_filas(
    hoy: Optional[date] = None,
    oficina: Optional[str] = None,
    alias_transferencia: Optional[str] = None,
    medio_cobro_id: Optional[int] = None,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Devuelve (filas, meta) donde:
      filas: lista de dicts (1 por cuota a contactar)
      meta:  dict con info global (oficina, alias, totales por delta, etc.)
    """
    hoy = hoy or timezone.localdate()
    oficina_norm = _normalize_oficina_bucket(oficina) if oficina else None
    alias_resuelto, titular = _resolver_alias_transferencia(alias_transferencia, medio_cobro_id)

    cuotas = list(obtener_cuotas_candidatas(hoy, oficina=oficina_norm))

    filas: List[Dict[str, Any]] = []
    totales_por_delta: Dict[int, int] = {d: 0 for d in REPORT_DELTAS}

    for cuota in cuotas:
        fv = getattr(cuota, "fecha_vencimiento", None)
        if not fv:
            continue
        # 🎯 Disparamos por la fecha en que hay que PAGAR (fin de cobertura de la
        # cuota anterior), igual que el WhatsApp. El vto propio queda de referencia.
        fecha_objetivo = _fecha_pago_objetivo(cuota)
        if not fecha_objetivo:
            continue
        delta = (fecha_objetivo - hoy).days
        if delta not in REPORT_DELTAS:
            continue

        poliza = getattr(cuota, "poliza", None)
        cliente = getattr(poliza, "cliente", None) if poliza else None
        if not cliente:
            continue

        patente, vehiculo = _descripcion_cuota(cuota)
        telefono = _get_numero_whatsapp(cuota) or ""

        nombre = (
            _safe(getattr(cliente, "nombre_apellido", None))
            or " ".join(filter(None, [
                _safe(getattr(cliente, "nombre", "")),
                _safe(getattr(cliente, "apellido", "")),
            ])).strip()
            or "—"
        )

        compania = _safe(getattr(poliza, "compania", ""))
        nro_pol = _safe(getattr(poliza, "numero_poliza", "")) or f"#{getattr(poliza, 'id', '')}"

        # 🆕 ¿Es la primera cuota de una póliza renovada?
        es_renov = _es_primera_cuota_renovacion(cuota)

        filas.append({
            "delta":       delta,
            # Si es renovación, el detalle lo marca; el "Estado" igual muestra los días.
            "delta_label": "RENOVACIÓN" if es_renov else _delta_label(delta),
            "es_renovacion": es_renov,
            "estado":      f"{delta:+d}" if delta != 0 else "0",
            "cliente":     nombre,
            "telefono":    telefono,
            "patente":     patente,
            "vehiculo":    vehiculo,
            "compania":    compania,
            "nro_poliza":  nro_pol,
            "cuota_nro":   _safe(getattr(cuota, "cuota_nro", "")),
            "fecha_pago":  fecha_objetivo.strftime("%d/%m/%Y") if fecha_objetivo else "",
            "vencimiento": fv.strftime("%d/%m/%Y") if fv else "",
            "monto":       float(getattr(cuota, "monto", 0) or 0),
            "monto_fmt":   _fmt_money(getattr(cuota, "monto", 0)),
        })
        totales_por_delta[delta] = totales_por_delta.get(delta, 0) + 1

    # ── 🆕 Oportunidad de venta cruzada ──────────────────────────────────
    # Clientes que PAGARON hace OFERTA_DIAS_DESPUES_PAGO días → buen momento
    # para ofrecerles otros seguros (celular, comercio, profesionales).
    fecha_pago_oferta = hoy - timedelta(days=OFERTA_DIAS_DESPUES_PAGO)
    qs_ofertas = (
        Cuota.objects
        .select_related("poliza", "poliza__cliente")
        .filter(pagado=True, fecha_pago=fecha_pago_oferta)
        .exclude(poliza__estado__in=["cancelada", "anulada", "baja", "eliminada"])
        .order_by("poliza__cliente_id")
    )
    if oficina_norm:
        qs_ofertas = _apply_oficina_filter(qs_ofertas, oficina_norm)

    clientes_oferta = set()
    for cuota in qs_ofertas:
        poliza = getattr(cuota, "poliza", None)
        cliente = getattr(poliza, "cliente", None) if poliza else None
        if not cliente or cliente.id in clientes_oferta:
            continue
        clientes_oferta.add(cliente.id)

        patente, vehiculo = _descripcion_cuota(cuota)
        telefono = _get_numero_whatsapp(cuota) or ""
        nombre = (
            _safe(getattr(cliente, "nombre_apellido", None))
            or " ".join(filter(None, [
                _safe(getattr(cliente, "nombre", "")),
                _safe(getattr(cliente, "apellido", "")),
            ])).strip()
            or "—"
        )
        compania = _safe(getattr(poliza, "compania", ""))
        fv = getattr(cuota, "fecha_vencimiento", None)
        fp = getattr(cuota, "fecha_pago", None)

        filas.append({
            "delta":       None,            # no es cobranza → sin delta de vencimiento
            "delta_label": "Ofrecer seguros",
            "es_renovacion": False,
            "es_oferta":   True,
            "estado":      "VENTA",
            "cliente":     nombre,
            "telefono":    telefono,
            "patente":     patente,
            "vehiculo":    vehiculo,
            "compania":    compania,
            "nro_poliza":  _safe(getattr(poliza, "numero_poliza", "")) or f"#{getattr(poliza, 'id', '')}",
            "cuota_nro":   _safe(getattr(cuota, "cuota_nro", "")),
            "fecha_pago":  fp.strftime("%d/%m/%Y") if fp else "",
            "vencimiento": fv.strftime("%d/%m/%Y") if fv else "",
            "monto":       float(getattr(cuota, "monto", 0) or 0),
            "monto_fmt":   _fmt_money(getattr(cuota, "monto", 0)),
        })

    # Orden: cobranza por urgencia (-7 → +7), después las ofertas (al final), luego por nombre
    orden_delta = {-30: 0, -7: 1, -3: 2, -2: 3, 0: 4, 1: 5, 3: 6, 7: 7}
    filas.sort(key=lambda r: (orden_delta.get(r["delta"], 99), r["cliente"].lower()))

    meta = {
        "hoy":             hoy.strftime("%d/%m/%Y"),
        "oficina":         oficina_norm or "Todas",
        "alias":           alias_resuelto or "",
        "titular":         titular or "",
        "total":           len(filas),
        "total_renovaciones": sum(1 for f in filas if f.get("es_renovacion")),
        "total_ofertas":   len(clientes_oferta),
        "oferta_dias":     OFERTA_DIAS_DESPUES_PAGO,
        "mensaje_oferta":  MENSAJE_OFERTA,
        "totales_delta":   totales_por_delta,
        "deltas":          REPORT_DELTAS,
    }
    return filas, meta


# ─────────────────────────────────────────────────────────────────────
#  Export EXCEL (openpyxl)
# ─────────────────────────────────────────────────────────────────────

def generar_excel(filas: List[Dict[str, Any]], meta: Dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos pendientes"

    # Estilos
    header_fill = PatternFill("solid", fgColor="1F2937")  # gris oscuro
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="111827")
    sub_font   = Font(size=10, color="6B7280", italic=True)
    border     = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center")
    right  = Alignment(horizontal="right", vertical="center")

    # ── Cabecera del reporte ──
    ws.merge_cells("A1:J1")
    ws["A1"] = "Reporte de contactos pendientes"
    ws["A1"].font = title_font
    ws["A1"].alignment = left

    ws.merge_cells("A2:J2")
    ws["A2"] = (
        f"Fecha: {meta['hoy']}   •   Oficina: {meta['oficina']}   •   "
        f"Total: {meta['total']}   •   Renovaciones: {meta.get('total_renovaciones', 0)}"
        f"   •   Ofertas: {meta.get('total_ofertas', 0)}"
    )
    ws["A2"].font = sub_font
    ws["A2"].alignment = left

    if meta.get("alias"):
        ws.merge_cells("A3:J3")
        ws["A3"] = (
            f"Alias / medio de pago: {meta['alias']}"
            + (f"   •   Titular: {meta['titular']}" if meta.get("titular") else "")
        )
        ws["A3"].font = sub_font
        ws["A3"].alignment = left

    # ── Header de tabla ──
    headers = [
        "Estado", "Detalle", "Cliente", "Teléfono", "Patente",
        "Vehículo", "Compañía", "Cuota", "Fecha de pago", "Vencimiento",
    ]
    row_header = 5
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row_header, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    ws.row_dimensions[row_header].height = 24

    # ── Filas ──
    r = row_header + 1
    for fila in filas:
        values = [
            fila["estado"],
            fila["delta_label"],
            fila["cliente"],
            fila["telefono"],
            fila["patente"],
            fila["vehiculo"],
            fila["compania"],
            fila["cuota_nro"],
            fila["fecha_pago"],
            fila["vencimiento"],
        ]
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = border
            if i == 1:  # Estado
                c.alignment = center
                c.font = Font(bold=True, color=_color_delta(fila["delta"]))
            elif i == 5:  # Patente (resaltada para buscarla rápido en la app)
                c.alignment = center
                c.font = Font(bold=True, color="111827")
            else:
                c.alignment = left
                # 🆕 Marca de renovación en la columna Detalle
                if i == 2 and fila.get("es_renovacion"):
                    c.font = Font(bold=True, color="7C3AED")  # violeta
        # Zebra
        if (r - row_header) % 2 == 0:
            fill = PatternFill("solid", fgColor="F9FAFB")
            for col in range(1, len(values) + 1):
                ws.cell(row=r, column=col).fill = fill
        r += 1

    # Fila de totales por delta
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    totales_txt = "   |   ".join(
        f"{d:+d}: {meta['totales_delta'].get(d, 0)}" for d in meta["deltas"]
    )
    ws.cell(row=r, column=1, value=f"Totales por estado:   {totales_txt}")
    ws.cell(row=r, column=1).font = Font(bold=True, color="374151")
    ws.cell(row=r, column=1).alignment = left

    # 🆕 Nota sobre las filas de OFERTA (venta cruzada)
    if meta.get("total_ofertas"):
        r += 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        ws.cell(
            row=r, column=1,
            value=(f"Filas VENTA: clientes que pagaron hace {meta.get('oferta_dias', 14)} días "
                   f"— buen momento para ofrecerles otros seguros."),
        )
        ws.cell(row=r, column=1).font = Font(bold=True, color="0E7490")
        ws.cell(row=r, column=1).alignment = left

        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        ws.cell(row=r, column=1, value=f"Mensaje sugerido: {meta.get('mensaje_oferta', '')}")
        ws.cell(row=r, column=1).font = sub_font
        ws.cell(row=r, column=1).alignment = left

    # Ancho de columnas
    widths = [10, 22, 28, 16, 12, 22, 18, 8, 13, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes (mantiene visible el header al scrollear)
    ws.freeze_panes = ws[f"A{row_header + 1}"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _color_delta(delta) -> str:
    """Hex sin # para openpyxl."""
    if delta is None:
        return "0E7490"  # teal → fila de OFERTA (venta cruzada)
    if delta <= -7:
        return "991B1B"  # rojo oscuro
    if delta == -3:
        return "DC2626"  # rojo
    if delta == -2:
        return "B91C1C"  # rojo intenso → último aviso (mañana se da de baja)
    if delta == 0:
        return "D97706"  # naranja
    if delta == 1:
        return "EA580C"  # naranja fuerte → vence mañana (urgente)
    if delta == 3:
        return "059669"  # verde
    return "2563EB"      # azul (+7)


# ─────────────────────────────────────────────────────────────────────
#  Export PDF (reportlab)
# ─────────────────────────────────────────────────────────────────────

def generar_pdf(filas: List[Dict[str, Any]], meta: Dict[str, Any]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title="Reporte de contactos pendientes",
    )

    styles = getSampleStyleSheet()
    title_st = ParagraphStyle(
        "title", parent=styles["Heading1"],
        fontSize=16, textColor=colors.HexColor("#111827"), spaceAfter=4,
    )
    sub_st = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#6B7280"), spaceAfter=2,
    )
    foot_st = ParagraphStyle(
        "foot", parent=styles["Normal"],
        fontSize=8, textColor=colors.HexColor("#6B7280"),
    )

    story = []
    story.append(Paragraph("Reporte de contactos pendientes", title_st))
    story.append(Paragraph(
        f"<b>Fecha:</b> {meta['hoy']} &nbsp;&nbsp;•&nbsp;&nbsp; "
        f"<b>Oficina:</b> {meta['oficina']} &nbsp;&nbsp;•&nbsp;&nbsp; "
        f"<b>Total:</b> {meta['total']} &nbsp;&nbsp;•&nbsp;&nbsp; "
        f"<b>Renovaciones:</b> {meta.get('total_renovaciones', 0)} &nbsp;&nbsp;•&nbsp;&nbsp; "
        f"<b>Ofertas:</b> {meta.get('total_ofertas', 0)}",
        sub_st,
    ))
    if meta.get("alias"):
        story.append(Paragraph(
            f"<b>Alias:</b> {meta['alias']}"
            + (f" &nbsp;&nbsp;•&nbsp;&nbsp; <b>Titular:</b> {meta['titular']}" if meta.get("titular") else ""),
            sub_st,
        ))

    # Totales por delta
    tot_txt = "  |  ".join(
        f"<b>{d:+d}</b>: {meta['totales_delta'].get(d, 0)}" for d in meta["deltas"]
    )
    story.append(Paragraph(f"<b>Totales por estado:</b> &nbsp; {tot_txt}", sub_st))
    story.append(Spacer(1, 6 * mm))

    # Tabla
    headers = [
        "Estado", "Detalle", "Cliente", "Teléfono", "Patente",
        "Vehículo", "Compañía", "Cuota", "Fecha pago", "Vencimiento",
    ]

    # Anchos de columna (definidos ANTES para poder acortar el texto que no entra
    # y que no pise la columna de al lado). Suma ≈ 259mm (A4 landscape, márgenes 12mm).
    col_widths = [15*mm, 34*mm, 40*mm, 26*mm, 24*mm, 36*mm, 26*mm, 12*mm, 23*mm, 23*mm]
    _PAD = 8  # LEFTPADDING + RIGHTPADDING del estilo de la tabla
    w_detalle  = col_widths[1] - _PAD
    w_cliente  = col_widths[2] - _PAD
    w_vehiculo = col_widths[5] - _PAD
    w_compania = col_widths[6] - _PAD

    data = [headers]
    for f in filas:
        data.append([
            f["estado"],
            _truncar_ancho(f["delta_label"], w_detalle),
            _truncar_ancho(f["cliente"], w_cliente),     # nombre y apellido: se acorta si no entra
            f["telefono"],
            f["patente"],
            _truncar_ancho(f["vehiculo"], w_vehiculo),   # marca + modelo: se acorta si no entra
            _truncar_ancho(f["compania"], w_compania),
            f["cuota_nro"], f["fecha_pago"], f["vencimiento"],
        ])

    if not filas:
        data.append(["—"] * len(headers))

    tabla = Table(data, colWidths=col_widths, repeatRows=1)

    # Estilo base
    ts = TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",     (0, 0), (-1, 0), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING",    (0, 0), (-1, 0), 8),
        # Body
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("VALIGN",   (0, 1), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 1), (-1, -1), 4),
        ("RIGHTPADDING", (0, 1), (-1, -1), 4),
        ("TOPPADDING",   (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 1), (-1, -1), 4),
        # Estado: centrado y bold
        ("ALIGN",    (0, 1), (0, -1), "CENTER"),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        # 🆕 Patente: centrada y en negrita (para encontrarla rápido en la app)
        ("ALIGN",    (4, 1), (4, -1), "CENTER"),
        ("FONTNAME", (4, 1), (4, -1), "Helvetica-Bold"),
        # Bordes finos
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
        ("LINEBELOW", (0, 0), (-1, 0), 1.2, colors.HexColor("#1F2937")),
    ])

    # Zebra + color de estado fila por fila
    for i, f in enumerate(filas, start=1):
        if i % 2 == 0:
            ts.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F9FAFB"))
        ts.add("TEXTCOLOR", (0, i), (0, i), colors.HexColor(f"#{_color_delta(f['delta'])}"))
        # 🆕 Marca de renovación: columna Detalle en violeta y negrita
        if f.get("es_renovacion"):
            ts.add("TEXTCOLOR", (1, i), (1, i), colors.HexColor("#7C3AED"))
            ts.add("FONTNAME", (1, i), (1, i), "Helvetica-Bold")

    tabla.setStyle(ts)
    story.append(tabla)

    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph(
        "Generado automáticamente — usar este listado para gestión manual de contactos. "
        "Las filas marcadas como RENOVACIÓN corresponden a la primera cuota de una póliza renovada.",
        foot_st,
    ))
    if meta.get("total_ofertas"):
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(
            f"<b>Filas VENTA:</b> clientes que pagaron hace {meta.get('oferta_dias', 14)} días "
            f"— buen momento para ofrecerles otros seguros (celular, comercio, profesionales).",
            foot_st,
        ))
        story.append(Paragraph(
            f"<b>Mensaje sugerido:</b> {meta.get('mensaje_oferta', '')}",
            foot_st,
        ))

    doc.build(story)
    return buf.getvalue()