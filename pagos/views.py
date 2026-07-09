# pagos/views.py
import calendar
from datetime import timedelta, date
from decimal import Decimal, InvalidOperation
import csv
from io import StringIO, BytesIO

from django.db import models, transaction
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db.models import Q
from django.db.models import OuterRef, Subquery, DateField
from django.db.models import Max, Avg, Count, F, ExpressionWrapper, FloatField
from django.core.exceptions import FieldDoesNotExist
from django.db.models.functions import Coalesce, ExtractDay, ExtractHour, ExtractMinute

from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

from .models import Pago, Cuota, MedioCobro, AlertaEnviada
from polizas.models import Poliza
from balanzes.models import Ingreso as BalanceIngreso
from solicitudes.models import Empleado

from pagos.management.commands.enviar_alertas import ejecutar_alertas
from pagos.handlers.registrar_pago import registrar_pago_handler, _enviar_gracias_portal
from pagos.utils.factura import generar_factura_pdf

MAX_HISTORIAL_ALL_ROWS = 50000


# -------------------------
# 🚀 SUPER TRADUCTOR MULTI-TENANT (A PRUEBA DE FALLOS)
# -------------------------
def _get_seguridad_oficina_brute(request, requested_oficina=""):
    user = request.user
    if not user.is_authenticated:
        return ["BLOQUEADO"]
        
    es_admin = user.is_superuser or (hasattr(user, 'perfil') and user.perfil.rol == 'ADMIN')
    
    target = None
    if es_admin:
        val = str(requested_oficina or "").strip()
        if not val or val.upper() == "ALL":
            return [] # Sin filtro
        target = val
    else:
        if hasattr(user, 'perfil') and user.perfil.oficina:
            target = user.perfil.oficina
        else:
            return ["BLOQUEADO"]
            
    synonyms = set()
    
    if hasattr(target, 'codigo') and target.codigo:
        synonyms.add(str(target.codigo).strip().lower())
    if hasattr(target, 'id') and target.id:
        synonyms.add(str(target.id).strip().lower())
    if hasattr(target, 'nombre') and target.nombre:
        synonyms.add(str(target.nombre).strip().lower())
        
    if isinstance(target, str):
        s = target.strip().lower()
        synonyms.add(s)
        try:
            from django.apps import apps
            Oficina = apps.get_model("usuarios", "Oficina")
            if s.isdigit():
                ofi = Oficina.objects.filter(Q(codigo=s) | Q(id=s)).first()
            else:
                ofi = Oficina.objects.filter(nombre__icontains=s).first()
            if ofi:
                synonyms.add(str(ofi.codigo).strip().lower())
                synonyms.add(str(ofi.id).strip().lower())
                synonyms.add(str(ofi.nombre).strip().lower())
        except Exception:
            pass
            
    final_synonyms = set(synonyms)
    for s in synonyms:
        if "1" == s or "esquina" in s or "5 esquinas" in s:
            final_synonyms.update(["1", "5 esquinas", "ofi 1", "ofi1"])
        elif "2" == s or "axion" in s:
            final_synonyms.update(["2", "axion", "ofi 2", "ofi2"])
        elif "3" == s or "39" in s or "kilometro" in s:
            final_synonyms.update(["3", "39", "kilometro 39", "ofi 3", "ofi3"])
            
    return list(final_synonyms)


def _is_poliza_oficina_fk() -> bool:
    try:
        f = Poliza._meta.get_field("oficina")
        return bool(getattr(f, "is_relation", False))
    except (FieldDoesNotExist, Exception):
        return False


def _build_oficina_q_from_keys(keys):
    if not keys: return Q()
    if "BLOQUEADO" in keys: return Q(pk__isnull=True)
    
    is_fk = _is_poliza_oficina_fk()
    q_final = Q()
    
    for k in keys:
        s = str(k).strip()
        if not s: continue
        
        if is_fk:
            if s.isdigit():
                q_final |= Q(poliza__oficina_id=int(s))
            q_final |= Q(poliza__oficina__nombre__icontains=s)
            try: q_final |= Q(poliza__oficina__codigo=s)
            except Exception: pass
        else:
            q_final |= Q(poliza__oficina__icontains=s)
            q_final |= Q(poliza__oficina__iexact=s)
            
    return q_final


def _oficina_q_sobre_poliza(keys):
    """Igual que _build_oficina_q_from_keys pero para querysets de Poliza
    (sin el prefijo 'poliza__', porque ya estamos parados en Poliza)."""
    if not keys: return Q()
    if "BLOQUEADO" in keys: return Q(pk__isnull=True)
    is_fk = _is_poliza_oficina_fk()
    q_final = Q()
    for k in keys:
        s = str(k).strip()
        if not s: continue
        if is_fk:
            if s.isdigit():
                q_final |= Q(oficina_id=int(s))
            q_final |= Q(oficina__nombre__icontains=s)
            try: q_final |= Q(oficina__codigo=s)
            except Exception: pass
        else:
            q_final |= Q(oficina__icontains=s)
            q_final |= Q(oficina__iexact=s)
    return q_final


def _parse_mes_yyyy_mm(raw: str):
    s = str(raw or "").strip()
    if not s:
        return None, None
    try:
        parts = s.split("-")
        if len(parts) != 2:
            return None, None
        y = int(parts[0])
        m = int(parts[1])
        if m < 1 or m > 12:
            return None, None
        first = date(y, m, 1)
        if m == 12:
            nxt = date(y + 1, 1, 1)
        else:
            nxt = date(y, m + 1, 1)
        return first, nxt
    except Exception:
        return None, None


def _parse_ymd(raw: str):
    s = str(raw or "").strip()
    if not s:
        return None
    return parse_date(s)


def _to_bool(v):
    s = str(v or "").strip().lower()
    return s in {"1", "true", "t", "yes", "y", "on", "si", "sí"}


def _to_int(v, default=None):
    try:
        if v is None or v == "":
            return default
        return int(str(v).strip())
    except Exception:
        return default


def _only_digits(s: str) -> str:
    return "".join([c for c in str(s or "") if c.isdigit()])


def _compania_nombre_robusto(poliza):
    try:
        if not poliza:
            return ""
        comp = getattr(poliza, "compania", None)
        if comp is None:
            cn = getattr(poliza, "compania_nombre", None)
            return str(cn or "").strip()
        if hasattr(comp, "nombre"):
            return str(getattr(comp, "nombre", "") or "").strip()
        return str(comp).strip()
    except Exception:
        return ""


class MedioCobroViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = MedioCobro.objects.all()

    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ["proveedor", "tipo", "activo"]
    search_fields = ["valor", "etiqueta", "titular_nombre"]
    ordering_fields = ["creado", "actualizado", "ultimo_uso", "usos_totales"]
    ordering = ["-activo", "etiqueta", "proveedor", "tipo"]

    def get_serializer_class(self):
        from .serializers import MedioCobroSerializer
        return MedioCobroSerializer

    @action(detail=True, methods=["post"], url_path="activar")
    def activar(self, request, pk=None):
        obj = self.get_object()
        obj.activo = True
        obj.save(update_fields=["activo"])
        return Response({"detail": "Activado"}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="desactivar")
    def desactivar(self, request, pk=None):
        obj = self.get_object()
        obj.activo = False
        obj.save(update_fields=["activo"])
        return Response({"detail": "Desactivado"}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="marcar-uso")
    def marcar_uso(self, request, pk=None):
        obj = self.get_object()
        mark = getattr(obj, "marcar_uso", None)
        if callable(mark):
            mark()
        else:
            from django.utils import timezone as _tz
            obj.ultimo_uso = _tz.now()
            obj.usos_totales = (obj.usos_totales or 0) + 1
            obj.save(update_fields=["ultimo_uso", "usos_totales"])
        return Response({"detail": "Uso registrado"}, status=status.HTTP_200_OK)


class PagoViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Pago.objects.all().select_related("poliza", "cuota")

    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ["poliza", "cuota", "cuota_nro", "metodo", "registrado_en_balance"]
    search_fields = ["poliza__numero_poliza"]
    ordering_fields = ["fecha", "monto"]
    ordering = ["-fecha", "poliza_id", "cuota_nro"]

    def get_serializer_class(self):
        from .serializers import PagoSerializer
        return PagoSerializer

    @action(detail=False, methods=["post"], url_path="registrar")
    def registrar_pago(self, request):
        try:
            result = registrar_pago_handler(request.data, request)
            if isinstance(result, Response):
                return result
            return Response(result, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path="enviar-alertas")
    def enviar_alertas_manual(self, request):
        user = request.user
        es_admin = user.is_superuser or (hasattr(user, 'perfil') and user.perfil.rol == 'ADMIN')
        
        if not es_admin:
            return Response(
                {"error": "Acceso denegado. Solo el Administrador puede disparar los mensajes recordatorios masivos."}, 
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            cantidad = ejecutar_alertas()
            return Response({"mensaje": f"{cantidad} alertas enviadas"}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # 🚀 ENDPOINT ACTUALIZADO: Efectividad con Filtros de Fecha para el Tablero
    @action(detail=False, methods=["get"], url_path="reporte-efectividad")
    def reporte_efectividad(self, request):
        oficina_keys = _get_seguridad_oficina_brute(request, request.query_params.get("oficina", ""))
        if "BLOQUEADO" in oficina_keys:
            return Response({"detail": "Acceso denegado"}, status=403)

        alertas_qs = AlertaEnviada.objects.select_related(
            "cuota", 
            "cuota__poliza", 
            "cuota__poliza__cliente"
        ).filter(enviada=True)

        if oficina_keys:
            alertas_qs = alertas_qs.filter(_build_oficina_q_from_keys(oficina_keys))

        # 🚀 APLICAR FILTROS DE FECHA
        desde = request.query_params.get("desde")
        hasta = request.query_params.get("hasta")
        anio = request.query_params.get("anio")
        mes = request.query_params.get("mes")

        if desde or hasta:
            d1 = _parse_ymd(desde) if desde else None
            d2 = _parse_ymd(hasta) if hasta else None
            if d1:
                alertas_qs = alertas_qs.filter(fecha__gte=d1)
            if d2:
                alertas_qs = alertas_qs.filter(fecha__lt=(d2 + timedelta(days=1)))
        elif anio and mes:
            try:
                y = int(anio)
                m = int(mes)
                first = date(y, m, 1)
                if m == 12:
                    nxt = date(y + 1, 1, 1)
                else:
                    nxt = date(y, m + 1, 1)
                alertas_qs = alertas_qs.filter(fecha__gte=first, fecha__lt=nxt)
            except Exception:
                pass
        elif anio:
             try:
                y = int(anio)
                first = date(y, 1, 1)
                nxt = date(y + 1, 1, 1)
                alertas_qs = alertas_qs.filter(fecha__gte=first, fecha__lt=nxt)
             except Exception:
                pass

        # 🚀 ORDENAR DE MÁS NUEVO A MÁS VIEJO
        alertas_qs = alertas_qs.order_by("-fecha")

        resultados_pagados = []
        resultados_pendientes = []
        
        horas_totales = 0
        pagos_recuperados = 0
        total_enviadas = 0
        
        hoy = timezone.now()

        for alerta in alertas_qs:
            total_enviadas += 1
            cli = alerta.cuota.poliza.cliente
            nombre_cliente = f"{getattr(cli, 'nombre', '')} {getattr(cli, 'apellido', '')}".strip()

            if alerta.cuota.pagado and alerta.cuota.pago_registrado_en:
                delta = alerta.cuota.pago_registrado_en - alerta.fecha
                if delta.total_seconds() >= 0:
                    horas_pasadas = delta.total_seconds() / 3600.0
                    resultados_pagados.append({
                        "alerta_id": alerta.id,
                        "tipo_mensaje": alerta.tipo,
                        "fecha_mensaje": timezone.localtime(alerta.fecha).strftime("%d/%m/%Y %H:%M"),
                        "fecha_pago": timezone.localtime(alerta.cuota.pago_registrado_en).strftime("%d/%m/%Y %H:%M"),
                        "horas_tardanza": round(horas_pasadas, 1),
                        "cuota_nro": alerta.cuota.cuota_nro,
                        "monto_recuperado": float(alerta.cuota.monto) if alerta.cuota.monto else 0.0,
                        "cliente": nombre_cliente,
                        "patente": alerta.cuota.poliza.patente
                    })
                    horas_totales += horas_pasadas
                    pagos_recuperados += 1
            else:
                delta_pendiente = hoy - alerta.fecha
                dias_pasados = delta_pendiente.total_seconds() / 86400.0
                resultados_pendientes.append({
                    "alerta_id": alerta.id,
                    "tipo_mensaje": alerta.tipo,
                    "fecha_mensaje": timezone.localtime(alerta.fecha).strftime("%d/%m/%Y %H:%M"),
                    "dias_sin_pagar": round(max(0, dias_pasados), 1),
                    "cuota_nro": alerta.cuota.cuota_nro,
                    "monto_adeudado": float(alerta.cuota.monto) if alerta.cuota.monto else 0.0,
                    "cliente": nombre_cliente,
                    "patente": alerta.cuota.poliza.patente
                })

        promedio_horas = round(horas_totales / pagos_recuperados, 1) if pagos_recuperados > 0 else 0
        tasa_conversion = round((pagos_recuperados / total_enviadas) * 100, 1) if total_enviadas > 0 else 0

        return Response({
            "kpis": {
                "total_mensajes_enviados": total_enviadas,
                "pagos_recuperados": pagos_recuperados,
                "tasa_conversion": f"{tasa_conversion}%",
                "tiempo_promedio_respuesta_horas": promedio_horas
            },
            "detalle_pagados": resultados_pagados,
            "detalle_pendientes": resultados_pendientes 
        }, status=status.HTTP_200_OK)


    @action(detail=False, methods=["get"], url_path="buscar-cliente")
    def buscar_cliente(self, request):
        dni_raw = (request.query_params.get("dni") or request.query_params.get("q") or "").strip()
        dni = _only_digits(dni_raw)

        if not dni:
            return Response({"detail": "Falta parámetro dni (solo números)."}, status=status.HTTP_400_BAD_REQUEST)

        # 🆕 Refrescamos el estado de las pólizas antes de buscar, para que las que ya
        # vencieron y están pagas pasen a "finalizada" (y muestren el botón Renovar).
        # Sin esto, el estado solo se actualiza al entrar al módulo de Pólizas.
        try:
            from polizas.views.poliza import auto_marcar_vencidas
            auto_marcar_vencidas()
        except Exception:
            pass

        pol_qs = (
            Poliza.objects.select_related("cliente", "oficina")
            .only(
                "id",
                "oficina",
                "compania",
                "patente",
                "marca",
                "modelo",
                "estado",
                "fecha_baja",
                "fecha_vencimiento",
                "cliente_id",
                "cliente__apellido",
                "cliente__nombre",
                "cliente__dni_cuit_cuil",
            )
            .filter(
                Q(cliente__dni_cuit_cuil__iexact=dni_raw)
                | Q(cliente__dni_cuit_cuil__iexact=dni)
                | Q(cliente__dni_cuit_cuil__icontains=dni)
            )
            # 🆕 Fecha de la ÚLTIMA cuota (la mayor): es el "fin real" que usa el sistema
            # para finalizar la póliza. La tolerancia de 3 días se mide contra esto,
            # no contra poliza.fecha_vencimiento (que puede no estar cargada).
            .annotate(
                ult_cuota_vto=Subquery(
                    Cuota.objects.filter(poliza_id=OuterRef("pk"))
                    .order_by("-fecha_vencimiento")
                    .values("fecha_vencimiento")[:1],
                    output_field=DateField(),
                )
            )
        )
        
        # 🚀 MULTI-TENANT: la búsqueda por DNI ve TODAS las oficinas. Cualquier
        # empleado puede atender y cobrar a cualquier cliente.
        if not request.user.is_authenticated:
            return Response({"detail": "Acceso denegado"}, status=403)

        pol_qs = pol_qs.order_by("-id")[:50]

        first = pol_qs.first()
        if not first or not getattr(first, "cliente", None):
            return Response({"detail": "No se encontró cliente con ese DNI."}, status=status.HTTP_404_NOT_FOUND)

        cli = first.cliente
        nombre_apellido = f"{(cli.apellido or '').strip()} {(cli.nombre or '').strip()}".strip()

        def get_ofi_str(ofi_obj):
            if not ofi_obj: return ""
            if hasattr(ofi_obj, 'nombre'): return str(ofi_obj.nombre)
            return str(ofi_obj)

        oficina_str = get_ofi_str(getattr(first, "oficina", None))
        try:
            counts = {}
            for p in pol_qs:
                o_str = get_ofi_str(getattr(p, "oficina", None))
                if not o_str:
                    continue
                counts[o_str] = counts.get(o_str, 0) + 1
            if counts:
                oficina_str = sorted(counts.items(), key=lambda x: (-x[1], str(x[0])))[0][0]
        except Exception:
            pass

        # 🆕 Marcar cuáles ya tienen una renovación hecha: existe una póliza "hija"
        #    que las apunta como origen. El front usa esto para NO ofrecer "Renovar"
        #    sobre una póliza que ya fue renovada (evita el 409 POLIZA_YA_RENOVADA).
        pol_ids = [p.id for p in pol_qs]
        renovadas_origen_ids = set(
            Poliza.objects.filter(poliza_origen_id__in=pol_ids)
            .values_list("poliza_origen_id", flat=True)
        )

        polizas = []
        for p in pol_qs:
            modelo_txt = f"{(getattr(p, 'marca', '') or '').strip()} {(getattr(p, 'modelo', '') or '').strip()}".strip()
            fecha_baja = getattr(p, "fecha_baja", None)
            # Fin real = última cuota; si por algún motivo no hay, caemos al campo de la póliza.
            fecha_fin = getattr(p, "ult_cuota_vto", None) or getattr(p, "fecha_vencimiento", None)
            polizas.append(
                {
                    "poliza_id":  p.id,
                    "compania":   _compania_nombre_robusto(p),
                    "patente":    getattr(p, "patente", "") or "",
                    "modelo":     modelo_txt,
                    "oficina":    get_ofi_str(getattr(p, "oficina", None)),
                    "estado":     getattr(p, "estado", "") or "",
                    "fecha_baja": fecha_baja.isoformat() if fecha_baja else None,
                    "fecha_fin":  fecha_fin.isoformat() if fecha_fin else None,
                    "tiene_renovacion": p.id in renovadas_origen_ids,
                }
            )

        return Response(
            {
                "cliente": {"dni": getattr(cli, "dni_cuit_cuil", "") or dni, "nombre_apellido": nombre_apellido, "oficina": oficina_str},
                "polizas": polizas,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"], url_path="buscar")
    def buscar(self, request):
        q = (request.query_params.get("q") or request.query_params.get("search") or "").strip()
        
        oficina_raw = (request.query_params.get("oficina") or "").strip()
        # 🚀 MULTI-TENANT: la búsqueda de pagos ve TODAS las oficinas. Cualquier
        # empleado puede buscar y cobrar a cualquier cliente (el ingreso se registra
        # en la oficina del que cobra, no en la de la póliza). Solo filtramos si se
        # pide una oficina explícita en la URL (?oficina=...), útil para el admin.
        if not request.user.is_authenticated:
            return Response({"detail": "Acceso denegado"}, status=403)
        oficina_keys = []
        if oficina_raw and oficina_raw.upper() != "ALL":
            _keys = _get_seguridad_oficina_brute(request, oficina_raw)
            if "BLOQUEADO" not in _keys:
                oficina_keys = _keys

        poliza_id = _to_int(request.query_params.get("poliza_id") or request.query_params.get("poliza"), None)
        ocultar_pagadas = _to_bool(request.query_params.get("ocultar_pagadas") or request.query_params.get("solo_pendientes"))

        pagado_raw = request.query_params.get("pagado")
        pagado_filter = None
        if pagado_raw not in (None, ""):
            pagado_filter = _to_bool(pagado_raw)

        vencidas = _to_bool(request.query_params.get("vencidas"))
        por_vencer = _to_bool(request.query_params.get("por_vencer"))
        por_vencer_dias = _to_int(request.query_params.get("por_vencer_dias"), 7)
        por_vencer_dias = max(1, min(365, por_vencer_dias or 7))

        ordering = (request.query_params.get("ordering") or "").strip()

        page_size = _to_int(request.query_params.get("page_size"), None)
        limit = _to_int(request.query_params.get("limit"), None)
        if page_size is None and limit is not None:
            page_size = limit
        if page_size is None:
            page_size = 150
        page_size = max(1, min(500, page_size))

        _PAGE_SIZE = page_size

        class _SearchPagination(PageNumberPagination):
            page_size = _PAGE_SIZE
            page_size_query_param = "page_size"
            max_page_size = 500

        total_cuotas_sq = Subquery(
            Cuota.objects.filter(poliza_id=OuterRef("poliza_id"))
            .order_by()
            .values("poliza_id")
            .annotate(mx=Max("cuota_nro"))
            .values("mx")[:1]
        )

        qs = (
            Cuota.objects.all()
            .select_related("poliza", "poliza__cliente")
            .annotate(total_cuotas=total_cuotas_sq)
            .only(
                "id",
                "cuota_nro",
                "monto",
                "pagado",
                "fecha_vencimiento",
                "fecha_pago",
                "pago_registrado_en",
                "forma_pago",
                "observaciones_pago",
                "ultima_observacion_pago",
                "poliza_id",
                "poliza__numero_poliza",
                "poliza__patente",
                "poliza__marca",
                "poliza__modelo",
                "poliza__cobertura",
                "poliza__oficina",
                "poliza__compania",
                "poliza__cantidad_cuotas",
                "poliza__cliente_id",
                "poliza__cliente__apellido",
                "poliza__cliente__nombre",
                "poliza__cliente__dni_cuit_cuil",
                "poliza__cliente__telefono",
            )
        )

        if poliza_id:
            qs = qs.filter(poliza_id=poliza_id)

        if pagado_filter is not None:
            qs = qs.filter(pagado=pagado_filter)
        elif ocultar_pagadas:
            qs = qs.filter(pagado=False)

        if oficina_keys:
            qs = qs.filter(_build_oficina_q_from_keys(oficina_keys))

        if (not poliza_id) and q:
            terminos = q.split()
            for t in terminos:
                qs = qs.filter(
                    Q(poliza__numero_poliza__icontains=t)
                    | Q(poliza__patente__icontains=t)
                    | Q(poliza__cliente__apellido__icontains=t)
                    | Q(poliza__cliente__nombre__icontains=t)
                    | Q(poliza__cliente__dni_cuit_cuil__icontains=t)
                )

        hoy = timezone.localdate()
        if vencidas:
            qs = qs.filter(pagado=False, fecha_vencimiento__lt=hoy)

        if por_vencer:
            hasta = hoy + timedelta(days=por_vencer_dias)
            qs = qs.filter(pagado=False, fecha_vencimiento__gte=hoy, fecha_vencimiento__lte=hasta)

        if ordering in ("vencimiento", "fecha_vencimiento"):
            qs = qs.order_by("fecha_vencimiento", "poliza_id", "cuota_nro", "id")
        elif ordering in ("-vencimiento", "-fecha_vencimiento"):
            qs = qs.order_by("-fecha_vencimiento", "poliza_id", "cuota_nro", "id")
        elif ordering == "poliza":
            qs = qs.order_by("poliza__numero_poliza", "cuota_nro", "id")
        elif ordering == "-poliza":
            qs = qs.order_by("-poliza__numero_poliza", "cuota_nro", "id")
        elif ordering == "cuota":
            qs = qs.order_by("cuota_nro", "poliza_id", "id")
        elif ordering == "-cuota":
            qs = qs.order_by("-cuota_nro", "poliza_id", "id")
        else:
            qs = qs.order_by("poliza_id", "cuota_nro", "id")

        from .serializers import CuotaFlatSerializer

        paginator = _SearchPagination()
        page = paginator.paginate_queryset(qs, request, view=self)
        if page is not None:
            ser = CuotaFlatSerializer(page, many=True, context={"only_cuotas": bool(poliza_id)})
            return paginator.get_paginated_response(ser.data)

        ser = CuotaFlatSerializer(qs[:page_size], many=True)
        return Response({"count": len(ser.data), "results": ser.data}, status=status.HTTP_200_OK)

    # ===========================================================
    # 🚀 VERIFICACIÓN DE MICAELA — endpoints nuevos
    # ===========================================================

    @action(detail=True, methods=["post"], url_path="cambiar_estado_verificacion")
    def cambiar_estado_verificacion(self, request, pk=None):
        """
        POST /api/pagos/{id}/cambiar_estado_verificacion/
        Body: {"estado_verificacion": "verificado", "nota": "opcional"}

        Estados válidos:
          - pendiente
          - verificado
          - falta_emitir
          - pago_post_baja
          - avisar_vendedor
          - revisar_mariano
        """
        from .models import ESTADO_VERIFICACION_CHOICES

        pago = self.get_object()
        nuevo_estado = (request.data.get("estado_verificacion") or "").strip()
        nota = (request.data.get("nota") or "").strip()

        estados_validos = {k for k, _ in ESTADO_VERIFICACION_CHOICES}
        if nuevo_estado not in estados_validos:
            return Response(
                {"detail": f"Estado inválido. Debe ser uno de: {', '.join(estados_validos)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        pago.estado_verificacion = nuevo_estado
        if nota:
            pago.verificacion_nota = nota
        pago.verificado_por = request.user if request.user.is_authenticated else None
        pago.verificado_en = timezone.now()
        pago.save(update_fields=[
            "estado_verificacion",
            "verificacion_nota",
            "verificado_por",
            "verificado_en",
            "actualizado",
        ])

        return Response({
            "id": pago.id,
            "estado_verificacion": pago.estado_verificacion,
            "verificacion_nota": pago.verificacion_nota,
            "verificado_por": getattr(pago.verificado_por, "username", None),
            "verificado_en": pago.verificado_en,
            "requiere_atencion": pago.requiere_atencion,
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="atencion_count")
    def atencion_count(self, request):
        """
        GET /api/pagos/atencion_count/?oficina=ALL
        Devuelve cantidad de pagos en estados de atención.
        Si el user no es admin, solo cuenta los de su oficina.
        """
        from .models import Pago, ESTADOS_ATENCION
        from django.db.models import Count

        qs = Pago.objects.filter(estado_verificacion__in=ESTADOS_ATENCION)

        user = request.user
        perfil = getattr(user, "perfil", None)
        rol = getattr(perfil, "rol", None) if perfil else None

        if rol != "ADMIN":
            ofi_propia = getattr(perfil, "oficina_id", None) if perfil else None
            if ofi_propia:
                qs = qs.filter(poliza__oficina_id=ofi_propia)
        else:
            oficina_param = (request.query_params.get("oficina") or "").strip()
            if oficina_param and oficina_param.upper() != "ALL":
                qs = qs.filter(poliza__oficina_id=oficina_param)

        total = qs.count()

        por_estado = dict(
            qs.values_list("estado_verificacion")
              .annotate(c=Count("id"))
              .values_list("estado_verificacion", "c")
        )

        por_oficina = {}
        if rol == "ADMIN":
            por_oficina = dict(
                qs.values_list("poliza__oficina__nombre")
                  .annotate(c=Count("id"))
                  .values_list("poliza__oficina__nombre", "c")
            )

        return Response({
            "total": total,
            "por_estado": por_estado,
            "por_oficina": por_oficina,
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="atencion_list")
    def atencion_list(self, request):
        """
        GET /api/pagos/atencion_list/
        Lista los pagos en atención (máximo 50, ordenados por más reciente).
        """
        from .models import Pago, ESTADOS_ATENCION

        qs = Pago.objects.filter(estado_verificacion__in=ESTADOS_ATENCION)
        qs = qs.select_related("poliza", "poliza__cliente").order_by("-registrado_en")

        user = request.user
        perfil = getattr(user, "perfil", None)
        rol = getattr(perfil, "rol", None) if perfil else None

        if rol != "ADMIN":
            ofi_propia = getattr(perfil, "oficina_id", None) if perfil else None
            if ofi_propia:
                qs = qs.filter(poliza__oficina_id=ofi_propia)
        else:
            oficina_param = (request.query_params.get("oficina") or "").strip()
            if oficina_param and oficina_param.upper() != "ALL":
                qs = qs.filter(poliza__oficina_id=oficina_param)

        qs = qs[:50]
        serializer = PagoSerializer(qs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class CuotaViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Cuota.objects.all().select_related("poliza", "poliza__cliente")

    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ["poliza", "pagado", "fecha_vencimiento"]

    search_fields = [
        "poliza__numero_poliza",
        "poliza__patente",
        "poliza__cliente__apellido",
        "poliza__cliente__nombre",
        "poliza__cliente__dni_cuit_cuil",
    ]

    ordering_fields = ["fecha_vencimiento", "cuota_nro", "monto"]
    ordering = ["poliza_id", "cuota_nro"]

    def get_serializer_class(self):
        from .serializers import CuotaSerializer, CuotaPagoHistorialSerializer
        if getattr(self, "action", "") == "historial_pagos":
            return CuotaPagoHistorialSerializer
        return CuotaSerializer

    @action(detail=True, methods=["patch"], url_path="pagar")
    def pagar(self, request, pk=None):
        cuota: Cuota = self.get_object()

        if cuota.pagado:
            return Response({"detail": "La cuota ya figura como pagada."}, status=status.HTTP_409_CONFLICT)

        ahora = timezone.now()

        fecha_pago_raw = request.data.get("fecha_pago")
        if fecha_pago_raw:
            if isinstance(fecha_pago_raw, str):
                _fecha_pago = parse_date(fecha_pago_raw)
                if not _fecha_pago:
                    return Response({"fecha_pago": "Formato inválido. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
                fecha_pago = _fecha_pago
            else:
                fecha_pago = fecha_pago_raw
        else:
            fecha_pago = ahora.date()

        forma_pago = request.data.get("forma_pago")
        if forma_pago and forma_pago not in ("efectivo", "transferencia"):
            return Response({"forma_pago": 'Valor inválido. Use "efectivo" o "transferencia".'}, status=status.HTTP_400_BAD_REQUEST)

        metodo = request.data.get("metodo")
        if metodo in ("mercado_pago", "tarjeta"):
            metodo = "transferencia"
        if metodo and metodo not in ("efectivo", "transferencia"):
            return Response({"metodo": 'Valor inválido. Use "efectivo" o "transferencia".'}, status=status.HTTP_400_BAD_REQUEST)

        if not forma_pago and metodo:
            forma_pago = "efectivo" if metodo == "efectivo" else "transferencia"

        monto_raw = request.data.get("monto")
        monto_decimal = None
        if monto_raw not in (None, ""):
            try:
                monto_decimal = Decimal(str(monto_raw))
                if monto_decimal < 0:
                    return Response({"monto": "Debe ser un número positivo."}, status=status.HTTP_400_BAD_REQUEST)
            except (InvalidOperation, TypeError, ValueError):
                return Response({"monto": "Monto inválido."}, status=status.HTTP_400_BAD_REQUEST)

        observaciones = request.data.get("observaciones", "")
        registrar_en_balance = request.data.get("registrar_en_balance", True)

        # 🆕 Quién cobró (opcional; si no lo mandan, no rompe nada).
        responsable_id = request.data.get("responsable_empleado") or request.data.get("responsable_empleado_id")
        responsable_obj = None
        if responsable_id:
            responsable_obj = Empleado.objects.filter(id=responsable_id).first()
            if not responsable_obj:
                return Response({"responsable_empleado": "Empleado no encontrado."}, status=status.HTTP_400_BAD_REQUEST)
        responsable_nombre = (request.data.get("responsable_nombre") or "").strip() or getattr(responsable_obj, "nombre", "")

        with transaction.atomic():
            txt_obs = str(observaciones or "").strip()
            if txt_obs:
                try:
                    cuota.observaciones_pago = txt_obs
                    cuota.ultima_observacion_pago = txt_obs
                except Exception:
                    pass

            cuota.pagado = True
            cuota.fecha_pago = fecha_pago
            cuota.pago_registrado_en = ahora

            if forma_pago:
                cuota.forma_pago = forma_pago
            if monto_decimal is not None:
                cuota.monto = monto_decimal

            update_fields = ["pagado", "fecha_pago", "pago_registrado_en", "forma_pago", "monto"]
            if txt_obs:
                update_fields += ["observaciones_pago", "ultima_observacion_pago"]
            if responsable_obj is not None or responsable_nombre:
                cuota.responsable_empleado = responsable_obj
                cuota.responsable_nombre = responsable_nombre
                update_fields += ["responsable_empleado", "responsable_nombre"]

            cuota.save(update_fields=update_fields)

            pago_defaults = {
                "fecha": fecha_pago,
                "monto": monto_decimal if monto_decimal is not None else cuota.monto,
                "metodo": (metodo if metodo else (forma_pago if forma_pago in ("efectivo", "transferencia") else "transferencia")),
                "observaciones": observaciones,
                # 🆕 Detalle de la transferencia — ya viajaba en el body del wizard,
                #    faltaba guardarlo en el Pago (por eso nunca llegaba a Balances).
                "destino_cuenta": request.data.get("destino_cuenta") or request.data.get("medio_cobro_valor") or "",
                "enviado_por": request.data.get("enviado_por") or "",
                "cuit_remitente": request.data.get("cuit_remitente") or "",
                "nro_operacion": request.data.get("nro_operacion") or "",
                # 🆕 Quién cobró.
                "responsable_empleado": responsable_obj,
                "responsable_nombre": responsable_nombre,
            }
            pago, creado = Pago.objects.get_or_create(
                poliza=cuota.poliza,
                cuota=cuota,
                cuota_nro=cuota.cuota_nro,
                defaults=pago_defaults,
            )
            if not creado:
                for k, v in pago_defaults.items():
                    setattr(pago, k, v)
                pago.save()

            medio_id = request.data.get("medio_cobro_id")
            medio_valor = request.data.get("medio_cobro_valor") or request.data.get("destino_cuenta")
            try:
                medio = None
                if medio_id:
                    medio = MedioCobro.objects.filter(id=medio_id).first()
                if not medio and medio_valor:
                    medio = MedioCobro.objects.filter(valor=medio_valor).first() or MedioCobro.objects.filter(etiqueta=medio_valor).first()
                if medio:
                    mark = getattr(medio, "marcar_uso", None)
                    if callable(mark):
                        mark()
                    else:
                        medio.ultimo_uso = timezone.now()
                        medio.usos_totales = (medio.usos_totales or 0) + 1
                        medio.save(update_fields=["ultimo_uso", "usos_totales"])
            except Exception:
                pass

            if registrar_en_balance and not getattr(pago, "registrado_en_balance", False):
                pago.registrado_en_balance = True
                pago.save(update_fields=["registrado_en_balance"])

                poliza_obj = cuota.poliza
                ofi_code = str(getattr(poliza_obj.oficina, 'id', poliza_obj.oficina or ""))
                
                forma_balance = "efectivo" if (pago.metodo == "efectivo") else "transferencia"

                cliente_nombre = ""
                try:
                    c = poliza_obj.cliente
                    if c:
                        nom = (getattr(c, "nombre", "") or "").strip()
                        ape = (getattr(c, "apellido", "") or "").strip()
                        if ape and nom:
                            cliente_nombre = f"{ape}, {nom}"
                        else:
                            cliente_nombre = ape or nom
                except Exception:
                    pass

                # ── Datos de transferencia del wizard ──────────────
                enviado_por    = request.data.get("enviado_por") or cliente_nombre or ""
                destino_cuenta = request.data.get("destino_cuenta") or request.data.get("medio_cobro_valor") or ""
                cuit_remitente = request.data.get("cuit_remitente") or ""
                nro_operacion  = request.data.get("nro_operacion")  or ""

                # Observaciones con trazabilidad completa
                obs_partes = []
                if str(observaciones or "").strip(): obs_partes.append(str(observaciones).strip())
                if cuit_remitente: obs_partes.append(f"CUIT: {cuit_remitente}")
                if nro_operacion:  obs_partes.append(f"Op: {nro_operacion}")
                obs_completo = " | ".join(obs_partes) or ""

                ingreso_data = {
                    "monto":          pago.monto,
                    "categoria":      "Cobro de Cuota",
                    "forma_pago":     forma_balance,
                    "pagado_por":     enviado_por,
                    "billetera":      destino_cuenta,
                    "cuit_remitente": cuit_remitente,
                    "nro_operacion":  nro_operacion,
                    "observaciones":  obs_completo,
                    "descripcion":    f"Pago cuota {pago.cuota_nro} - Póliza {poliza_obj.numero_poliza}"
                }
                
                try:
                    ingreso_data["fecha"] = fecha_pago
                except Exception: pass
                
                try:
                    # ✅ CORRECCIÓN: Le agregamos _id para que Django entienda que es el número identificador
                    ingreso_data["oficina_id"] = ofi_code
                except Exception: pass
                
                try:
                    ingreso_data["usuario"] = request.user
                except Exception: pass
                
                BalanceIngreso.objects.create(**ingreso_data)

            # 🚀===================================================
            # 🚀 LÓGICA DE REACTIVACIÓN INTELIGENTE
            # ===================================================
            poliza: Poliza = cuota.poliza
            hoy = timezone.localdate()
            hay_vencidas = poliza.cuotas.filter(pagado=False, fecha_vencimiento__lt=hoy).exists()
            
            estado_actual = str(getattr(poliza, "estado", "")).strip().upper()

            if estado_actual in ("CANCELADA", "ANULADA"):
                pass
            elif estado_actual == "VENCIDA" and not hay_vencidas:
                poliza.estado = "activa"
                poliza.save(update_fields=["estado"])
            elif not hay_vencidas and estado_actual != "ACTIVA" and estado_actual not in ("CANCELADA", "ANULADA", "BAJA"):
                poliza.estado = "activa"
                poliza.save(update_fields=["estado"])

            # Si la póliza está dada de baja y se cobró → marcar en verificación
            if estado_actual in ("BAJA", "BAJA_RECIENTE"):
                poliza.estado = "en_verificacion"
                poliza.save(update_fields=["estado"])
            # ===================================================

        # 🆕 Agradecimiento + link al portal (no rompe el pago si falla). Cubre los dos botones.
        try:
            _enviar_gracias_portal(cuota.poliza)
        except Exception as e:
            print(f"[cuota.pagar] WhatsApp de gracias falló: {e}")

        serializer = self.get_serializer(cuota)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post", "patch"], url_path="cambiar-fecha")
    def cambiar_fecha(self, request, pk=None):
        cuota = self.get_object()
        nueva_fecha_raw = request.data.get("nueva_fecha")
        
        if not nueva_fecha_raw:
            return Response({"nueva_fecha": "Este campo es requerido."}, status=status.HTTP_400_BAD_REQUEST)

        nueva_fecha = _parse_ymd(nueva_fecha_raw)
        if not nueva_fecha:
            return Response({"nueva_fecha": "Formato inválido. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

        ajustar_siguientes = _to_bool(request.data.get("ajustar_siguientes", False))

        def add_months(sourcedate, months):
            month = sourcedate.month - 1 + months
            year = sourcedate.year + month // 12
            month = month % 12 + 1
            day = min(sourcedate.day, calendar.monthrange(year, month)[1])
            return date(year, month, day)

        with transaction.atomic():
            cuota.fecha_vencimiento = nueva_fecha
            cuota.save(update_fields=["fecha_vencimiento"])
            modificadas = 1

            if ajustar_siguientes:
                siguientes = Cuota.objects.filter(
                    poliza=cuota.poliza,
                    cuota_nro__gt=cuota.cuota_nro
                ).order_by("cuota_nro")

                meses_a_sumar = 1
                for sig in siguientes:
                    sig.fecha_vencimiento = add_months(nueva_fecha, meses_a_sumar)
                    sig.save(update_fields=["fecha_vencimiento"])
                    meses_a_sumar += 1
                    modificadas += 1

        return Response({
            "detail": "Fechas actualizadas correctamente.",
            "cuotas_modificadas": modificadas
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="factura")
    def factura(self, request, pk=None):
        cuota = self.get_object()
        
        oficina_keys = _get_seguridad_oficina_brute(request)
        if "BLOQUEADO" in oficina_keys:
            return Response({"detail": "No tienes acceso a esta factura."}, status=403)
            
        pdf_filelike = generar_factura_pdf(cuota)
        return FileResponse(
            pdf_filelike,
            as_attachment=True,
            filename=f"factura_cuota_{cuota.id}.pdf",
            content_type="application/pdf",
        )

    @action(detail=False, methods=["get"], url_path="a-vencer")
    def cuotas_a_vencer(self, request):
        hoy = timezone.localdate()
        hitos = {hoy - timedelta(days=30), hoy - timedelta(days=7), hoy - timedelta(days=3), hoy, hoy + timedelta(days=3)}
        
        oficina_keys = _get_seguridad_oficina_brute(request, request.query_params.get("oficina", ""))
        if "BLOQUEADO" in oficina_keys:
            return Response({"detail": "Acceso denegado"}, status=403)
            
        qs = Cuota.objects.filter(pagado=False, fecha_vencimiento__in=hitos).select_related("poliza", "poliza__cliente")
        
        if oficina_keys:
            qs = qs.filter(_build_oficina_q_from_keys(oficina_keys))
            
        qs = qs.order_by("fecha_vencimiento", "poliza_id", "cuota_nro")
        
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def _historial_pagos_base_queryset(self):
        last_pago_registrado_en_sq = Subquery(
            Pago.objects.filter(
                poliza_id=OuterRef("poliza_id"),
                cuota_nro=OuterRef("cuota_nro"),
            )
            .order_by("-registrado_en", "-id")
            .values("registrado_en")[:1]
        )

        return (
            Cuota.objects.filter(pagado=True)
            .select_related("poliza", "poliza__cliente")
            .annotate(
                pago_ts=Coalesce(F("pago_registrado_en"), last_pago_registrado_en_sq)
            )
            .only(
                "id",
                "fecha_pago",
                "pago_registrado_en",
                "cuota_nro",
                "monto",
                "forma_pago",
                "observaciones_pago",
                "ultima_observacion_pago",
                "poliza_id",
                "poliza__numero_poliza",
                "poliza__patente",
                "poliza__marca",
                "poliza__modelo",
                "poliza__oficina",
                "poliza__compania",
                "poliza__cantidad_cuotas",
                "poliza__cliente_id",
                "poliza__cliente__apellido",
                "poliza__cliente__nombre",
                "poliza__cliente__dni_cuit_cuil",
                "poliza__cliente__telefono",
            )
        )

    def _apply_historial_filters(self, qs, request, *, mes="", dia="", desde="", hasta="", oficina="", search=""):
        mes = (mes or "").strip()
        dia = (dia or "").strip()
        desde = (desde or "").strip()
        hasta = (hasta or "").strip()
        search = (search or "").strip()
        
        oficina_keys = _get_seguridad_oficina_brute(request, (oficina or "").strip())
        if "BLOQUEADO" in oficina_keys:
            raise ValueError("Acceso denegado a esta oficina.")

        if dia:
            d = _parse_ymd(dia)
            if not d:
                raise ValueError("Parámetro 'dia' inválido. Use YYYY-MM-DD (ej: 2026-02-10).")
            qs = qs.filter(fecha_pago=d)

        elif desde or hasta:
            d1 = _parse_ymd(desde) if desde else None
            d2 = _parse_ymd(hasta) if hasta else None
            if desde and not d1:
                raise ValueError("Parámetro 'desde' inválido. Use YYYY-MM-DD.")
            if hasta and not d2:
                raise ValueError("Parámetro 'hasta' inválido. Use YYYY-MM-DD.")
            if d1 and d2 and d2 < d1:
                raise ValueError("Rango inválido: 'hasta' no puede ser menor que 'desde'.")

            if d1:
                qs = qs.filter(fecha_pago__gte=d1)
            if d2:
                qs = qs.filter(fecha_pago__lt=(d2 + timedelta(days=1)))

        elif mes:
            first, nxt = _parse_mes_yyyy_mm(mes)
            if not first:
                raise ValueError("Parámetro 'mes' inválido. Use YYYY-MM (ej: 2026-02).")
            qs = qs.filter(fecha_pago__gte=first, fecha_pago__lt=nxt)

        if oficina_keys:
            qs = qs.filter(_build_oficina_q_from_keys(oficina_keys))

        if search:
            terminos = search.split()
            for t in terminos:
                qs = qs.filter(
                    Q(poliza__numero_poliza__icontains=t)
                    | Q(poliza__patente__icontains=t)
                    | Q(poliza__cliente__apellido__icontains=t)
                    | Q(poliza__cliente__nombre__icontains=t)
                    | Q(poliza__cliente__dni_cuit_cuil__icontains=t)
                )

        return qs

    def _apply_historial_ordering(self, qs, ordering_raw: str):
        ordering = (ordering_raw or "-fecha_pago").strip()
        allowed = {"fecha_pago", "-fecha_pago", "monto", "-monto", "cuota_nro", "-cuota_nro"}
        if ordering not in allowed:
            ordering = "-fecha_pago"

        if ordering == "monto":
            return qs.order_by("monto", "poliza_id", "cuota_nro")
        if ordering == "-monto":
            return qs.order_by("-monto", "-fecha_pago", "poliza_id", "cuota_nro")
        if ordering == "cuota_nro":
            return qs.order_by("cuota_nro", "-fecha_pago", "poliza_id")
        if ordering == "-cuota_nro":
            return qs.order_by("-cuota_nro", "-fecha_pago", "poliza_id")

        return qs.order_by("-fecha_pago", "poliza_id", "cuota_nro")

    def _render_historial_csv(self, qs, filename="historial_pagos.csv"):
        out = StringIO()
        writer = csv.writer(out)

        writer.writerow(
            [
                "fecha_pago",
                "cuota_nro",
                "monto",
                "forma_pago",
                "numero_poliza",
                "patente",
                "compania",
                "oficina",
                "cliente_apellido",
                "cliente_nombre",
                "cliente_dni",
                "cliente_telefono",
            ]
        )

        for c in qs.iterator(chunk_size=2000):
            pol = getattr(c, "poliza", None)
            cli = getattr(pol, "cliente", None) if pol else None

            writer.writerow(
                [
                    getattr(c, "fecha_pago", "") or "",
                    getattr(c, "cuota_nro", "") or "",
                    getattr(c, "monto", "") or "",
                    getattr(c, "forma_pago", "") or "",
                    getattr(pol, "numero_poliza", "") if pol else "",
                    getattr(pol, "patente", "") if pol else "",
                    _compania_nombre_robusto(pol),
                    getattr(pol, "oficina", "") if pol else "",
                    getattr(cli, "apellido", "") if cli else "",
                    getattr(cli, "nombre", "") if cli else "",
                    getattr(cli, "dni_cuit_cuil", "") if cli else "",
                    getattr(cli, "telefono", "") if cli else "",
                ]
            )

        resp = HttpResponse(out.getvalue(), content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    def _render_historial_pdf(self, qs, filename="historial_pagos.pdf", titulo="Historial de pagos"):
        buff = BytesIO()
        pdf = canvas.Canvas(buff, pagesize=landscape(A4))
        width, height = landscape(A4)

        left = 12 * mm
        top = height - 12 * mm
        line_h = 6.2 * mm

        pdf.setTitle(filename)

        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(left, top, titulo)
        pdf.setFont("Helvetica", 9)
        pdf.drawString(left, top - 8 * mm, f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}")

        y = top - 16 * mm

        cols = [
            ("Fecha", 22 * mm),
            ("Asegurado", 60 * mm),
            ("DNI", 26 * mm),
            ("Patente", 24 * mm),
            ("Póliza", 26 * mm),
            ("Compañía", 40 * mm),
            ("Oficina", 22 * mm),
            ("Importe", 24 * mm),
            ("Medio", 26 * mm),
        ]

        def draw_row(values, y, bold=False):
            x = left
            pdf.setFont("Helvetica-Bold" if bold else "Helvetica", 9)
            for (txt, w), v in zip(cols, values):
                s = str(v or "")
                if len(s) > 38 and w <= 60 * mm:
                    s = s[:35] + "…"
                pdf.drawString(x, y, s)
                x += w

        draw_row([t for (t, _) in cols], y, bold=True)
        y -= line_h
        pdf.setLineWidth(0.4)
        pdf.line(left, y + 2 * mm, width - left, y + 2 * mm)

        for item in qs.iterator(chunk_size=2000):
            if y < 12 * mm:
                pdf.showPage()
                pdf.setFont("Helvetica-Bold", 14)
                pdf.drawString(left, top, titulo)
                pdf.setFont("Helvetica", 9)
                pdf.drawString(left, top - 8 * mm, f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}")
                y = top - 16 * mm
                draw_row([t for (t, _) in cols], y, bold=True)
                y -= line_h
                pdf.line(left, y + 2 * mm, width - left, y + 2 * mm)

            pol = getattr(item, "poliza", None)
            cli = getattr(pol, "cliente", None) if pol else None

            fecha = getattr(item, "fecha_pago", None) or ""
            if fecha:
                try:
                    fecha = fecha.strftime("%d/%m/%Y")
                except Exception:
                    fecha = str(fecha)

            ape = getattr(cli, "apellido", "") if cli else ""
            nom = getattr(cli, "nombre", "") if cli else ""
            asegurado = f"{ape}, {nom}".strip(", ").strip()

            dni = getattr(cli, "dni_cuit_cuil", "") if cli else ""
            patente = getattr(pol, "patente", "") if pol else ""
            numero_poliza = getattr(pol, "numero_poliza", "") if pol else ""
            compania = _compania_nombre_robusto(pol)
            oficina = getattr(pol, "oficina", "") if pol else ""
            monto = getattr(item, "monto", "") or ""
            medio = getattr(item, "forma_pago", "") or ""

            draw_row([fecha, asegurado, dni, patente, numero_poliza, compania, oficina, monto, medio], y, bold=False)
            y -= line_h

        pdf.save()
        buff.seek(0)
        return FileResponse(buff, as_attachment=True, filename=filename, content_type="application/pdf")

    def _render_historial_xlsx(self, qs, filename="historial_pagos.xlsx", titulo="Historial de pagos"):
        """
        Excel profesional con TABLA REAL de openpyxl (objeto Table, con flechitas
        de filtro/ordenamiento nativas en cada header).

        Columnas (8): Apellido y Nombre · Patente · Compañía · Fecha de pago ·
                      Cuota · Oficina · Medio · Importe
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        wb = Workbook()
        ws = wb.active
        ws.title = "Historial de Pagos"

        # ── Helpers locales ─────────────────────────────────────────────────
        def _cuota_label(c):
            """Devuelve '1/3', '3/5', etc. Si no hay total, devuelve solo el nro."""
            nro = getattr(c, "cuota_nro", None)
            if not nro:
                return ""
            pol = getattr(c, "poliza", None)
            total = getattr(pol, "cantidad_cuotas", None) if pol else None
            if not total:
                return str(nro)
            return f"{nro}/{total}"

        def _apellido_nombre(cli):
            if not cli:
                return ""
            ape = (getattr(cli, "apellido", "") or "").strip()
            nom = (getattr(cli, "nombre", "") or "").strip()
            if ape and nom:
                return f"{ape}, {nom}"
            return ape or nom or ""

        def _fecha_pago_str(c):
            ts = getattr(c, "pago_ts", None) or getattr(c, "pago_registrado_en", None)
            if ts:
                try:
                    return timezone.localtime(ts).strftime("%d/%m/%Y %H:%M")
                except Exception:
                    pass
            fp = getattr(c, "fecha_pago", None)
            if fp:
                try:
                    return fp.strftime("%d/%m/%Y")
                except Exception:
                    return str(fp)
            return ""

        def _medio(c):
            f = (getattr(c, "forma_pago", "") or "").upper().strip()
            if f == "EFECTIVO":
                return "Efectivo"
            if f == "TRANSFERENCIA":
                return "Transferencia"
            return f or "—"

        def _oficina_nombre(pol):
            if not pol:
                return ""
            ofi = getattr(pol, "oficina", None)
            if not ofi:
                return ""
            # Si es FK a usuarios.Oficina, tiene .nombre
            nombre = getattr(ofi, "nombre", None)
            if nombre:
                return str(nombre).strip()
            return str(ofi).strip()

        # ── Estilos ─────────────────────────────────────────────────────────
        title_font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
        title_fill = PatternFill("solid", fgColor="1E3A8A")
        title_align = Alignment(horizontal="left", vertical="center", indent=1)

        subtitle_font = Font(color="475569", size=10, italic=True, name="Calibri")
        subtitle_align = Alignment(horizontal="left", vertical="center", indent=1)

        header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        header_fill = PatternFill("solid", fgColor="1E3A8A")
        header_align = Alignment(horizontal="center", vertical="center")

        thin = Side(border_style="thin", color="E2E8F0")
        cell_border = Border(top=thin, bottom=thin, left=thin, right=thin)
        cell_font = Font(color="1E293B", size=10, name="Calibri")
        cell_align_left = Alignment(horizontal="left", vertical="center", indent=1)
        cell_align_right = Alignment(horizontal="right", vertical="center", indent=1)
        cell_align_center = Alignment(horizontal="center", vertical="center")

        total_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        total_fill = PatternFill("solid", fgColor="064E3B")
        total_align = Alignment(horizontal="right", vertical="center", indent=1)

        # ── Estructura ──────────────────────────────────────────────────────
        HEADERS = [
            "Apellido y Nombre",
            "Patente",
            "Compañía",
            "Fecha de pago",
            "Cuota",
            "Oficina",
            "Medio",
            "Importe",
        ]
        n_cols = len(HEADERS)
        last_col_letter = get_column_letter(n_cols)

        # Fila 1: Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        c = ws.cell(row=1, column=1, value=titulo)
        c.font = title_font
        c.fill = title_fill
        c.alignment = title_align
        ws.row_dimensions[1].height = 28

        # Fila 2: Subtítulo
        total_count = qs.count() if hasattr(qs, "count") else 0
        generado_str = timezone.localtime().strftime("%d/%m/%Y %H:%M")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        c = ws.cell(row=2, column=1, value=f"Generado: {generado_str}   ·   Total de pagos: {total_count}")
        c.font = subtitle_font
        c.alignment = subtitle_align
        ws.row_dimensions[2].height = 18

        # Fila 3 vacía
        ws.row_dimensions[3].height = 6

        # Fila 4: Headers
        header_row = 4
        for col_idx, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = cell_border
        ws.row_dimensions[header_row].height = 24

        # Filas de datos
        row_alt_fill = PatternFill("solid", fgColor="F8FAFC")
        total_general = 0.0
        current_row = header_row

        for c_obj in qs.iterator(chunk_size=2000):
            current_row += 1
            pol = getattr(c_obj, "poliza", None)
            cli = getattr(pol, "cliente", None) if pol else None

            apellido_nombre = _apellido_nombre(cli)
            patente = (getattr(pol, "patente", "") or "").strip().upper() if pol else ""
            compania = _compania_nombre_robusto(pol) if pol else ""
            fecha_pago = _fecha_pago_str(c_obj)
            cuota_lbl = _cuota_label(c_obj)
            oficina = _oficina_nombre(pol)
            medio = _medio(c_obj)
            try:
                monto = float(getattr(c_obj, "monto", 0) or 0)
            except (TypeError, ValueError):
                monto = 0.0
            total_general += monto

            valores = [
                apellido_nombre or "—",
                patente or "—",
                compania or "—",
                fecha_pago or "—",
                cuota_lbl or "—",
                oficina or "—",
                medio,
                monto,
            ]

            is_alt = (current_row - header_row) % 2 == 0
            for col_idx, val in enumerate(valores, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = cell_font
                cell.border = cell_border
                if is_alt:
                    cell.fill = row_alt_fill

                if col_idx == 8:
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = cell_align_right
                elif col_idx in (2, 5, 7):
                    cell.alignment = cell_align_center
                else:
                    cell.alignment = cell_align_left

        data_last_row = current_row
        has_data = data_last_row > header_row

        # Total general
        if has_data:
            total_row = data_last_row + 1
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
            c = ws.cell(row=total_row, column=1, value="TOTAL GENERAL")
            c.font = total_font
            c.fill = total_fill
            c.alignment = total_align
            c = ws.cell(row=total_row, column=8, value=total_general)
            c.font = total_font
            c.fill = total_fill
            c.alignment = total_align
            c.number_format = '"$"#,##0.00'
            ws.row_dimensions[total_row].height = 22

        # 🚀 TABLA REAL de Excel (con flechitas de filtro nativas)
        if has_data:
            table_ref = f"A{header_row}:{last_col_letter}{data_last_row}"
            table = Table(displayName="HistorialPagos", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

        # Anchos de columna
        anchos = [32, 14, 26, 20, 10, 18, 16, 16]
        for i, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze panes (congela título + header)
        ws.freeze_panes = "A5"

        # Respuesta HTTP
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=["get"], url_path="pagos")
    def historial_pagos(self, request):
        mes = (request.query_params.get("mes") or "").strip()
        dia = (request.query_params.get("dia") or "").strip()
        desde = (request.query_params.get("desde") or "").strip()
        hasta = (request.query_params.get("hasta") or "").strip()
        oficina = (request.query_params.get("oficina") or "").strip()
        search = (request.query_params.get("search") or request.query_params.get("q") or "").strip()
        ordering = (request.query_params.get("ordering") or "-fecha_pago").strip()

        export = (request.query_params.get("export") or "").strip().lower()
        all_flag = _to_bool(request.query_params.get("all") or request.query_params.get("todos"))

        qs = self._historial_pagos_base_queryset()

        try:
            qs = self._apply_historial_filters(qs, request, mes=mes, dia=dia, desde=desde, hasta=hasta, oficina=oficina, search=search)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        qs = self._apply_historial_ordering(qs, ordering)

        if export == "csv":
            filename = "historial_pagos.csv"
            if dia:
                filename = f"pagos_{dia}.csv"
            elif desde or hasta:
                filename = f"pagos_{desde or 'inicio'}_a_{hasta or 'hoy'}.csv"
            elif mes:
                filename = f"pagos_{mes}.csv"
            return self._render_historial_csv(qs, filename=filename)

        if export in ("xlsx", "excel"):
            filename = "historial_pagos.xlsx"
            titulo = "Historial de pagos"
            if dia:
                filename = f"pagos_{dia}.xlsx"
                titulo = f"Pagos del día {dia}"
            elif desde or hasta:
                filename = f"pagos_{desde or 'inicio'}_a_{hasta or 'hoy'}.xlsx"
                titulo = f"Pagos {desde or 'inicio'} a {hasta or 'hoy'}"
            elif mes:
                filename = f"pagos_{mes}.xlsx"
                titulo = f"Pagos del mes {mes}"
            return self._render_historial_xlsx(qs, filename=filename, titulo=titulo)

        if export == "pdf":
            filename = "historial_pagos.pdf"
            titulo = "Historial de pagos"
            if dia:
                filename = f"pagos_{dia}.pdf"
                titulo = f"Pagos del día {dia}"
            elif desde or hasta:
                filename = f"pagos_{desde or 'inicio'}_a_{hasta or 'hoy'}.pdf"
                titulo = f"Pagos {desde or 'inicio'} a {hasta or 'hoy'}"
            elif mes:
                filename = f"pagos_{mes}.pdf"
                titulo = f"Pagos del mes {mes}"
            return self._render_historial_pdf(qs, filename=filename, titulo=titulo)

        if all_flag:
            from .serializers import CuotaPagoHistorialSerializer
            items = list(qs[:MAX_HISTORIAL_ALL_ROWS])
            ser = CuotaPagoHistorialSerializer(items, many=True)
            return Response(
                {"count": len(items), "results": ser.data, "all": True, "max_rows": MAX_HISTORIAL_ALL_ROWS},
                status=status.HTTP_200_OK,
            )

        page = self.paginate_queryset(qs)
        if page is not None:
            ser = self.get_serializer(page, many=True)
            return self.get_paginated_response(ser.data)

        ser = self.get_serializer(qs, many=True)
        return Response(ser.data, status=status.HTTP_200_OK)