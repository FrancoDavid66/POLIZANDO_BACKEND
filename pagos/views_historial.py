# pagos/views_historial.py
#
# Mixin con el historial de pagos (tabla + exportación CSV/PDF/Excel) de
# CuotaViewSet, separado de pagos/views.py para que ese archivo no sea un
# solo bloque enorme. Se usa por herencia — mismo comportamiento, misma URL
# ("pagos", mapeada a historial_pagos), solo cambia en qué archivo vive el
# código. Es el bloque más grande que se movió (~500 líneas).

import csv
from datetime import timedelta
from io import StringIO, BytesIO

from django.db.models import F, OuterRef, Subquery, Q
from django.db.models.functions import Coalesce
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

from .models import Pago, Cuota
from pagos.views_helpers import (
    _get_seguridad_oficina_brute,
    _build_oficina_q_from_keys,
    _parse_mes_yyyy_mm,
    _parse_ymd,
    _to_bool,
    _compania_nombre_robusto,
    MAX_HISTORIAL_ALL_ROWS,
)


class HistorialPagosMixin:
    def _historial_pagos_base_queryset(self):
        last_pago_registrado_en_sq = Subquery(
            Pago.objects.filter(
                poliza_id=OuterRef("poliza_id"),
                cuota_nro=OuterRef("cuota_nro"),
            )
            .order_by("-registrado_en", "-id")
            .values("registrado_en")[:1]
        )

        return (
            Cuota.objects.filter(pagado=True)
            .select_related("poliza", "poliza__cliente")
            .annotate(
                pago_ts=Coalesce(F("pago_registrado_en"), last_pago_registrado_en_sq)
            )
            .only(
                "id",
                "fecha_pago",
                "pago_registrado_en",
                "cuota_nro",
                "monto",
                "forma_pago",
                "observaciones_pago",
                "ultima_observacion_pago",
                "poliza_id",
                "poliza__numero_poliza",
                "poliza__patente",
                "poliza__marca",
                "poliza__modelo",
                "poliza__oficina",
                "poliza__compania",
                "poliza__cantidad_cuotas",
                "poliza__cliente_id",
                "poliza__cliente__apellido",
                "poliza__cliente__nombre",
                "poliza__cliente__dni_cuit_cuil",
                "poliza__cliente__telefono",
            )
        )

    def _apply_historial_filters(self, qs, request, *, mes="", dia="", desde="", hasta="", oficina="", search=""):
        mes = (mes or "").strip()
        dia = (dia or "").strip()
        desde = (desde or "").strip()
        hasta = (hasta or "").strip()
        search = (search or "").strip()

        oficina_keys = _get_seguridad_oficina_brute(request, (oficina or "").strip())
        if "BLOQUEADO" in oficina_keys:
            raise ValueError("Acceso denegado a esta oficina.")

        if dia:
            d = _parse_ymd(dia)
            if not d:
                raise ValueError("Parámetro 'dia' inválido. Use YYYY-MM-DD (ej: 2026-02-10).")
            qs = qs.filter(fecha_pago=d)

        elif desde or hasta:
            d1 = _parse_ymd(desde) if desde else None
            d2 = _parse_ymd(hasta) if hasta else None
            if desde and not d1:
                raise ValueError("Parámetro 'desde' inválido. Use YYYY-MM-DD.")
            if hasta and not d2:
                raise ValueError("Parámetro 'hasta' inválido. Use YYYY-MM-DD.")
            if d1 and d2 and d2 < d1:
                raise ValueError("Rango inválido: 'hasta' no puede ser menor que 'desde'.")

            if d1:
                qs = qs.filter(fecha_pago__gte=d1)
            if d2:
                qs = qs.filter(fecha_pago__lt=(d2 + timedelta(days=1)))

        elif mes:
            first, nxt = _parse_mes_yyyy_mm(mes)
            if not first:
                raise ValueError("Parámetro 'mes' inválido. Use YYYY-MM (ej: 2026-02).")
            qs = qs.filter(fecha_pago__gte=first, fecha_pago__lt=nxt)

        if oficina_keys:
            qs = qs.filter(_build_oficina_q_from_keys(oficina_keys))

        if search:
            terminos = search.split()
            for t in terminos:
                qs = qs.filter(
                    Q(poliza__numero_poliza__icontains=t)
                    | Q(poliza__patente__icontains=t)
                    | Q(poliza__cliente__apellido__icontains=t)
                    | Q(poliza__cliente__nombre__icontains=t)
                    | Q(poliza__cliente__dni_cuit_cuil__icontains=t)
                )

        return qs

    def _apply_historial_ordering(self, qs, ordering_raw: str):
        ordering = (ordering_raw or "-fecha_pago").strip()
        allowed = {"fecha_pago", "-fecha_pago", "monto", "-monto", "cuota_nro", "-cuota_nro"}
        if ordering not in allowed:
            ordering = "-fecha_pago"

        if ordering == "monto":
            return qs.order_by("monto", "poliza_id", "cuota_nro")
        if ordering == "-monto":
            return qs.order_by("-monto", "-fecha_pago", "poliza_id", "cuota_nro")
        if ordering == "cuota_nro":
            return qs.order_by("cuota_nro", "-fecha_pago", "poliza_id")
        if ordering == "-cuota_nro":
            return qs.order_by("-cuota_nro", "-fecha_pago", "poliza_id")

        return qs.order_by("-fecha_pago", "poliza_id", "cuota_nro")

    def _render_historial_csv(self, qs, filename="historial_pagos.csv"):
        out = StringIO()
        writer = csv.writer(out)

        writer.writerow(
            [
                "fecha_pago",
                "cuota_nro",
                "monto",
                "forma_pago",
                "numero_poliza",
                "patente",
                "compania",
                "oficina",
                "cliente_apellido",
                "cliente_nombre",
                "cliente_dni",
                "cliente_telefono",
            ]
        )

        for c in qs.iterator(chunk_size=2000):
            pol = getattr(c, "poliza", None)
            cli = getattr(pol, "cliente", None) if pol else None

            writer.writerow(
                [
                    getattr(c, "fecha_pago", "") or "",
                    getattr(c, "cuota_nro", "") or "",
                    getattr(c, "monto", "") or "",
                    getattr(c, "forma_pago", "") or "",
                    getattr(pol, "numero_poliza", "") if pol else "",
                    getattr(pol, "patente", "") if pol else "",
                    _compania_nombre_robusto(pol),
                    getattr(pol, "oficina", "") if pol else "",
                    getattr(cli, "apellido", "") if cli else "",
                    getattr(cli, "nombre", "") if cli else "",
                    getattr(cli, "dni_cuit_cuil", "") if cli else "",
                    getattr(cli, "telefono", "") if cli else "",
                ]
            )

        resp = HttpResponse(out.getvalue(), content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    def _render_historial_pdf(self, qs, filename="historial_pagos.pdf", titulo="Historial de pagos"):
        buff = BytesIO()
        pdf = canvas.Canvas(buff, pagesize=landscape(A4))
        width, height = landscape(A4)

        left = 12 * mm
        top = height - 12 * mm
        line_h = 6.2 * mm

        pdf.setTitle(filename)

        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(left, top, titulo)
        pdf.setFont("Helvetica", 9)
        pdf.drawString(left, top - 8 * mm, f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}")

        y = top - 16 * mm

        cols = [
            ("Fecha", 22 * mm),
            ("Asegurado", 60 * mm),
            ("DNI", 26 * mm),
            ("Patente", 24 * mm),
            ("Póliza", 26 * mm),
            ("Compañía", 40 * mm),
            ("Oficina", 22 * mm),
            ("Importe", 24 * mm),
            ("Medio", 26 * mm),
        ]

        def draw_row(values, y, bold=False):
            x = left
            pdf.setFont("Helvetica-Bold" if bold else "Helvetica", 9)
            for (txt, w), v in zip(cols, values):
                s = str(v or "")
                if len(s) > 38 and w <= 60 * mm:
                    s = s[:35] + "…"
                pdf.drawString(x, y, s)
                x += w

        draw_row([t for (t, _) in cols], y, bold=True)
        y -= line_h
        pdf.setLineWidth(0.4)
        pdf.line(left, y + 2 * mm, width - left, y + 2 * mm)

        for item in qs.iterator(chunk_size=2000):
            if y < 12 * mm:
                pdf.showPage()
                pdf.setFont("Helvetica-Bold", 14)
                pdf.drawString(left, top, titulo)
                pdf.setFont("Helvetica", 9)
                pdf.drawString(left, top - 8 * mm, f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}")
                y = top - 16 * mm
                draw_row([t for (t, _) in cols], y, bold=True)
                y -= line_h
                pdf.line(left, y + 2 * mm, width - left, y + 2 * mm)

            pol = getattr(item, "poliza", None)
            cli = getattr(pol, "cliente", None) if pol else None

            fecha = getattr(item, "fecha_pago", None) or ""
            if fecha:
                try:
                    fecha = fecha.strftime("%d/%m/%Y")
                except Exception:
                    fecha = str(fecha)

            ape = getattr(cli, "apellido", "") if cli else ""
            nom = getattr(cli, "nombre", "") if cli else ""
            asegurado = f"{ape}, {nom}".strip(", ").strip()

            dni = getattr(cli, "dni_cuit_cuil", "") if cli else ""
            patente = getattr(pol, "patente", "") if pol else ""
            numero_poliza = getattr(pol, "numero_poliza", "") if pol else ""
            compania = _compania_nombre_robusto(pol)
            oficina = getattr(pol, "oficina", "") if pol else ""
            monto = getattr(item, "monto", "") or ""
            medio = getattr(item, "forma_pago", "") or ""

            draw_row([fecha, asegurado, dni, patente, numero_poliza, compania, oficina, monto, medio], y, bold=False)
            y -= line_h

        pdf.save()
        buff.seek(0)
        return FileResponse(buff, as_attachment=True, filename=filename, content_type="application/pdf")

    def _render_historial_xlsx(self, qs, filename="historial_pagos.xlsx", titulo="Historial de pagos"):
        """
        Excel profesional con TABLA REAL de openpyxl (objeto Table, con flechitas
        de filtro/ordenamiento nativas en cada header).

        Columnas (8): Apellido y Nombre · Patente · Compañía · Fecha de pago ·
                      Cuota · Oficina · Medio · Importe
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        wb = Workbook()
        ws = wb.active
        ws.title = "Historial de Pagos"

        # ── Helpers locales ─────────────────────────────────────────────────
        def _cuota_label(c):
            """Devuelve '1/3', '3/5', etc. Si no hay total, devuelve solo el nro."""
            nro = getattr(c, "cuota_nro", None)
            if not nro:
                return ""
            pol = getattr(c, "poliza", None)
            total = getattr(pol, "cantidad_cuotas", None) if pol else None
            if not total:
                return str(nro)
            return f"{nro}/{total}"

        def _apellido_nombre(cli):
            if not cli:
                return ""
            ape = (getattr(cli, "apellido", "") or "").strip()
            nom = (getattr(cli, "nombre", "") or "").strip()
            if ape and nom:
                return f"{ape}, {nom}"
            return ape or nom or ""

        def _fecha_pago_str(c):
            ts = getattr(c, "pago_ts", None) or getattr(c, "pago_registrado_en", None)
            if ts:
                try:
                    return timezone.localtime(ts).strftime("%d/%m/%Y %H:%M")
                except Exception:
                    pass
            fp = getattr(c, "fecha_pago", None)
            if fp:
                try:
                    return fp.strftime("%d/%m/%Y")
                except Exception:
                    return str(fp)
            return ""

        def _medio(c):
            f = (getattr(c, "forma_pago", "") or "").upper().strip()
            if f == "EFECTIVO":
                return "Efectivo"
            if f == "TRANSFERENCIA":
                return "Transferencia"
            return f or "—"

        def _oficina_nombre(pol):
            if not pol:
                return ""
            ofi = getattr(pol, "oficina", None)
            if not ofi:
                return ""
            # Si es FK a usuarios.Oficina, tiene .nombre
            nombre = getattr(ofi, "nombre", None)
            if nombre:
                return str(nombre).strip()
            return str(ofi).strip()

        # ── Estilos ─────────────────────────────────────────────────────────
        title_font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
        title_fill = PatternFill("solid", fgColor="1E3A8A")
        title_align = Alignment(horizontal="left", vertical="center", indent=1)

        subtitle_font = Font(color="475569", size=10, italic=True, name="Calibri")
        subtitle_align = Alignment(horizontal="left", vertical="center", indent=1)

        header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        header_fill = PatternFill("solid", fgColor="1E3A8A")
        header_align = Alignment(horizontal="center", vertical="center")

        thin = Side(border_style="thin", color="E2E8F0")
        cell_border = Border(top=thin, bottom=thin, left=thin, right=thin)
        cell_font = Font(color="1E293B", size=10, name="Calibri")
        cell_align_left = Alignment(horizontal="left", vertical="center", indent=1)
        cell_align_right = Alignment(horizontal="right", vertical="center", indent=1)
        cell_align_center = Alignment(horizontal="center", vertical="center")

        total_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        total_fill = PatternFill("solid", fgColor="064E3B")
        total_align = Alignment(horizontal="right", vertical="center", indent=1)

        # ── Estructura ──────────────────────────────────────────────────────
        HEADERS = [
            "Apellido y Nombre",
            "Patente",
            "Compañía",
            "Fecha de pago",
            "Cuota",
            "Oficina",
            "Medio",
            "Importe",
        ]
        n_cols = len(HEADERS)
        last_col_letter = get_column_letter(n_cols)

        # Fila 1: Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        c = ws.cell(row=1, column=1, value=titulo)
        c.font = title_font
        c.fill = title_fill
        c.alignment = title_align
        ws.row_dimensions[1].height = 28

        # Fila 2: Subtítulo
        total_count = qs.count() if hasattr(qs, "count") else 0
        generado_str = timezone.localtime().strftime("%d/%m/%Y %H:%M")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        c = ws.cell(row=2, column=1, value=f"Generado: {generado_str}   ·   Total de pagos: {total_count}")
        c.font = subtitle_font
        c.alignment = subtitle_align
        ws.row_dimensions[2].height = 18

        # Fila 3 vacía
        ws.row_dimensions[3].height = 6

        # Fila 4: Headers
        header_row = 4
        for col_idx, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = cell_border
        ws.row_dimensions[header_row].height = 24

        # Filas de datos
        row_alt_fill = PatternFill("solid", fgColor="F8FAFC")
        total_general = 0.0
        current_row = header_row

        for c_obj in qs.iterator(chunk_size=2000):
            current_row += 1
            pol = getattr(c_obj, "poliza", None)
            cli = getattr(pol, "cliente", None) if pol else None

            apellido_nombre = _apellido_nombre(cli)
            patente = (getattr(pol, "patente", "") or "").strip().upper() if pol else ""
            compania = _compania_nombre_robusto(pol) if pol else ""
            fecha_pago = _fecha_pago_str(c_obj)
            cuota_lbl = _cuota_label(c_obj)
            oficina = _oficina_nombre(pol)
            medio = _medio(c_obj)
            try:
                monto = float(getattr(c_obj, "monto", 0) or 0)
            except (TypeError, ValueError):
                monto = 0.0
            total_general += monto

            valores = [
                apellido_nombre or "—",
                patente or "—",
                compania or "—",
                fecha_pago or "—",
                cuota_lbl or "—",
                oficina or "—",
                medio,
                monto,
            ]

            is_alt = (current_row - header_row) % 2 == 0
            for col_idx, val in enumerate(valores, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = cell_font
                cell.border = cell_border
                if is_alt:
                    cell.fill = row_alt_fill

                if col_idx == 8:
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = cell_align_right
                elif col_idx in (2, 5, 7):
                    cell.alignment = cell_align_center
                else:
                    cell.alignment = cell_align_left

        data_last_row = current_row
        has_data = data_last_row > header_row

        # Total general
        if has_data:
            total_row = data_last_row + 1
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
            c = ws.cell(row=total_row, column=1, value="TOTAL GENERAL")
            c.font = total_font
            c.fill = total_fill
            c.alignment = total_align
            c = ws.cell(row=total_row, column=8, value=total_general)
            c.font = total_font
            c.fill = total_fill
            c.alignment = total_align
            c.number_format = '"$"#,##0.00'
            ws.row_dimensions[total_row].height = 22

        # 🚀 TABLA REAL de Excel (con flechitas de filtro nativas)
        if has_data:
            table_ref = f"A{header_row}:{last_col_letter}{data_last_row}"
            table = Table(displayName="HistorialPagos", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

        # Anchos de columna
        anchos = [32, 14, 26, 20, 10, 18, 16, 16]
        for i, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze panes (congela título + header)
        ws.freeze_panes = "A5"

        # Respuesta HTTP
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=["get"], url_path="pagos")
    def historial_pagos(self, request):
        mes = (request.query_params.get("mes") or "").strip()
        dia = (request.query_params.get("dia") or "").strip()
        desde = (request.query_params.get("desde") or "").strip()
        hasta = (request.query_params.get("hasta") or "").strip()
        oficina = (request.query_params.get("oficina") or "").strip()
        search = (request.query_params.get("search") or request.query_params.get("q") or "").strip()
        ordering = (request.query_params.get("ordering") or "-fecha_pago").strip()

        export = (request.query_params.get("export") or "").strip().lower()
        all_flag = _to_bool(request.query_params.get("all") or request.query_params.get("todos"))

        qs = self._historial_pagos_base_queryset()

        try:
            qs = self._apply_historial_filters(qs, request, mes=mes, dia=dia, desde=desde, hasta=hasta, oficina=oficina, search=search)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        qs = self._apply_historial_ordering(qs, ordering)

        if export == "csv":
            filename = "historial_pagos.csv"
            if dia:
                filename = f"pagos_{dia}.csv"
            elif desde or hasta:
                filename = f"pagos_{desde or 'inicio'}_a_{hasta or 'hoy'}.csv"
            elif mes:
                filename = f"pagos_{mes}.csv"
            return self._render_historial_csv(qs, filename=filename)

        if export in ("xlsx", "excel"):
            filename = "historial_pagos.xlsx"
            titulo = "Historial de pagos"
            if dia:
                filename = f"pagos_{dia}.xlsx"
                titulo = f"Pagos del día {dia}"
            elif desde or hasta:
                filename = f"pagos_{desde or 'inicio'}_a_{hasta or 'hoy'}.xlsx"
                titulo = f"Pagos {desde or 'inicio'} a {hasta or 'hoy'}"
            elif mes:
                filename = f"pagos_{mes}.xlsx"
                titulo = f"Pagos del mes {mes}"
            return self._render_historial_xlsx(qs, filename=filename, titulo=titulo)

        if export == "pdf":
            filename = "historial_pagos.pdf"
            titulo = "Historial de pagos"
            if dia:
                filename = f"pagos_{dia}.pdf"
                titulo = f"Pagos del día {dia}"
            elif desde or hasta:
                filename = f"pagos_{desde or 'inicio'}_a_{hasta or 'hoy'}.pdf"
                titulo = f"Pagos {desde or 'inicio'} a {hasta or 'hoy'}"
            elif mes:
                filename = f"pagos_{mes}.pdf"
                titulo = f"Pagos del mes {mes}"
            return self._render_historial_pdf(qs, filename=filename, titulo=titulo)

        if all_flag:
            from .serializers import CuotaPagoHistorialSerializer
            items = list(qs[:MAX_HISTORIAL_ALL_ROWS])
            ser = CuotaPagoHistorialSerializer(items, many=True)
            return Response(
                {"count": len(items), "results": ser.data, "all": True, "max_rows": MAX_HISTORIAL_ALL_ROWS},
                status=status.HTTP_200_OK,
            )

        page = self.paginate_queryset(qs)
        if page is not None:
            ser = self.get_serializer(page, many=True)
            return self.get_paginated_response(ser.data)

        ser = self.get_serializer(qs, many=True)
        return Response(ser.data, status=status.HTTP_200_OK)