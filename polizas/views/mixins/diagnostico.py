# polizas/views/mixins/diagnostico.py
"""
🔍 MIXIN DE DIAGNÓSTICO DE DATOS

Provee 2 endpoints SOLO LECTURA para detectar problemas de calidad de datos:

  GET /api/polizas/diagnostico-datos/
      → JSON con conteos + ejemplos de cada categoría

  GET /api/polizas/diagnostico-datos/exportar/?categoria=CATEGORIA&formato=csv|xlsx
      → Descarga CSV o Excel con los IDs detectados

⚠️  NINGÚN ENDPOINT MODIFICA DATOS. Solo lee y reporta.

🛡️ Acceso restringido: solo usuarios con rol=ADMIN o is_superuser.
"""

from datetime import timedelta
import csv
from io import BytesIO, StringIO

from django.db.models import Exists, OuterRef, Count, Max, Min, Q
from django.http import HttpResponse
from django.utils import timezone

from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from clientes.models import Cliente
from polizas.models import Poliza
from pagos.models import Cuota


# ════════════════════════════════════════════════════
# Helpers
# ════════════════════════════════════════════════════

def _is_admin(user):
    """¿El usuario tiene permiso para ver el diagnóstico?"""
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    perfil = getattr(user, "perfil", None)
    if perfil and getattr(perfil, "rol", "") == "ADMIN":
        return True
    return False


def _cliente_label(c):
    """Formato corto de un cliente para mostrar en ejemplos."""
    if not c:
        return "—"
    nombre = (c.nombre or "").strip()
    apellido = (c.apellido or "").strip()
    full = ", ".join(filter(None, [apellido, nombre]))
    return full or "—"


def _serialize_cliente_basic(c):
    """Datos mínimos de un cliente huérfano para el frontend."""
    return {
        "id": c.id,
        "nombre": c.nombre or "",
        "apellido": c.apellido or "",
        "dni_cuit_cuil": c.dni_cuit_cuil or "",
        "telefono": c.telefono or "",
    }


def _serialize_poliza_basic(p):
    """Datos mínimos de una póliza para el frontend."""
    return {
        "id": p.id,
        "numero_poliza": p.numero_poliza or "",
        "patente": (p.patente or "").upper(),
        "estado": p.estado or "",
        "cliente_id": p.cliente_id,
        "cliente_label": _cliente_label(p.cliente if p.cliente_id else None),
        "compania": p.compania or "",
        "oficina_id": p.oficina_id,
    }


# ════════════════════════════════════════════════════
# Categorías diagnosticables
# ════════════════════════════════════════════════════
# Cada categoría es una clave: el frontend la usa para pedir
# ejemplos detallados o exportar.

CATEGORIAS = {
    "clientes_sin_oficina": "Clientes sin oficina asignada",
    "polizas_sin_oficina": "Pólizas sin oficina asignada",
    "polizas_activas_morosas": "Pólizas 'activas' con mora",
    "polizas_vencidas_al_dia": "Pólizas 'vencidas' sin cuotas atrasadas",
    "polizas_canceladas_con_pago": "Pólizas 'canceladas' con pagos recientes",
}


def _get_queryset_categoria(categoria, dias_mora=30):
    """
    Devuelve el queryset correspondiente a una categoría.

    Esta función es la ÚNICA fuente de verdad de qué cuenta cada categoría.
    Se usa tanto para el reporte como para la exportación.
    """
    hoy = timezone.localdate()
    cutoff_mora = hoy - timedelta(days=dias_mora)
    hace_30 = hoy - timedelta(days=30)

    if categoria == "clientes_sin_oficina":
        return Cliente.objects.filter(oficina__isnull=True).order_by("-id")

    if categoria == "polizas_sin_oficina":
        return (
            Poliza.objects.filter(oficina__isnull=True)
            .select_related("cliente")
            .order_by("-id")
        )

    if categoria == "polizas_activas_morosas":
        # Morosa = la cobertura (vto de la última cuota PAGADA) ya venció y quedan impagas.
        # Si nunca pagó nada, la mora arranca en el vto de su primera cuota impaga.
        return (
            Poliza.objects.filter(estado__iexact="activa")
            .annotate(
                _cobertura=Max("cuotas__fecha_vencimiento", filter=Q(cuotas__pagado=True)),
                _impagas=Count("cuotas", filter=Q(cuotas__pagado=False)),
                _primer_impaga=Min("cuotas__fecha_vencimiento", filter=Q(cuotas__pagado=False)),
            )
            .filter(_impagas__gt=0)
            .filter(Q(_cobertura__lt=cutoff_mora) | Q(_cobertura__isnull=True, _primer_impaga__lt=cutoff_mora))
            .select_related("cliente")
            .order_by("-id")
        )

    if categoria == "polizas_vencidas_al_dia":
        # Pólizas marcadas "vencida" pero que en realidad están al día:
        # no tienen impagas, o su cobertura llega hasta hoy o más adelante.
        return (
            Poliza.objects.filter(estado__iexact="vencida")
            .annotate(
                _cobertura=Max("cuotas__fecha_vencimiento", filter=Q(cuotas__pagado=True)),
                _impagas=Count("cuotas", filter=Q(cuotas__pagado=False)),
            )
            .filter(Q(_impagas=0) | Q(_cobertura__gte=hoy))
            .select_related("cliente")
            .order_by("-id")
        )

    if categoria == "polizas_canceladas_con_pago":
        cuotas_pagadas_recientes = Cuota.objects.filter(
            poliza=OuterRef("pk"),
            pagado=True,
            fecha_pago__gte=hace_30,
        )
        return (
            Poliza.objects.filter(estado__iexact="cancelada")
            .annotate(tiene_pago_reciente=Exists(cuotas_pagadas_recientes))
            .filter(tiene_pago_reciente=True)
            .select_related("cliente")
            .order_by("-id")
        )

    return None


# ════════════════════════════════════════════════════
# Mixin
# ════════════════════════════════════════════════════

class PolizaDiagnosticoMixin:
    """
    Endpoints de diagnóstico de calidad de datos.
    Conectados al PolizaViewSet, accesibles solo para admin.
    """

    @action(
        detail=False,
        methods=["get"],
        url_path="diagnostico-datos",
        permission_classes=[IsAuthenticated],
    )
    def diagnostico_datos(self, request):
        """
        GET /api/polizas/diagnostico-datos/?dias_mora=30&ejemplos=10

        Devuelve conteos + primeros N ejemplos de cada categoría.
        """
        if not _is_admin(request.user):
            return Response(
                {"error": "Solo administradores pueden acceder al diagnóstico."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # ── Parámetros ──
        try:
            dias_mora = int(request.query_params.get("dias_mora", 30))
        except (TypeError, ValueError):
            dias_mora = 30
        dias_mora = max(1, min(365, dias_mora))

        try:
            ejemplos = int(request.query_params.get("ejemplos", 10))
        except (TypeError, ValueError):
            ejemplos = 10
        ejemplos = max(0, min(50, ejemplos))

        hoy = timezone.localdate()

        # ── Conteos por categoría ──
        resultado = {}

        # 1. Clientes sin oficina
        qs = _get_queryset_categoria("clientes_sin_oficina", dias_mora)
        resultado["clientes_sin_oficina"] = {
            "label": CATEGORIAS["clientes_sin_oficina"],
            "count": qs.count(),
            "ejemplos": [_serialize_cliente_basic(c) for c in qs[:ejemplos]],
        }

        # 2. Pólizas sin oficina
        qs = _get_queryset_categoria("polizas_sin_oficina", dias_mora)
        resultado["polizas_sin_oficina"] = {
            "label": CATEGORIAS["polizas_sin_oficina"],
            "count": qs.count(),
            "ejemplos": [_serialize_poliza_basic(p) for p in qs[:ejemplos]],
        }

        # 3. Pólizas activas con mora
        qs = _get_queryset_categoria("polizas_activas_morosas", dias_mora)
        resultado["polizas_activas_morosas"] = {
            "label": f"Pólizas 'activas' con mora de +{dias_mora} días",
            "count": qs.count(),
            "ejemplos": [_serialize_poliza_basic(p) for p in qs[:ejemplos]],
        }

        # 4. Pólizas vencidas al día
        qs = _get_queryset_categoria("polizas_vencidas_al_dia", dias_mora)
        resultado["polizas_vencidas_al_dia"] = {
            "label": CATEGORIAS["polizas_vencidas_al_dia"],
            "count": qs.count(),
            "ejemplos": [_serialize_poliza_basic(p) for p in qs[:ejemplos]],
        }

        # 5. Pólizas canceladas con pago reciente
        qs = _get_queryset_categoria("polizas_canceladas_con_pago", dias_mora)
        resultado["polizas_canceladas_con_pago"] = {
            "label": CATEGORIAS["polizas_canceladas_con_pago"],
            "count": qs.count(),
            "ejemplos": [_serialize_poliza_basic(p) for p in qs[:ejemplos]],
        }

        # ── Distribución de estados (informativa) ──
        por_estado = list(
            Poliza.objects.values("estado")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
        total_polizas = Poliza.objects.count()
        total_clientes = Cliente.objects.count()

        # ── Totales globales ──
        total_problemas = sum(cat["count"] for cat in resultado.values())

        return Response(
            {
                "ok": True,
                "ejecutado_en": timezone.now().isoformat(),
                "fecha_base": hoy.isoformat(),
                "dias_mora": dias_mora,
                "totales": {
                    "polizas_en_base": total_polizas,
                    "clientes_en_base": total_clientes,
                    "problemas_detectados": total_problemas,
                },
                "categorias": resultado,
                "distribucion_estados": [
                    {
                        "estado": (row["estado"] or "SIN_ESTADO").lower(),
                        "total": row["total"],
                    }
                    for row in por_estado
                ],
            },
            status=status.HTTP_200_OK,
        )

    @action(
        detail=False,
        methods=["get"],
        url_path="diagnostico-datos/exportar",
        permission_classes=[IsAuthenticated],
    )
    def diagnostico_datos_exportar(self, request):
        """
        GET /api/polizas/diagnostico-datos/exportar/?categoria=...&formato=csv|xlsx&dias_mora=30

        Descarga CSV o Excel con TODOS los IDs detectados de una categoría.
        """
        if not _is_admin(request.user):
            return Response(
                {"error": "Solo administradores pueden exportar."},
                status=status.HTTP_403_FORBIDDEN,
            )

        categoria = (request.query_params.get("categoria") or "").strip()
        if categoria not in CATEGORIAS:
            return Response(
                {
                    "error": "Categoría inválida.",
                    "categorias_validas": list(CATEGORIAS.keys()),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        formato = (request.query_params.get("formato") or "csv").lower().strip()
        if formato not in ("csv", "xlsx"):
            return Response(
                {"error": "Formato inválido. Use csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            dias_mora = int(request.query_params.get("dias_mora", 30))
        except (TypeError, ValueError):
            dias_mora = 30
        dias_mora = max(1, min(365, dias_mora))

        qs = _get_queryset_categoria(categoria, dias_mora)

        # ── Armar filas ──
        if categoria == "clientes_sin_oficina":
            headers = ["id", "apellido", "nombre", "dni_cuit_cuil", "telefono"]
            rows = [
                [c.id, c.apellido or "", c.nombre or "", c.dni_cuit_cuil or "", c.telefono or ""]
                for c in qs
            ]
        else:
            # Todas las demás categorías son de pólizas
            headers = [
                "id",
                "numero_poliza",
                "patente",
                "estado",
                "cliente_id",
                "cliente",
                "compania",
                "oficina_id",
            ]
            rows = [
                [
                    p.id,
                    p.numero_poliza or "",
                    (p.patente or "").upper(),
                    p.estado or "",
                    p.cliente_id or "",
                    _cliente_label(p.cliente if p.cliente_id else None),
                    p.compania or "",
                    p.oficina_id or "",
                ]
                for p in qs
            ]

        # ── Nombre del archivo ──
        fecha = timezone.localdate().isoformat()
        nombre_base = f"diagnostico_{categoria}_{fecha}"

        # ── Exportar CSV ──
        if formato == "csv":
            buf = StringIO()
            writer = csv.writer(buf)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)

            response = HttpResponse(
                buf.getvalue().encode("utf-8-sig"),  # BOM para Excel
                content_type="text/csv; charset=utf-8",
            )
            response["Content-Disposition"] = f'attachment; filename="{nombre_base}.csv"'
            return response

        # ── Exportar XLSX ──
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return Response(
                {"error": "openpyxl no está instalado. Use formato=csv."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Diagnóstico"

        # Encabezados con estilo
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="0F172A")
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Filas de datos
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Ajustar ancho de columnas
        for col_idx, header in enumerate(headers, start=1):
            max_len = max(
                [len(str(header))]
                + [len(str(row[col_idx - 1])) for row in rows[:100]],
                default=12,
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(
                max_len + 2, 40
            )

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{nombre_base}.xlsx"'
        return response